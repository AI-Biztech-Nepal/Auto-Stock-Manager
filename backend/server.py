from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from pathlib import Path
import os, logging, jwt, uuid, json, base64, io, asyncio, re, copy, contextvars, secrets, hashlib
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures.process import BrokenProcessPool
from pydantic import BaseModel, Field, EmailStr
from PIL import Image, ImageOps
import pillow_heif
pillow_heif.register_heif_opener()  # lets Image.open() decode HEIC/HEIF from iPhone cameras
from pypdf import PdfReader, PdfWriter
from google import genai
from google.genai import types as genai_types
from google.genai import errors as genai_errors
import httpx
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import certifi
from pymongo.server_api import ServerApi
from pymongo import ReturnDocument

# Force Python SSL to use certifi CA bundle (fixes Atlas TLS on Docker/Render)
os.environ["SSL_CERT_FILE"] = certifi.where()
os.environ["REQUESTS_CA_BUNDLE"] = certifi.where()

# DB_BACKEND=mysql points the app at Hostinger MySQL (backend/sqldb.py, a thin
# shim over the same db.<collection>.<method>(...) calls below) instead of
# MongoDB Atlas. Defaults to mongo so this switch is safe to deploy before cutover.
DB_BACKEND = os.environ.get('DB_BACKEND', 'mongo').lower()
if DB_BACKEND == 'mysql':
    from sqldb import MySQLDatabase
    db = MySQLDatabase()
    client = db  # shutdown() below calls client.close(); same object serves both roles
else:
    mongo_url = os.environ['MONGO_URL']
    _is_atlas = "mongodb+srv" in mongo_url or "mongodb.net" in mongo_url
    client = AsyncIOMotorClient(
        mongo_url,
        tls=True,
        tlsCAFile=certifi.where(),
        server_api=ServerApi('1')
    ) if _is_atlas else AsyncIOMotorClient(mongo_url)
    db = client[os.environ['DB_NAME']]

# ── Multi-tenant data isolation ─────────────────────────────────────────
# Every business that signs up (see /auth/signup) gets its own company_id. The previous
# attempt at this scoped each endpoint by hand (merging company_id into every filter/insert
# across ~100 call sites) -- one missed call site was exactly how company_id corruption and
# cross-tenant data leaks happened before. Instead, scoping is enforced once, here, at the DB
# access layer: get_current_user sets current_company_id from the caller's JWT for the
# duration of the request, and _ScopedCollection transparently merges/stamps company_id onto
# every query and write for tenant collections -- an endpoint would have to go out of its way
# to bypass this, rather than having to remember to apply it.
current_company_id: contextvars.ContextVar = contextvars.ContextVar("current_company_id", default=None)

# Every collection that holds one business's own data. Deliberately excludes:
#  - "users": login accounts need both company-scoped queries (list/edit/delete my company's
#    accounts) AND global ones (login-by-username, uniqueness check) -- see the handful of
#    /auth/* routes below, which apply company_scope() by hand instead.
#  - "companies": the tenant registry itself, never scoped to a tenant.
TENANT_COLLECTIONS = {
    "customers", "emi_payments", "emi_records", "expenses", "job_cards", "kit_components",
    "leads", "legal_documents", "part_transactions", "partners", "sales", "spare_parts",
    "sync_logs", "team_members", "vehicle_photos", "vehicles", "vendor_payments", "vendors",
    "audit_logs", "ai_chat_sessions", "settings",
}

def company_scope(cu: dict) -> dict:
    """For the handful of "users" queries that need explicit company scoping (that
    collection is deliberately excluded from the automatic wrapper -- see above)."""
    return {"company_id": cu.get("company_id")}

class _ScopedCollection:
    def __init__(self, coll):
        self._coll = coll

    def _scoped(self, filt):
        filt = dict(filt or {})
        cid = current_company_id.get()
        if cid is not None:
            filt["company_id"] = cid
        return filt

    def _stamped(self, doc):
        cid = current_company_id.get()
        if "company_id" not in doc:
            if cid is None:
                # A session with no company_id (e.g. a token minted before this account
                # had one assigned) must never silently write untagged data -- that data
                # becomes invisible to every company-scoped read afterward with no error
                # raised anywhere, which is exactly how 3 real sales went missing on
                # 2026-08-18. Fail loudly instead so the user knows to log out/in again.
                raise HTTPException(401, "Your session is missing company info -- please log out and log back in, then retry.")
            doc = {**doc, "company_id": cid}
        return doc

    async def find_one(self, filt=None, *a, **k):
        return await self._coll.find_one(self._scoped(filt), *a, **k)

    def find(self, filt=None, *a, **k):
        return self._coll.find(self._scoped(filt), *a, **k)

    async def count_documents(self, filt=None, *a, **k):
        return await self._coll.count_documents(self._scoped(filt), *a, **k)

    def distinct(self, field, filt=None, *a, **k):
        return self._coll.distinct(field, self._scoped(filt), *a, **k)

    async def insert_one(self, doc, *a, **k):
        return await self._coll.insert_one(self._stamped(doc), *a, **k)

    async def insert_many(self, docs, *a, **k):
        return await self._coll.insert_many([self._stamped(d) for d in docs], *a, **k)

    async def update_one(self, filt, update, *a, **k):
        return await self._coll.update_one(self._scoped(filt), update, *a, **k)

    async def update_many(self, filt, update, *a, **k):
        return await self._coll.update_many(self._scoped(filt), update, *a, **k)

    async def delete_one(self, filt, *a, **k):
        return await self._coll.delete_one(self._scoped(filt), *a, **k)

    async def delete_many(self, filt, *a, **k):
        return await self._coll.delete_many(self._scoped(filt), *a, **k)

    async def find_one_and_update(self, filt, update, *a, **k):
        return await self._coll.find_one_and_update(self._scoped(filt), update, *a, **k)

    def create_index(self, *a, **k):
        return self._coll.create_index(*a, **k)

class _ScopedDB:
    """Drop-in wrapper: db.vehicles.find(...) etc. all still work unchanged. Only
    collections in TENANT_COLLECTIONS get company_id auto-merged/stamped."""
    def __init__(self, real_db):
        self._real_db = real_db

    def __getattr__(self, name):
        coll = getattr(self._real_db, name)
        return _ScopedCollection(coll) if name in TENANT_COLLECTIONS else coll

db = _ScopedDB(db)  # db.users / db.companies pass through unscoped automatically (not in TENANT_COLLECTIONS)

# Fails startup rather than silently signing tokens with a known, public default --
# 'hamro-gng-2024' was the code fallback for years and is effectively a published secret
# (it's right here in git history), so any deployment still relying on it can have its
# tokens forged by anyone who's read this file. Set a real random JWT_SECRET in the
# environment before deploying; see DEPLOYMENT.md.
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET or JWT_SECRET == 'hamro-gng-2024':
    raise RuntimeError(
        "JWT_SECRET is not set (or is still the old insecure default 'hamro-gng-2024'). "
        "Set it to a long random value in the environment before starting the app -- "
        "e.g. `python -c \"import secrets; print(secrets.token_hex(32))\"` -- "
        "see DEPLOYMENT.md's Security Checklist."
    )

# ── AI Assistant (Google Gemini) ────────────────────────────────────────
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
AI_MODEL = "gemini-3.5-flash"
ai_client = genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None

app = FastAPI(title="Hamro G&G Auto OS", version="1.0.0")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto", bcrypt__rounds=10)

# ── Rate limiting (brute-force login / signup-spam guard) ──────────────
# Keyed by client IP, in-memory -- fine for this app's single-process deployment (see
# ecosystem.config.js: one PM2/uvicorn process, no horizontal scaling). Reads
# X-Forwarded-For first, same as _public_photo_url below, since the VPS puts nginx in
# front of this app; falls back to the direct connection IP for local/dev use.
def _client_ip(request: Request) -> str:
    xff = request.headers.get("x-forwarded-for")
    return xff.split(",")[0].strip() if xff else get_remote_address(request)

limiter = Limiter(key_func=_client_ip)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Storefront revalidation ─────────────────────────────────────────────
# The Next.js storefront (hamroauto.com.np) caches its public pages for 60s (ISR) and
# otherwise has no way to know inventory changed. This pings its on-demand revalidation
# route right after a change so listings/photos update within seconds instead of up to
# a minute later. Both env vars are optional — with either unset this silently no-ops,
# so it's safe to deploy before the storefront side is configured.
STOREFRONT_REVALIDATE_URL = os.environ.get("STOREFRONT_REVALIDATE_URL")
STOREFRONT_REVALIDATE_SECRET = os.environ.get("STOREFRONT_REVALIDATE_SECRET")

async def _notify_storefront():
    if not STOREFRONT_REVALIDATE_URL:
        return
    try:
        async with httpx.AsyncClient(timeout=5) as http:
            await http.post(STOREFRONT_REVALIDATE_URL, params={"secret": STOREFRONT_REVALIDATE_SECRET})
    except Exception:
        logger.warning("Storefront revalidation ping failed", exc_info=True)

# ── Transactional email (Resend) ─────────────────────────────────────────
# Powers email verification and password reset below. Optional like the storefront ping
# above: with RESEND_API_KEY unset, sending silently no-ops (just logs) instead of
# raising, so this is safe to deploy before a Resend account/domain is set up -- signup
# and login just won't be able to email anyone until it's configured. See DEPLOYMENT.md.
RESEND_API_KEY = os.environ.get("RESEND_API_KEY")
RESEND_FROM = os.environ.get("RESEND_FROM", "Hamro G&G Auto OS <onboarding@resend.dev>")
# Used to build the links inside verification/reset emails (they open in the browser,
# not the API) -- must be the frontend's real URL in production. Defaults to local dev.
FRONTEND_URL = os.environ.get("FRONTEND_URL", "http://localhost:3000")

async def _send_email(to: str, subject: str, html: str):
    if not RESEND_API_KEY:
        logger.warning(f"RESEND_API_KEY not set — skipped sending {subject!r} to {to}")
        return
    try:
        async with httpx.AsyncClient(timeout=10) as http:
            resp = await http.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {RESEND_API_KEY}"},
                json={"from": RESEND_FROM, "to": [to], "subject": subject, "html": html},
            )
            if resp.status_code >= 400:
                logger.warning(f"Resend API error sending to {to}: {resp.status_code} {resp.text}")
    except Exception:
        logger.warning(f"Failed to send email to {to}", exc_info=True)

# ── Email-verification / password-reset tokens ──────────────────────────
# See schema.sql's auth_tokens table. Only the raw token's SHA-256 hash is ever stored --
# same reasoning as password hashing: a DB leak alone can't be used to claim a still-valid
# link. Not company-scoped (excluded from TENANT_COLLECTIONS, same as users/companies) --
# looked up by hash alone, before any tenant context is known.
async def _create_auth_token(user_id: str, purpose: str, ttl_minutes: int) -> str:
    raw = secrets.token_urlsafe(32)
    now = datetime.now(timezone.utc)
    await db.auth_tokens.insert_one({
        "id": str(uuid.uuid4()), "user_id": user_id, "purpose": purpose,
        "token_hash": hashlib.sha256(raw.encode()).hexdigest(),
        "expires_at": (now + timedelta(minutes=ttl_minutes)).isoformat(),
        "used_at": None, "created_at": now.isoformat(),
    })
    return raw

async def _consume_auth_token(raw_token: str, purpose: str) -> Optional[dict]:
    """Validates and single-use-marks a token in one step. Returns None (rather than
    raising) for any invalid/expired/already-used/wrong-purpose token -- callers can't
    tell those apart, which is deliberate: distinguishing them would let an attacker probe
    which failure mode a guessed token hit."""
    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
    rec = await db.auth_tokens.find_one({"token_hash": token_hash, "purpose": purpose})
    if not rec or rec.get("used_at"):
        return None
    if datetime.now(timezone.utc) > datetime.fromisoformat(rec["expires_at"]):
        return None
    await db.auth_tokens.update_one({"id": rec["id"]}, {"$set": {"used_at": datetime.now(timezone.utc).isoformat()}})
    return rec

@app.get("/api/health")
async def health(): return {"status": "ok", "service": "Hamro G&G Auto OS"}

# ── Auth Helpers ──────────────────────────────────────────────────────
def hash_pw(pw: str) -> str: return pwd_context.hash(pw)
def verify_pw(pw: str, hashed: str) -> bool: return pwd_context.verify(pw, hashed)

# bcrypt is CPU-bound and blocks the event loop; run it in a worker thread so one
# slow login doesn't stall every other concurrent request on this single-worker server.
async def hash_pw_async(pw: str) -> str: return await asyncio.to_thread(hash_pw, pw)
async def verify_pw_async(pw: str, hashed: str) -> bool: return await asyncio.to_thread(verify_pw, pw, hashed)

def create_token(user_id: str, username: str, role: str = "admin", company_id: Optional[str] = None) -> str:
    return jwt.encode(
        {"user_id": user_id, "username": username, "role": role, "company_id": company_id,
         "exp": datetime.now(timezone.utc) + timedelta(days=7)},
        JWT_SECRET, algorithm="HS256"
    )

async def get_current_user(creds: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    # platform_owner is the one role that legitimately has no company_id -- its endpoints
    # pass explicit company_id filters per call instead of relying on auto-scoping. Any
    # other role showing up here with no company_id means its token was minted while the
    # account had none set (e.g. a stale token from before this account was assigned to a
    # company) -- letting that through would make _ScopedCollection silently stop scoping
    # this session's reads/writes at all (see _scoped/_stamped above), which is exactly
    # how 3 real sales went untagged and invisible on 2026-08-18. Reject it here, at the
    # first possible point, instead of letting a broken session touch any data.
    if payload.get("role") != "platform_owner" and not payload.get("company_id"):
        raise HTTPException(401, "Your session is missing company info -- please log out and log back in.")
    # Every tenant-scoped db.<collection> call for the rest of this request now
    # automatically filters/stamps by this company_id -- see _ScopedCollection above.
    current_company_id.set(payload.get("company_id"))
    return payload

# ── Role-based access control ─────────────────────────────────────────
# "admin" (Admin) always has full access. Other roles only get what's
# listed here as {resource: {allowed actions}}; anything not listed is denied.
ROLE_PERMISSIONS = {
    "stock_supervisor": {  # Front desk stock
        "vehicles": {"view", "create", "edit"},
        "vehicle_media": {"view", "create", "delete"},
        "expenses": {"view", "create", "delete"},
        "jobs": {"view"},
        "customers": {"view", "create", "edit", "delete"},
        "sales": {"view", "create"},
        "team": {"view", "create", "edit", "delete"},
        "vendor_lookup": {"view", "create"},  # vendor picker + inline "add new vendor" when picking a vehicle's purchase source
    },
    "parts_supervisor": {  # Parts department
        "spare_parts": {"view", "create", "edit", "delete"},
        "jobs": {"view", "create", "edit", "delete"},
        "vehicles": {"view", "edit_status"},  # read-only vehicle data, plus limited pipeline-status changes (see PARTS_ALLOWED_STATUSES)
        "vehicle_media": {"view"},  # read-only, so opening a vehicle's detail page doesn't 403 loading photos/documents
        "vendor_lookup": {"view", "create"},  # supplier picker + inline "add new vendor" on a part
        "vendors": {"view", "edit", "delete", "manage_payments"},  # full Vendor Management tab access
        "team": {"view", "create", "edit", "delete"},
    },
}

# parts_supervisor's "edit_status" permission is intentionally narrower than full vehicle "edit":
# they can only move a vehicle between Available / In Repair (job cards require In Repair, see
# POST /jobs) and mark it Scrap. Sold/Reserved/Unlisted and all other vehicle fields stay
# restricted to admin / stock_supervisor.
PARTS_ALLOWED_STATUSES = {"available", "in_repair", "scrap"}

def require(resource: str, action: str):
    async def _checker(cu: dict = Depends(get_current_user)):
        role = cu.get("role", "admin")
        if role == "admin":
            return cu
        if action in ROLE_PERMISSIONS.get(role, {}).get(resource, set()):
            return cu
        raise HTTPException(403, "You do not have permission to perform this action")
    return _checker

def require_any(resource: str, actions: set):
    async def _checker(cu: dict = Depends(get_current_user)):
        role = cu.get("role", "admin")
        if role == "admin":
            return cu
        if ROLE_PERMISSIONS.get(role, {}).get(resource, set()) & actions:
            return cu
        raise HTTPException(403, "You do not have permission to perform this action")
    return _checker

async def admin_only(cu: dict = Depends(get_current_user)):
    if cu.get("role", "admin") != "admin":
        raise HTTPException(403, "This section is restricted to Admin accounts")
    return cu

# Not signup-able and not the old Super Admin console -- a role that only ever gets granted
# by hand (UPDATE users SET role='platform_owner' ...) to one specific, already-existing
# account. Logs in through the exact same plain username/password form as everyone else, no
# separate flow -- the previous multi-tenant attempt's "company code" login field is exactly
# what confused people, so this deliberately doesn't reintroduce anything like it. A
# platform_owner's own token has no company_id (see /platform/* routes below, which take an
# explicit company_id instead of relying on the automatic per-request scoping).
async def platform_owner_only(cu: dict = Depends(get_current_user)):
    if cu.get("role") != "platform_owner":
        raise HTTPException(403, "This section is restricted to the platform owner")
    return cu

def stock_aging(purchase_date_str: str) -> dict:
    try:
        s = str(purchase_date_str)
        d = datetime.fromisoformat(s.replace('Z', '+00:00')) if 'T' in s else datetime.strptime(s, '%Y-%m-%d').replace(tzinfo=timezone.utc)
        days = (datetime.now(timezone.utc) - d).days
    except:
        days = 0
    if days <= 30:   return {"days": days, "category": "fresh",  "label": "Fresh Stock"}
    elif days <= 45: return {"days": days, "category": "normal", "label": "Normal"}
    elif days <= 60: return {"days": days, "category": "slow",   "label": "Slow Moving"}
    else:            return {"days": days, "category": "dead",   "label": "Dead Stock Alert"}

# Front desk stock can record purchase price at intake and manage expenses,
# but must never see the numbers that reveal profit on existing stock. They do see
# minimum_selling_price, though — they need the negotiation floor to work a sale.
FRONT_DESK_HIDDEN_VEHICLE_FIELDS = {
    "purchase_price", "total_investment", "expected_profit", "profit_margin",
    "low_margin", "accessories_cost",
}

# Parts department doesn't handle pricing or sales at all, so on top of the front-desk-hidden
# fields they also never see the selling price or minimum selling price.
PARTS_HIDDEN_VEHICLE_FIELDS = FRONT_DESK_HIDDEN_VEHICLE_FIELDS | {"selling_price", "minimum_selling_price"}

def _hide_financials_for_role(v: dict, role: str) -> dict:
    if role == "stock_supervisor":
        for f in FRONT_DESK_HIDDEN_VEHICLE_FIELDS:
            v.pop(f, None)
    elif role == "parts_supervisor":
        for f in PARTS_HIDDEN_VEHICLE_FIELDS:
            v.pop(f, None)
    return v

# A job card's contribution to vehicle cost — the settled actual_cost once the job is
# marked complete, or its estimate while still pending/in progress so the work isn't
# invisible from the vehicle's expense total until someone closes it out.
def _job_card_cost(jc: dict) -> float:
    ac = jc.get("actual_cost")
    return ac if ac is not None else jc.get("estimated_cost", 0)

async def enrich_vehicle(v: dict) -> dict:
    v["aging"] = stock_aging(v.get("purchase_date", ""))
    exps = await db.expenses.find({"vehicle_id": v["id"]}, {"_id": 0}).to_list(200)
    jobs = await db.job_cards.find({"vehicle_id": v["id"]}, {"_id": 0}).to_list(200)
    total_exp = sum(e["amount"] for e in exps) + sum(_job_card_cost(j) for j in jobs)
    v["total_expenses"] = total_exp
    v["total_investment"] = v.get("purchase_price", 0) + total_exp + v.get("accessories_cost", 0)
    sp = v.get("selling_price") or 0
    if sp > 0:
        v["expected_profit"] = sp - v["total_investment"]
        v["profit_margin"] = round((v["expected_profit"] / sp) * 100, 2)
        v["low_margin"] = v["profit_margin"] < 8
    else:
        v["expected_profit"] = None; v["profit_margin"] = None; v["low_margin"] = False
    return v

# Sold vehicles carry a 6-month warranty — a job card can still be opened against
# one after sale, as long as it's within that window from sold_date.
VEHICLE_WARRANTY_DAYS = 182  # ~6 months

def _within_warranty(vehicle: dict) -> bool:
    sold_date = vehicle.get("sold_date")
    if not sold_date: return False
    try:
        d = datetime.strptime(str(sold_date)[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except (ValueError, TypeError):
        return False
    return 0 <= (datetime.now(timezone.utc) - d).days <= VEHICLE_WARRANTY_DAYS

# ── Helper: compute total investment for a vehicle ────────────────────
async def _vehicle_investment(vehicle_id: str, vehicle: dict) -> float:
    """Returns purchase_price + accessories + all expenses + all job card costs for a vehicle."""
    exps = await db.expenses.find({"vehicle_id": vehicle_id}, {"_id": 0}).to_list(200)
    jobs = await db.job_cards.find({"vehicle_id": vehicle_id}, {"_id": 0}).to_list(200)
    return (vehicle.get("purchase_price", 0) + vehicle.get("accessories_cost", 0)
            + sum(e["amount"] for e in exps) + sum(_job_card_cost(j) for j in jobs))

# Job card cost for a sale's Extra Expenses card — display only. Warranty job cards
# (opened after the sale) are excluded, same reasoning as investment: that's free
# post-sale service, not a cost of the sale itself. This number is NOT part of
# expenses_total/total_amount/due_amount — see _strip_job_card_extra_expenses — so it
# never changes what the customer is billed; it only tells the shop what to expect the
# vehicle's job cards to cost against margin.
async def _job_card_cost_total(vehicle_id: str) -> float:
    jobs = await db.job_cards.find({"vehicle_id": vehicle_id, "is_warranty": {"$ne": True}}, {"_id": 0}).to_list(200)
    return sum(_job_card_cost(j) for j in jobs)

# For a sale returned via POST /vehicles/{vid}/return, only the retained (unrefunded) portion
# is real revenue — the rest went back to the customer. Every revenue/profit aggregate should
# read a sale's amount through this instead of "total_amount" directly.
#
# Deliberately uses "sale_price", not "total_amount": total_amount also folds in extra_expenses
# like registration transfer or insurance transfer — fees the shop collects from the buyer
# purely to forward to a third party (RTO, insurer...), not markup on the vehicle. Counting
# those as revenue would inflate profit/margin with no matching cost on the other side.
#
# Job card costs are NEVER added to a sale's extra_expenses (see _strip_job_card_extra_expenses) —
# they reduce margin purely by being counted in _vehicle_investment, without ever touching
# what the customer is billed. So sale_price alone is always the right revenue figure here.
def _sale_revenue(s: dict) -> float:
    return s.get("retained_amount", 0) if s.get("returned") else s.get("sale_price", 0)

# A vehicle can be tied to a vendor two ways: the legacy vendor_id field, or the
# newer linked_contact_type/id (set via the Customer/Vendor picker on the stock form).
# Vendor aggregates (list page, ledger, payables) must count both, or a vehicle linked
# only through the picker silently vanishes from that vendor's numbers.
def _vendor_vehicle_filter(vendor_id: str) -> dict:
    return {"$or": [{"vendor_id": vendor_id}, {"linked_contact_type": "vendor", "linked_contact_id": vendor_id}]}

# ── Helper: compute total amount owed to a vendor (payable) ──────────
async def _vendor_payable(vendor_id: str) -> float:
    """Returns max(0, total_purchased - total_paid) for a vendor."""
    veh = await db.vehicles.find(_vendor_vehicle_filter(vendor_id), {"_id": 0}).to_list(200)
    owed = sum(v.get("purchase_price", 0) for v in veh)
    pmts = await db.vendor_payments.find({"vendor_id": vendor_id}, {"_id": 0}).to_list(200)
    paid = sum(p["amount"] for p in pmts)
    return max(0.0, owed - paid)

# ── Batched variants of the above — same results, one round-trip per list
# instead of one per item. Used by endpoints that loop over many vehicles/
# vendors (dashboard, accounting summary) where per-item queries add up to
# a very visible delay once the DB isn't in the same region as the backend.
async def _batch_vehicle_investment(vehicles: list) -> dict:
    vehicle_ids = [v["id"] for v in vehicles]
    if not vehicle_ids:
        return {}
    all_exps = await db.expenses.find({"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0}).to_list(20000)
    exps_by_vehicle: dict = {}
    for e in all_exps:
        exps_by_vehicle.setdefault(e["vehicle_id"], []).append(e)
    all_jobs = await db.job_cards.find({"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0}).to_list(20000)
    job_cost_by_vehicle: dict = {}
    for j in all_jobs:
        job_cost_by_vehicle[j["vehicle_id"]] = job_cost_by_vehicle.get(j["vehicle_id"], 0) + _job_card_cost(j)
    return {
        v["id"]: v.get("purchase_price", 0) + v.get("accessories_cost", 0)
        + sum(e["amount"] for e in exps_by_vehicle.get(v["id"], []))
        + job_cost_by_vehicle.get(v["id"], 0)
        for v in vehicles
    }

async def _batch_vendor_payable(vendor_ids: list) -> dict:
    if not vendor_ids:
        return {}
    veh = await db.vehicles.find(
        {"$or": [{"vendor_id": {"$in": vendor_ids}}, {"linked_contact_type": "vendor", "linked_contact_id": {"$in": vendor_ids}}]},
        {"_id": 0, "vendor_id": 1, "linked_contact_type": 1, "linked_contact_id": 1, "purchase_price": 1},
    ).to_list(20000)
    owed: dict = {}
    for v in veh:
        vid = v.get("vendor_id") or (v.get("linked_contact_id") if v.get("linked_contact_type") == "vendor" else None)
        if vid:
            owed[vid] = owed.get(vid, 0) + v.get("purchase_price", 0)
    pmts = await db.vendor_payments.find({"vendor_id": {"$in": vendor_ids}}, {"_id": 0}).to_list(20000)
    paid: dict = {}
    for p in pmts:
        paid[p["vendor_id"]] = paid.get(p["vendor_id"], 0) + p["amount"]
    return {vid: max(0.0, owed.get(vid, 0) - paid.get(vid, 0)) for vid in vendor_ids}

# ── Helper: compute total remaining balance for an EMI record ─────────
async def _emi_remaining(emi_id: str, loan_amount: float) -> float:
    pmts = await db.emi_payments.find({"emi_id": emi_id}, {"_id": 0}).to_list(200)
    paid = sum(p["amount"] for p in pmts)
    return max(0.0, loan_amount - paid)

async def _batch_emi_remaining(emis: list) -> dict:
    emi_ids = [e["id"] for e in emis]
    if not emi_ids:
        return {}
    all_pmts = await db.emi_payments.find({"emi_id": {"$in": emi_ids}}, {"_id": 0}).to_list(20000)
    paid_by_emi: dict = {}
    for p in all_pmts:
        paid_by_emi[p["emi_id"]] = paid_by_emi.get(p["emi_id"], 0) + p["amount"]
    return {e["id"]: max(0.0, e.get("loan_amount", 0) - paid_by_emi.get(e["id"], 0)) for e in emis}

# ── Helper: classify aging counts for a list of available vehicles ────
def _aging_counts(vehicles: list) -> dict:
    counts = {"fresh": 0, "normal": 0, "slow": 0, "dead": 0}
    for v in vehicles:
        cat = stock_aging(v.get("purchase_date", ""))["category"]
        counts[cat] = counts.get(cat, 0) + 1
    return counts

# ── Helper: build ai_suggestions context data by type ────────────────
async def _build_suggestions_context(context_type: str) -> dict:
    if context_type == "inventory":
        vehicles = await db.vehicles.find({"status": "available"}, {"_id": 0}).to_list(50)
        slow_list = []
        for v in vehicles:
            ag = stock_aging(v.get("purchase_date", ""))
            if ag["category"] in ["slow", "dead"]:
                exps = await db.expenses.find({"vehicle_id": v["id"]}, {"_id": 0}).to_list(50)
                slow_list.append({
                    "brand": v.get("brand"), "model": v.get("model"),
                    "days": ag["days"], "category": ag["category"],
                    "purchase_price": v.get("purchase_price"), "selling_price": v.get("selling_price"),
                    "total_investment": v.get("purchase_price", 0) + sum(e["amount"] for e in exps)
                })
        return {"available_count": len(vehicles), "slow_and_dead_stock": slow_list[:6]}
    if context_type == "finance":
        avail = await db.vehicles.find({"status": "available"}, {"_id": 0}).to_list(500)
        locked = sum(v.get("purchase_price", 0) for v in avail)
        return {
            "locked_capital_NPR": locked,
            "total_vehicles": await db.vehicles.count_documents({}),
            "sold_vehicles": await db.vehicles.count_documents({"status": "sold"})
        }
    if context_type == "customer":
        custs = await db.customers.find({}, {"_id": 0}).to_list(20)
        return {"total_customers": len(custs), "customers": [{"name": c["name"], "contact": c.get("contact_number")} for c in custs[:8]]}
    if context_type == "festival":
        avail = await db.vehicles.find({"status": "available"}, {"_id": 0}).to_list(50)
        return {"available_stock": len(avail), "vehicles": [{"brand": v.get("brand"), "model": v.get("model"), "price": v.get("selling_price")} for v in avail[:10]]}
    if context_type == "vendor":
        vendors = await db.vendors.find({}, {"_id": 0}).to_list(20)
        return {"total_vendors": len(vendors), "vendors_with_due": sum(1 for _ in vendors)}
    return {}

# ── AI Assistant (Gemini-powered) ──────────────────────────────────────
class AIChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None

class AIVehicleInput(BaseModel):
    brand: str; model: str; year: int
    engine_cc: Optional[int] = None
    fuel_type: Optional[str] = "Petrol"
    condition: Optional[str] = "Good"
    ownership_number: Optional[int] = 1
    kilometer_run: Optional[int] = None
    purchase_price: Optional[float] = None

class AIPriceRequest(BaseModel):
    vehicle: AIVehicleInput

class AISuggestionsRequest(BaseModel):
    context_type: str
    additional_context: Optional[str] = None

MARKDOWN_NOTE = 'Format your reply as plain text with **bold** for emphasis and "- " for bullet points only — no headers, links, or tables.'

async def _ai_text(system: str, contents, max_tokens: int = 1024) -> str:
    if not ai_client:
        raise HTTPException(503, "AI Assistant is not configured. Set GEMINI_API_KEY on the server.")
    try:
        resp = await ai_client.aio.models.generate_content(
            model=AI_MODEL, contents=contents,
            config=genai_types.GenerateContentConfig(system_instruction=system, max_output_tokens=max_tokens),
        )
    except genai_errors.APIError as e:
        # Previously unhandled — any Gemini-side error (rate limit, transient outage,
        # quota) crashed through as a bare 500 with no JSON body, which is why the
        # frontend fell back to a generic "Sorry, I couldn't connect" with no real
        # cause. Surfacing the actual code/message here lets the UI show something
        # the user can act on (e.g. "rate limited, try again shortly").
        logger.warning(f"Gemini API error {e.code}: {e.message}", exc_info=True)
        if e.code == 429:
            raise HTTPException(429, "The AI is getting a lot of requests right now (Gemini rate limit) — please wait a few seconds and try again.")
        raise HTTPException(502, f"AI Assistant couldn't reach Gemini right now ({e.message or f'error {e.code}'}). Please try again.")
    return resp.text or ""

# Builds a comprehensive, read-only snapshot of live business data for the AI assistant's
# context — recomputed fresh on every chat message so it can answer virtually any factual
# question (inventory, sales/dues, customers, vendors, parts, EMI) without a separate
# tool-calling round trip. Every collection is bulk-fetched once and aggregated in Python
# (no N+1 per-record queries), so this stays fast even as the business grows.
async def _build_ai_business_snapshot() -> str:
    all_vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(3000)
    active_vehicles = [v for v in all_vehicles if v.get("status") != "sold"]
    sold_vehicles = [v for v in all_vehicles if v.get("status") == "sold"]
    active_ids = [v["id"] for v in active_vehicles]
    photo_ids = set(await db.vehicle_photos.distinct("vehicle_id", {"vehicle_id": {"$in": active_ids}}))

    by_status: dict = {}
    for v in active_vehicles:
        by_status[v.get("status", "?")] = by_status.get(v.get("status", "?"), 0) + 1
    avail = [v for v in active_vehicles if v.get("status") == "available"]
    no_photo = [v for v in active_vehicles if v["id"] not in photo_ids]
    no_price = [v for v in active_vehicles if not v.get("selling_price") and v.get("status") != "scrap"]

    def _vline(v):
        bits = [
            f"{v.get('brand')} {v.get('model')} {v.get('year')}", v.get("status", "?"),
            f"{v.get('kilometer_run') or '?'}km",
            f"Rs. {v['selling_price']}" if v.get("selling_price") else "no price set",
        ]
        if v.get("registration_number"): bits.append(f"reg {v['registration_number']}")
        if v["id"] not in photo_ids: bits.append("NO PHOTO")
        return "- " + " · ".join(bits)
    vehicle_lines = "\n".join(_vline(v) for v in active_vehicles[:250]) or "(none)"

    all_sales = await db.sales.find({}, {"_id": 0}).sort("sale_date", -1).to_list(3000)
    total_revenue = round(sum(_sale_revenue(s) for s in all_sales), 2)
    due_sales = [s for s in all_sales if s.get("due_amount", 0) > 0]
    total_due = round(sum(s.get("due_amount", 0) for s in all_sales), 2)
    today_iso = datetime.now(timezone.utc).date().isoformat()
    overdue_sales = [s for s in due_sales if s.get("due_date") and s.get("due_date") < today_iso]
    overdue_ids = {s["id"] for s in overdue_sales}
    due_lines = "\n".join(
        f"- {s.get('customer_name') or 'Walk-in'} · {s.get('vehicle_info')} · owes Rs. {s.get('due_amount')}"
        + (f" (OVERDUE, was due {s['due_date']})" if s["id"] in overdue_ids else (f" (due {s['due_date']})" if s.get("due_date") else ""))
        for s in due_sales[:100]
    ) or "(none)"

    customers = await db.customers.find({}, {"_id": 0}).to_list(3000)
    due_by_customer: dict = {}
    for s in all_sales:
        if s.get("customer_id") and s.get("due_amount", 0) > 0:
            due_by_customer[s["customer_id"]] = due_by_customer.get(s["customer_id"], 0) + s["due_amount"]
    customers_with_due = [
        f"- {c.get('name')} ({c.get('contact_number') or 'no phone'}) owes Rs. {round(due_by_customer[c['id']], 2)}"
        for c in customers if c.get("id") in due_by_customer
    ]
    customers_due_lines = "\n".join(customers_with_due[:100])
    if len(customers_with_due) > 100: customers_due_lines += f"\n(+{len(customers_with_due) - 100} more not listed)"

    vendors = await db.vendors.find({}, {"_id": 0}).to_list(500)
    all_parts = await db.spare_parts.find({}, {"_id": 0}).to_list(3000)
    all_vendor_payments = await db.vendor_payments.find({}, {"_id": 0}).to_list(3000)
    owed_by_vendor: dict = {}
    for v in all_vehicles:  # payable doesn't disappear once a vehicle sells
        vid = v.get("vendor_id") or (v.get("linked_contact_id") if v.get("linked_contact_type") == "vendor" else None)
        if vid: owed_by_vendor[vid] = owed_by_vendor.get(vid, 0) + v.get("purchase_price", 0)
    for p in all_parts:
        if p.get("vendor_id"):
            owed_by_vendor[p["vendor_id"]] = owed_by_vendor.get(p["vendor_id"], 0) + p.get("quantity", 0) * p.get("unit_cost", 0)
    paid_by_vendor: dict = {}
    for p in all_vendor_payments:
        paid_by_vendor[p["vendor_id"]] = paid_by_vendor.get(p["vendor_id"], 0) + p["amount"]
    vendors_with_due = []
    for v in vendors:
        due = max(0.0, owed_by_vendor.get(v["id"], 0) - paid_by_vendor.get(v["id"], 0))
        if due > 0: vendors_with_due.append(f"- {v.get('name')} ({v.get('phone') or 'no phone'}) — we owe Rs. {round(due, 2)}")
    vendor_due_lines = "\n".join(vendors_with_due[:100]) or "(none)"
    if len(vendors_with_due) > 100: vendor_due_lines += f"\n(+{len(vendors_with_due) - 100} more not listed)"

    low_stock = [p for p in all_parts if p.get("quantity", 0) <= p.get("min_stock_alert", 2)]
    low_stock_lines = "\n".join(
        f"- {p.get('name')} ({p.get('category') or '?'}): {p.get('quantity', 0)} left, alert at {p.get('min_stock_alert', 2)}"
        for p in low_stock[:80]
    ) or "(none)"

    emi_records = await db.emi_records.find({}, {"_id": 0}).to_list(1000)
    emi_payments = await db.emi_payments.find({}, {"_id": 0}).to_list(5000)
    paid_by_emi: dict = {}
    for p in emi_payments:
        paid_by_emi[p["emi_id"]] = paid_by_emi.get(p["emi_id"], 0) + p["amount"]
    for e in emi_records:
        e["remaining_balance"] = max(0, e.get("loan_amount", 0) - paid_by_emi.get(e["id"], 0))
    active_emis = [e for e in emi_records if e["remaining_balance"] > 0]
    emi_receivable = round(sum(e["remaining_balance"] for e in active_emis), 2)

    return f"""=== LIVE BUSINESS DATA (fetched fresh for this message — answer directly from this, never say you lack backend access) ===

INVENTORY — {len(active_vehicles)} active vehicles ({', '.join(f'{k}: {n}' for k, n in by_status.items()) or 'none'}), {len(sold_vehicles)} sold historically.
{len(avail)} currently listed available for sale. {len(no_photo)} have NO photo uploaded. {len(no_price)} have no selling price set.
Vehicle list:
{vehicle_lines}

SALES — {len(all_sales)} total sales, Rs. {total_revenue} total revenue.
{len(due_sales)} sale(s) have an outstanding due, Rs. {total_due} total due, {len(overdue_sales)} of those are overdue.
Sales with dues:
{due_lines}

CUSTOMERS — {len(customers)} total. {len(customers_with_due)} currently owe money:
{customers_due_lines or '(none)'}

VENDORS — {len(vendors)} total. We currently owe {len(vendors_with_due)} of them money:
{vendor_due_lines}

SPARE PARTS — {len(all_parts)} parts tracked. {len(low_stock)} at or below their low-stock threshold:
{low_stock_lines}

EMI / FINANCING — {len(emi_records)} plans total, {len(active_emis)} still active, Rs. {emi_receivable} still receivable.
=== END LIVE DATA ==="""

@api_router.post("/ai/chatbot")
async def ai_chatbot(req: AIChatRequest, cu: dict = Depends(admin_only)):
    session_id = req.session_id or str(uuid.uuid4())
    session = await db.ai_chat_sessions.find_one({"id": session_id}, {"_id": 0})
    history = session["messages"] if session else []  # stored as [{"role": "user"|"assistant", "text": "..."}]

    settings = await db.settings.find_one({}, {"_id": 0}) or {}
    business_snapshot = await _build_ai_business_snapshot()

    system = (
        f"You are the AI assistant for {settings.get('business_name', 'Hamro G&G Auto')}, "
        "a used motorbike/scooter dealership in Nepal. You're used both internally by dealership staff "
        "(who can ask about anything in the live data below — inventory, photos, prices, sales, dues, "
        "customers, vendors, parts stock, EMI) and to answer prospective-customer questions about "
        "available vehicles, prices, and financing. Prices are in NPR. Be warm, concise, and helpful, "
        "and answer directly from the data given — never claim you lack backend access, it's provided "
        "below. If something genuinely isn't in the data, say so honestly rather than guessing. "
        "Politely decline only requests entirely unrelated to this dealership (e.g. general trivia, "
        "coding help, other businesses).\n\n"
        f"{business_snapshot}\n\n{MARKDOWN_NOTE}"
    )
    messages = history + [{"role": "user", "text": req.message}]
    # Gemini's role name for assistant turns is "model", not "assistant".
    contents = [
        genai_types.Content(role="model" if m["role"] == "assistant" else "user", parts=[genai_types.Part.from_text(text=m["text"])])
        for m in messages
    ]
    # Higher cap than the other AI endpoints — with full business-data access, replies can
    # legitimately need to enumerate dozens of vehicles/customers/parts, and the default
    # 1024 was cutting long list answers off mid-sentence.
    reply = await _ai_text(system, contents, max_tokens=4096)
    messages.append({"role": "assistant", "text": reply})

    await db.ai_chat_sessions.update_one(
        {"id": session_id},
        {"$set": {"messages": messages[-20:], "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return {"reply": reply, "session_id": session_id}

@api_router.post("/ai/price-suggestion")
async def ai_price_suggestion(req: AIPriceRequest, cu: dict = Depends(admin_only)):
    v = req.vehicle
    similar = await db.vehicles.find(
        {"status": "sold", "brand": {"$regex": f"^{v.brand}$", "$options": "i"}, "model": {"$regex": f"^{v.model}$", "$options": "i"}},
        {"_id": 0, "year": 1, "selling_price": 1, "kilometer_run": 1, "condition": 1, "sold_date": 1},
    ).sort("sold_date", -1).to_list(10)

    history_lines = "\n".join(
        f"- {s.get('year')} · {s.get('kilometer_run') or '?'}km · {s.get('condition') or '?'} condition · sold for Rs. {s.get('selling_price')}"
        for s in similar
    ) or "No sold-history records for this brand/model in this dealership."

    system = (
        "You are a pricing analyst for a Nepali used-motorbike dealership. Recommend a selling price range in NPR "
        "for the vehicle described, reasoning from the sold-history comparables given and the general Nepal used-bike market. "
        f"{MARKDOWN_NOTE}"
    )
    # ownership_number > 90 encodes a transcript (bluebook lost/damaged, duplicate copy
    # issued): TRANSCRIPT_BASE(90) + issue count — mirrors TRANSCRIPT_BASE in helpers.js.
    def _ordinal(n): return f"{n}{'st' if n == 1 else 'nd' if n == 2 else 'rd' if n == 3 else 'th'}"
    ownership_desc = f"on its {_ordinal(v.ownership_number - 90)} transcript bluebook (original lost/damaged)" if v.ownership_number > 90 \
        else f"{_ordinal(v.ownership_number)} owner"
    prompt = (
        f"Vehicle to price: {v.brand} {v.model} {v.year}, {v.engine_cc or '?'}cc {v.fuel_type}, "
        f"{ownership_desc}, {v.kilometer_run or '?'}km, "
        f"condition: {v.condition}, purchase price: Rs. {v.purchase_price or 'unknown'}.\n\n"
        f"Similar sold vehicles from this dealership's history:\n{history_lines}\n\n"
        "Give a recommended selling price range and a brief justification."
    )
    suggestion = await _ai_text(system, prompt)
    return {"suggestion": suggestion, "sold_history_count": len(similar)}

@api_router.get("/ai/festival-intelligence")
async def ai_festival_intelligence(cu: dict = Depends(admin_only)):
    avail = await db.vehicles.find({"status": "available"}, {"_id": 0, "brand": 1}).to_list(1000)
    stock_snapshot: dict = {}
    for v in avail:
        b = v.get("brand") or "Other"
        stock_snapshot[b] = stock_snapshot.get(b, 0) + 1

    today = datetime.now(timezone.utc).date().isoformat()
    system = (
        "You are a business intelligence advisor for a Nepali used-motorbike dealership. Nepali festivals "
        "(Dashain, Tihar, etc.) fall on the lunar Bikram Sambat calendar and shift each Gregorian year — give "
        f"approximate timing and say so explicitly rather than exact dates. {MARKDOWN_NOTE}"
    )
    prompt = (
        f"Today's date: {today}. Current available stock by brand: {json.dumps(stock_snapshot)}.\n\n"
        "Identify the nearest major upcoming Nepali festival(s) and give stock and pricing strategy advice for "
        "a used-bike dealership heading into that season — which brands/segments to push, and any pricing moves."
    )
    intelligence = await _ai_text(system, prompt)
    return {"intelligence": intelligence, "stock_snapshot": stock_snapshot}

@api_router.post("/ai/suggestions")
async def ai_suggestions(req: AISuggestionsRequest, cu: dict = Depends(admin_only)):
    context = await _build_suggestions_context(req.context_type)
    system = (
        "You are a business advisor for a Nepali used-motorbike dealership. Give concrete, actionable "
        f"recommendations grounded only in the data provided — do not invent numbers not given. {MARKDOWN_NOTE}"
    )
    prompt = f"Context ({req.context_type}): {json.dumps(context)}\n"
    if req.additional_context:
        prompt += f"\nAdditional note from the owner: {req.additional_context}\n"
    prompt += "\nGive your recommendations."
    suggestions = await _ai_text(system, prompt)
    return {"suggestions": suggestions}


class LoginRequest(BaseModel):
    username: str; password: str

class RegisterRequest(BaseModel):
    # Same reasoning as SignUpRequest.email: stored in the `username` column, so this only
    # affects newly-created accounts -- existing plain-username staff accounts are unaffected.
    email: EmailStr; password: str = Field(min_length=8); name: str
    role: str = "sales_staff"

class SignUpRequest(BaseModel):
    # Email, not a free-text username: email is naturally globally unique (unlike everyone
    # wanting "admin"), and is what a real password-reset flow needs to send a link to.
    # Stored in the same `username` column existing accounts already use (plain usernames
    # like "admin" keep logging in unchanged) -- this only changes what NEW signups collect.
    name: str; company_name: str
    email: EmailStr; password: str = Field(min_length=8)

class ChangePasswordRequest(BaseModel):
    current_password: str; new_password: str = Field(min_length=8)

class ResendVerificationRequest(BaseModel):
    email: EmailStr

class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str; new_password: str = Field(min_length=8)

class UserUpdate(BaseModel):
    name: str; role: str
    password: Optional[str] = Field(default=None, min_length=8)

class VehicleCreate(BaseModel):
    brand: str; model: str
    variant: Optional[str] = None
    year: int; engine_cc: int
    fuel_type: str = "Petrol"
    vehicle_type: str = "bike"  # "bike" | "scooter" — drives hamroauto.com.np's type filter
    ownership_number: Optional[int] = None  # may not be known yet at initial stock entry (bluebook not on hand) — filled in later via edit
    chassis_number: Optional[str] = None
    engine_number: Optional[str] = None
    kilometer_run: Optional[int] = None
    condition: str = "Good"
    condition_rating: int = 7
    color: Optional[str] = None
    registration_number: str
    purchase_price: float
    accessories_cost: float = 0
    purchase_date: str
    purchase_source: str
    vendor_id: Optional[str] = None
    purchase_from: Optional[str] = None
    # Optional link to an existing/newly-created Customer or Vendor record, independent of
    # vendor_id/purchase_from above — lets stock entry reference a customer (e.g. a trade-in
    # / exchange source) as well as a vendor, picked via the Customer/Vendor combobox on the form.
    linked_contact_type: Optional[str] = None  # "customer" | "vendor"
    linked_contact_id: Optional[str] = None
    linked_contact_name: Optional[str] = None
    selling_price: Optional[float] = None
    minimum_selling_price: Optional[float] = None
    notes: Optional[str] = None
    status: str = "available"
    bluebook_status: str = "pending"
    insurance_status: str = "pending"
    tax_clearance_status: str = "pending"
    transfer_status: str = "pending"

class VehicleUpdate(BaseModel):
    brand: Optional[str] = None; model: Optional[str] = None
    variant: Optional[str] = None; year: Optional[int] = None
    engine_cc: Optional[int] = None; fuel_type: Optional[str] = None
    vehicle_type: Optional[str] = None
    ownership_number: Optional[int] = None
    chassis_number: Optional[str] = None; engine_number: Optional[str] = None
    kilometer_run: Optional[int] = None
    condition: Optional[str] = None; condition_rating: Optional[int] = None
    color: Optional[str] = None; registration_number: Optional[str] = None
    purchase_price: Optional[float] = None; accessories_cost: Optional[float] = None
    purchase_date: Optional[str] = None; purchase_source: Optional[str] = None
    vendor_id: Optional[str] = None; purchase_from: Optional[str] = None
    linked_contact_type: Optional[str] = None; linked_contact_id: Optional[str] = None
    linked_contact_name: Optional[str] = None
    selling_price: Optional[float] = None; minimum_selling_price: Optional[float] = None
    notes: Optional[str] = None; status: Optional[str] = None
    sold_date: Optional[str] = None; customer_id: Optional[str] = None
    salesperson_id: Optional[str] = None; salesperson_name: Optional[str] = None
    discount: Optional[float] = None
    bluebook_status: Optional[str] = None; insurance_status: Optional[str] = None
    tax_clearance_status: Optional[str] = None; transfer_status: Optional[str] = None

class VehicleStatusUpdate(BaseModel):
    status: str

class VehicleReturnRequest(BaseModel):
    refund_amount: float  # NPR amount of the sale's total_amount handed back to the customer, after condition assessment
    new_status: str = "available"  # where the vehicle re-enters stock: available / unlisted / reserved / scrap / in_repair
    notes: Optional[str] = None

class ExpenseCreate(BaseModel):
    vehicle_id: str; category: str; amount: float
    description: Optional[str] = None; date: Optional[str] = None

class JobCardCreate(BaseModel):
    vehicle_id: Optional[str] = None; is_external: bool = False
    # Only used when is_external is True — a walk-in repair vehicle we don't own,
    # so there's no inventory record to pull these from.
    vehicle_brand: Optional[str] = None; vehicle_model: Optional[str] = None
    vehicle_year: Optional[int] = None; registration_number: Optional[str] = None
    customer_name: Optional[str] = None; customer_contact: Optional[str] = None
    work_description: str
    mechanic_id: Optional[str] = None; mechanic_name: str
    estimated_cost: float; notes: Optional[str] = None
    coupon_no: int; job_date: str
    parts: List[dict] = []

class JobCardUpdate(BaseModel):
    work_description: Optional[str] = None; mechanic_name: Optional[str] = None
    estimated_cost: Optional[float] = None; actual_cost: Optional[float] = None
    status: Optional[str] = None; notes: Optional[str] = None
    parts: Optional[List[dict]] = None

class PartStockOut(BaseModel):
    quantity: int
    reason: str  # Sale | Used in Job Card | Damaged | Return
    date: Optional[str] = None
    job_id: Optional[str] = None
    notes: Optional[str] = None

class SaleCreate(BaseModel):
    vehicle_id: str
    customer_id: Optional[str] = None
    sale_price: float
    extra_expenses: List[dict] = []  # [{name: str, amount: float}]
    payment_method: str = "Cash"
    paid_cash: float = 0
    paid_bank: float = 0
    due_date: Optional[str] = None
    sale_date: Optional[str] = None
    notes: Optional[str] = None

class SaleUpdate(BaseModel):
    vehicle_id: str
    customer_id: Optional[str] = None
    sale_price: float
    extra_expenses: List[dict] = []  # [{name: str, amount: float}]
    payment_method: str = "Cash"
    paid_cash: float = 0
    paid_bank: float = 0
    due_date: Optional[str] = None
    sale_date: Optional[str] = None
    notes: Optional[str] = None

class CustomerCreate(BaseModel):
    name: str; contact_number: str
    address: Optional[str] = None
    occupation: Optional[str] = None
    budget_min: Optional[float] = None
    budget_max: Optional[float] = None
    interested_brands: Optional[str] = None
    notes: Optional[str] = None

class TeamMemberCreate(BaseModel):
    name: str; role: str
    contact: Optional[str] = None
    specialization: Optional[str] = None
    commission_rate: Optional[float] = None
    joining_date: Optional[str] = None

class PartnerCreate(BaseModel):
    name: str; capital_contribution: float
    stake_percentage: float; contact: Optional[str] = None

class VendorCreate(BaseModel):
    name: str; phone: str
    address: Optional[str] = None
    notes: Optional[str] = None
    vendor_type: Optional[str] = "both"

class VendorPaymentCreate(BaseModel):
    vendor_id: str; amount: float
    vehicle_id: Optional[str] = None
    payment_date: Optional[str] = None
    notes: Optional[str] = None

class EMICreate(BaseModel):
    customer_id: str; vehicle_id: str
    loan_amount: float; down_payment: float
    interest_rate: float; tenure_months: int
    start_date: str
    financer_name: Optional[str] = None
    notes: Optional[str] = None

class EMIPaymentCreate(BaseModel):
    emi_id: str; amount: float
    payment_date: Optional[str] = None
    notes: Optional[str] = None

class LeadCreate(BaseModel):
    type: str  # "sell" | "exchange" | "service"
    name: str; phone: str
    message: Optional[str] = None
    images: Optional[List[str]] = None
    requested_service: Optional[str] = None  # "service" leads: e.g. "Oil Change"
    vehicle_type: Optional[str] = None  # "service" leads: e.g. "Bike - Pulsar 150"
    preferred_date: Optional[str] = None  # "service" leads: preferred servicing date

class LeadUpdate(BaseModel):
    status: str  # "new" | "contacted" | "closed"

class SettingsUpdate(BaseModel):
    logo_url: Optional[str] = None
    business_name: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    address: Optional[str] = None
    hero_image_url: Optional[str] = None
    service_image_url: Optional[str] = None

# ── AUTH ──────────────────────────────────────────────────────────────
@api_router.post("/auth/signup")
@limiter.limit("3/hour")
async def signup(request: Request, req: SignUpRequest):
    """Public, unauthenticated — anyone can create their own company here. They become
    that company's sole Admin; everything they create from here on (vehicles, customers,
    employee accounts, ...) is automatically isolated to this company_id, enforced at the
    db access layer (see _ScopedCollection) rather than per-endpoint. Login identifiers are
    deliberately kept GLOBALLY unique (not per-company) -- the previous multi-tenant attempt
    required a separate "company code" field on login to disambiguate duplicates across
    companies, which is exactly the UX that confused people before. Using email rather than
    a free-text username sidesteps the "everyone wants to type admin" collision on top of
    that, and is what a real password-reset-by-email flow needs -- see /auth/forgot-password
    below. New signups must also verify their email before they can log in (see
    /auth/verify-email) -- staff accounts an admin creates via /auth/register are exempt
    (that admin is already vouching for the email); only anonymous self-signup is gated,
    since that's the actual abuse case (throwaway/fake emails claiming a company)."""
    email = req.email.lower()
    existing = await db.users.find_one({"username": email})
    if existing:
        raise HTTPException(400, "An account with this email already exists")
    company_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    await db.companies.insert_one({"id": company_id, "name": req.company_name, "created_at": now})
    user = {"id": str(uuid.uuid4()), "username": email,
            "password_hash": await hash_pw_async(req.password), "name": req.name, "role": "admin",
            "company_id": company_id, "email_verified_at": None, "created_at": now}
    await db.users.insert_one(user)
    # A settings row is looked up purely by company_id (see get_settings/update_settings) --
    # "id" here is just an arbitrary row id, not a lookup key.
    await db.settings.insert_one({"id": str(uuid.uuid4()), "company_id": company_id,
                                   "business_name": req.company_name})
    verify_token = await _create_auth_token(user["id"], "verify_email", ttl_minutes=24 * 60)
    link = f"{FRONTEND_URL}/verify-email?token={verify_token}"
    await _send_email(email, "Verify your email — Hamro G&G Auto OS",
                       f"<p>Hi {req.name},</p>"
                       f"<p>Click below to verify your email and activate <b>{req.company_name}</b>'s workspace:</p>"
                       f"<p><a href='{link}'>{link}</a></p>"
                       f"<p>This link expires in 24 hours.</p>")
    return {"message": "Account created — check your email to verify it before signing in.", "email": email}

@api_router.post("/auth/login")
@limiter.limit("5/minute")
async def login(request: Request, req: LoginRequest):
    user = await db.users.find_one({"username": req.username})
    if not user or not await verify_pw_async(req.password, user["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    if not user.get("email_verified_at"):
        raise HTTPException(403, "Please verify your email before signing in — check your inbox, or request a new link.")
    token = create_token(user.get("id", ""), user["username"], user.get("role", "admin"), user.get("company_id"))
    company = await db.companies.find_one({"id": user.get("company_id")}, {"_id": 0, "name": 1}) if user.get("company_id") else None
    return {"token": token, "username": user["username"], "name": user.get("name", user["username"]),
            "role": user.get("role", "admin"), "company_name": company["name"] if company else None}

@api_router.get("/auth/verify-email")
@limiter.limit("10/minute")
async def verify_email(request: Request, token: str):
    rec = await _consume_auth_token(token, "verify_email")
    if not rec:
        raise HTTPException(400, "This verification link is invalid or has expired — request a new one.")
    await db.users.update_one({"id": rec["user_id"]}, {"$set": {"email_verified_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Email verified — you can now sign in."}

@api_router.post("/auth/resend-verification")
@limiter.limit("3/hour")
async def resend_verification(request: Request, req: ResendVerificationRequest):
    # Always returns the same generic message regardless of whether the account exists or
    # is already verified -- otherwise this endpoint could be used to enumerate registered
    # emails (try an address, see if the response differs).
    user = await db.users.find_one({"username": req.email.lower()})
    if user and not user.get("email_verified_at"):
        verify_token = await _create_auth_token(user["id"], "verify_email", ttl_minutes=24 * 60)
        link = f"{FRONTEND_URL}/verify-email?token={verify_token}"
        await _send_email(user["username"], "Verify your email — Hamro G&G Auto OS",
                           f"<p>Click below to verify your email:</p><p><a href='{link}'>{link}</a></p>"
                           f"<p>This link expires in 24 hours.</p>")
    return {"message": "If that email has a pending account, a new verification link has been sent."}

@api_router.post("/auth/forgot-password")
@limiter.limit("3/hour")
async def forgot_password(request: Request, req: ForgotPasswordRequest):
    user = await db.users.find_one({"username": req.email.lower()})  # same enumeration reasoning as resend_verification
    if user:
        reset_token = await _create_auth_token(user["id"], "reset_password", ttl_minutes=60)
        link = f"{FRONTEND_URL}/reset-password?token={reset_token}"
        await _send_email(user["username"], "Reset your password — Hamro G&G Auto OS",
                           f"<p>Click below to choose a new password:</p><p><a href='{link}'>{link}</a></p>"
                           f"<p>This link expires in 1 hour. If you didn't request this, you can ignore this email.</p>")
    return {"message": "If that email exists, a password reset link has been sent."}

@api_router.post("/auth/reset-password")
@limiter.limit("5/hour")
async def reset_password(request: Request, req: ResetPasswordRequest):
    rec = await _consume_auth_token(req.token, "reset_password")
    if not rec:
        raise HTTPException(400, "This reset link is invalid or has expired — request a new one.")
    await db.users.update_one({"id": rec["user_id"]}, {"$set": {"password_hash": await hash_pw_async(req.new_password)}})
    return {"message": "Password reset — you can now sign in with your new password."}

@api_router.get("/auth/me")
async def me(cu: dict = Depends(get_current_user)):
    user = await db.users.find_one({"username": cu["username"]}, {"_id": 0, "password_hash": 0})
    return user

@api_router.post("/auth/change-password")
async def change_password(req: ChangePasswordRequest, cu: dict = Depends(get_current_user)):
    lookup = {"username": cu["username"]}
    user = await db.users.find_one(lookup)
    if not user or not await verify_pw_async(req.current_password, user["password_hash"]):
        raise HTTPException(400, "Current password incorrect")
    await db.users.update_one(lookup, {"$set": {"password_hash": await hash_pw_async(req.new_password)}})
    return {"message": "Password changed successfully"}

@api_router.get("/auth/users")
async def list_users(cu: dict = Depends(get_current_user)):
    users = await db.users.find(company_scope(cu), {"_id": 0, "password_hash": 0}).to_list(100)
    return users

@api_router.post("/auth/register")
async def register(req: RegisterRequest, cu: dict = Depends(get_current_user)):
    if cu.get("role") != "admin":
        raise HTTPException(403, "Only admin can register users")
    email = req.email.lower()
    existing = await db.users.find_one({"username": email})  # globally unique, see /auth/signup
    if existing:
        raise HTTPException(400, "An account with this email already exists")
    now = datetime.now(timezone.utc).isoformat()
    user = {"id": str(uuid.uuid4()), "username": email,
            "password_hash": await hash_pw_async(req.password), "name": req.name, "role": req.role,
            "company_id": cu.get("company_id"),
            # Pre-verified: an already-authenticated admin is vouching for this email by
            # creating the account, unlike anonymous /auth/signup -- see its docstring.
            "email_verified_at": now, "created_at": now}
    await db.users.insert_one(user)
    user.pop("_id", None); user.pop("password_hash", None)
    return user

@api_router.put("/auth/users/{uid}")
async def update_user(uid: str, req: UserUpdate, cu: dict = Depends(admin_only)):
    target = await db.users.find_one({"id": uid, **company_scope(cu)})
    if not target:
        raise HTTPException(404, "Account not found")
    if target.get("role") == "admin" and req.role != "admin":
        remaining_admins = await db.users.count_documents({"role": "admin", "id": {"$ne": uid}, **company_scope(cu)})
        if remaining_admins == 0:
            raise HTTPException(400, "Cannot change role: at least one Admin account must remain")
    update = {"name": req.name, "role": req.role}
    if req.password:
        update["password_hash"] = await hash_pw_async(req.password)
    await db.users.update_one({"id": uid, **company_scope(cu)}, {"$set": update})
    return await db.users.find_one({"id": uid, **company_scope(cu)}, {"_id": 0, "password_hash": 0})

@api_router.delete("/auth/users/{uid}")
async def delete_user(uid: str, cu: dict = Depends(admin_only)):
    if uid == cu.get("user_id"):
        raise HTTPException(400, "You cannot delete your own account")
    target = await db.users.find_one({"id": uid, **company_scope(cu)})
    if not target:
        raise HTTPException(404, "Account not found")
    if target.get("role") == "admin":
        remaining_admins = await db.users.count_documents({"role": "admin", "id": {"$ne": uid}, **company_scope(cu)})
        if remaining_admins == 0:
            raise HTTPException(400, "Cannot delete the only remaining Admin account")
    await db.users.delete_one({"id": uid, **company_scope(cu)})
    return {"message": "Account deleted"}

# ── PLATFORM (platform_owner only) ──────────────────────────────────────
# Read-only visibility across every company that's signed up -- not a management console,
# just what used to require going into phpMyAdmin by hand.
@api_router.get("/platform/companies")
async def platform_list_companies(cu: dict = Depends(platform_owner_only)):
    companies = await db.companies.find({}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    out = []
    for c in companies:
        cid = c["id"]
        # Storage: same reasoning as /admin/storage-usage -- vehicle_photos and
        # legal_documents are the only two collections holding binary file data
        # (base64, not disk); `size` is the decoded byte size recorded at upload time.
        photos = await db.vehicle_photos.find({"company_id": cid}, {"_id": 0, "size": 1}).to_list(100000)
        docs = await db.legal_documents.find({"company_id": cid}, {"_id": 0, "size": 1}).to_list(100000)
        storage_bytes = sum((p.get("size") or 0) for p in photos) + sum((d.get("size") or 0) for d in docs)
        out.append({
            **c,
            "vehicle_count": await db.vehicles.count_documents({"company_id": cid}),
            "customer_count": await db.customers.count_documents({"company_id": cid}),
            "user_count": await db.users.count_documents({"company_id": cid}),
            "storage_bytes": storage_bytes,
        })
    return out

@api_router.get("/platform/companies/{company_id}/users")
async def platform_list_company_users(company_id: str, cu: dict = Depends(platform_owner_only)):
    if not await db.companies.find_one({"id": company_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "Company not found")
    return await db.users.find({"company_id": company_id}, {"_id": 0, "password_hash": 0}).sort("username", 1).to_list(500)

@api_router.delete("/platform/companies/{company_id}")
async def platform_delete_company(company_id: str, cu: dict = Depends(platform_owner_only)):
    """Hard delete -- irreversible. Cascades across every tenant-scoped collection (same
    list _ScopedDB uses for automatic scoping) plus that company's users, not just the
    company row itself, so nothing is left orphaned under a company_id that no longer
    resolves to anything."""
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(404, "Company not found")
    for name in TENANT_COLLECTIONS:
        await getattr(db, name).delete_many({"company_id": company_id})
    await db.users.delete_many({"company_id": company_id})
    await db.companies.delete_one({"id": company_id})
    return {"message": f"{company['name']} and all its data have been permanently deleted"}

# ── VEHICLES ──────────────────────────────────────────────────────────
@api_router.get("/vehicles")
async def get_vehicles(status: Optional[str] = None, brand: Optional[str] = None, cu: dict = Depends(require("vehicles", "view"))):
    q = {}
    if status and status != "all":
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        q["status"] = {"$in": statuses} if len(statuses) > 1 else statuses[0]
    if brand and brand != "all": q["brand"] = brand
    vehicles = await db.vehicles.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    if not vehicles:
        return []
    # Batch-load all expenses in one query (avoids N+1)
    vehicle_ids = [v["id"] for v in vehicles]
    all_exps = await db.expenses.find({"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0}).to_list(10000)
    exps_by_vehicle: dict = {}
    for e in all_exps:
        exps_by_vehicle.setdefault(e["vehicle_id"], []).append(e)
    # Batch-load all job cards too — their cost counts toward a vehicle's expense total
    # the same way a plain expense entry does (see _job_card_cost).
    all_jobs = await db.job_cards.find({"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0}).to_list(10000)
    jobs_by_vehicle: dict = {}
    for j in all_jobs:
        jobs_by_vehicle.setdefault(j["vehicle_id"], []).append(j)
    # Batch-load each vehicle's photo ids in one query (never the base64 `data` field — same
    # projection as /public/vehicles), so the frontend can show a few thumbnails directly on the
    # inventory card (and filter for missing/insufficient photos) without an N+1 fetch of
    # /vehicles/{id}/photos per row, and without inflating this response with image bytes that
    # the browser could otherwise cache via a stable URL (see _public_photo_url).
    photo_docs = await db.vehicle_photos.find(
        {"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0, "id": 1, "vehicle_id": 1, "uploaded_at": 1}
    ).sort("uploaded_at", 1).to_list(10000)
    photos_by_vehicle: dict = {}
    for p in photo_docs:
        photos_by_vehicle.setdefault(p["vehicle_id"], []).append(p)
    # Enrich each vehicle using pre-loaded expenses + job cards
    def enrich_with_expenses(v: dict, exps: list, jobs: list) -> dict:
        v["aging"] = stock_aging(v.get("purchase_date", ""))
        total_exp = sum(e["amount"] for e in exps) + sum(_job_card_cost(j) for j in jobs)
        v["total_expenses"] = total_exp
        v["total_investment"] = v.get("purchase_price", 0) + total_exp + v.get("accessories_cost", 0)
        sp = v.get("selling_price") or 0
        if sp > 0:
            v["expected_profit"] = sp - v["total_investment"]
            v["profit_margin"] = round((v["expected_profit"] / sp) * 100, 2)
            v["low_margin"] = v["profit_margin"] < 8
        else:
            v["expected_profit"] = None; v["profit_margin"] = None; v["low_margin"] = False
        vehicle_photos = photos_by_vehicle.get(v["id"], [])
        v["photo_count"] = len(vehicle_photos)
        v["has_photo"] = v["photo_count"] > 0
        # Card only shows a handful of thumbnails — cap it so the list payload doesn't
        # balloon for vehicles with a large photo library. Real URLs (not embedded base64)
        # so the browser can cache each photo instead of re-downloading it on every load.
        # Relative (not absolute via _public_photo_url) since this endpoint is only ever
        # consumed by our own frontend — a relative path resolves against whichever origin
        # served the page (Vercel's proxy rewrite, or the VPS mirror directly), so the
        # browser never has to open a direct connection to the VPS's sslip.io hostname.
        v["thumb_photos"] = [
            {"id": p["id"], "url": f"/api/public/vehicles/{v['id']}/photos/{p['id']}"}
            for p in vehicle_photos[:3]
        ]
        return v
    role = cu.get("role", "admin")
    return [_hide_financials_for_role(enrich_with_expenses(v, exps_by_vehicle.get(v["id"], []), jobs_by_vehicle.get(v["id"], [])), role) for v in vehicles]

@api_router.post("/vehicles")
async def create_vehicle(vehicle: VehicleCreate, cu: dict = Depends(require("vehicles", "create"))):
    existing = await db.vehicles.find_one(
        {"registration_number": {"$regex": f"^{re.escape(vehicle.registration_number.strip())}$", "$options": "i"}},
        {"_id": 0, "id": 1},
    )
    if existing:
        raise HTTPException(400, f"Registration number '{vehicle.registration_number}' is already in stock")
    v = vehicle.model_dump()
    v["id"] = str(uuid.uuid4())
    v["created_at"] = datetime.now(timezone.utc).isoformat()
    v["updated_at"] = datetime.now(timezone.utc).isoformat()
    v["sold_date"] = None; v["customer_id"] = None
    v["salesperson_id"] = None; v["salesperson_name"] = None; v["discount"] = 0
    v["created_by"] = cu["username"]
    # Audit log
    await db.audit_logs.insert_one({"action": "vehicle_created", "vehicle_id": v["id"],
        "user": cu["username"], "timestamp": datetime.now(timezone.utc).isoformat(),
        "details": f"Added {v['brand']} {v['model']} {v['year']}"})
    await db.vehicles.insert_one(v)
    v.pop("_id", None)
    asyncio.create_task(_notify_storefront())
    return v

# ── Bulk Import (xlsx/csv) ─────────────────────────────────────────────
IMPORT_REQUIRED_FIELDS = ["brand", "model", "year", "purchase_price", "purchase_date", "purchase_source", "registration_number"]

def _import_cell_str(record: dict, key: str) -> Optional[str]:
    val = record.get(key)
    if val is None: return None
    s = str(val).strip()
    return s if s else None

def _import_cell_num(record: dict, key: str):
    val = record.get(key)
    if val is None or str(val).strip() == "": return None
    return float(val)

_IMPORT_DATE_FORMATS = [
    "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
    "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y",
    "%m-%d-%Y", "%m/%d/%Y",
    "%d-%b-%Y", "%d %b %Y", "%d %B %Y",
    "%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%B %d %Y",
    "%d-%m-%y", "%d/%m/%y",
]

def _parse_flexible_date(val) -> str:
    """Best-effort parse of a free-format date value into YYYY-MM-DD. Falls back to the raw trimmed string if unrecognized."""
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    for fmt in _IMPORT_DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s

def _parse_vehicle_import_rows(content: bytes, filename: str, created_by: str):
    filename = (filename or "").lower()
    rows: List[list] = []

    if filename.endswith(".csv"):
        import csv, io
        text = content.decode("utf-8-sig", errors="ignore")
        rows = [row for row in csv.reader(io.StringIO(text))]
    elif filename.endswith(".xlsx"):
        import openpyxl, io
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            raise HTTPException(400, "Could not read file. Make sure it's a valid .xlsx sheet.")
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        raise HTTPException(400, "Unsupported file type. Use .xlsx or .csv")

    if len(rows) < 2:
        raise HTTPException(400, "The sheet needs a header row plus at least one data row.")

    header_row = rows[0]
    headers = []
    for i, h in enumerate(header_row):
        h = str(h).strip() if h is not None else ""
        headers.append(h.lower().replace(" ", "_") if h else f"column_{i}")

    docs = []
    errors = []
    row_results = []
    total_data_rows = 0
    for sheet_row_num, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        total_data_rows += 1
        row = list(row) + [None] * (len(headers) - len(row))
        record = dict(zip(headers, row))
        summary = f"{_import_cell_str(record, 'brand') or ''} {_import_cell_str(record, 'model') or ''}".strip() or "(unnamed)"
        missing = [f for f in IMPORT_REQUIRED_FIELDS if _import_cell_str(record, f) is None]
        if missing:
            reason = f"Missing required field(s): {', '.join(missing)}"
            errors.append({"row": sheet_row_num, "reason": reason})
            row_results.append({"row": sheet_row_num, "vehicle": summary, "status": "error", "reason": reason})
            continue
        try:
            purchase_date = _parse_flexible_date(record.get("purchase_date"))
            selling_price = _import_cell_num(record, "selling_price")
            min_selling_price = _import_cell_num(record, "minimum_selling_price")
            km_run = _import_cell_num(record, "kilometer_run")
            status_val = (_import_cell_str(record, "status") or "available").lower()
            if status_val == "hidden":
                status_val = "unlisted"  # legacy alias, pre-rename data
            if status_val not in ("available", "reserved", "sold", "unlisted", "scrap", "in_repair"):
                status_val = "available"
            doc = {
                "id": str(uuid.uuid4()),
                "brand": _import_cell_str(record, "brand"),
                "model": _import_cell_str(record, "model"),
                "variant": _import_cell_str(record, "variant"),
                "year": int(float(record.get("year"))),
                "engine_cc": int(float(record.get("engine_cc"))) if _import_cell_str(record, "engine_cc") else 0,
                "fuel_type": _import_cell_str(record, "fuel_type") or "Petrol",
                "ownership_number": int(float(record.get("ownership_number"))) if _import_cell_str(record, "ownership_number") else 1,
                "chassis_number": _import_cell_str(record, "chassis_number"),
                "engine_number": _import_cell_str(record, "engine_number"),
                "kilometer_run": int(km_run) if km_run is not None else None,
                "condition": _import_cell_str(record, "condition") or "Good",
                "condition_rating": int(float(record.get("condition_rating"))) if _import_cell_str(record, "condition_rating") else 7,
                "color": _import_cell_str(record, "color"),
                "registration_number": _import_cell_str(record, "registration_number"),
                "purchase_price": float(record.get("purchase_price")),
                "accessories_cost": _import_cell_num(record, "accessories_cost") or 0,
                "purchase_date": purchase_date,
                "purchase_source": _import_cell_str(record, "purchase_source"),
                "vendor_id": None,
                "purchase_from": _import_cell_str(record, "purchase_from"),
                "selling_price": selling_price,
                "minimum_selling_price": min_selling_price,
                "notes": _import_cell_str(record, "notes"),
                "status": status_val,
                "bluebook_status": "pending", "insurance_status": "pending",
                "tax_clearance_status": "pending", "transfer_status": "pending",
                "sold_date": None, "customer_id": None,
                "salesperson_id": None, "salesperson_name": None, "discount": 0,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "created_by": created_by,
            }
            docs.append(doc)
            row_results.append({"row": sheet_row_num, "vehicle": summary, "status": "ok", "reason": None})
        except (ValueError, TypeError) as e:
            reason = f"Invalid value: {e}"
            errors.append({"row": sheet_row_num, "reason": reason})
            row_results.append({"row": sheet_row_num, "vehicle": summary, "status": "error", "reason": reason})

    return docs, errors, row_results, total_data_rows


@api_router.post("/vehicles/import")
async def import_vehicles(file: UploadFile = File(...), confirm: bool = False, cu: dict = Depends(admin_only)):
    content = await file.read()
    docs, errors, row_results, total_data_rows = _parse_vehicle_import_rows(content, file.filename, cu["username"])

    # Duplicate registration_number check — both within the sheet itself and against
    # vehicles already in stock (case-insensitive, since dealers key inventory off this number).
    ok_indices = [i for i, r in enumerate(row_results) if r["status"] == "ok"]
    existing_regs_lower = {d["registration_number"].strip().lower() for d in await db.vehicles.find({}, {"_id": 0, "registration_number": 1}).to_list(100000) if d.get("registration_number")}
    seen_in_sheet = {}
    kept_docs = []
    for doc, ri in zip(docs, ok_indices):
        row = row_results[ri]
        reg = (doc.get("registration_number") or "").strip()
        reg_lower = reg.lower()
        if reg_lower in existing_regs_lower:
            reason = f"Registration number '{reg}' is already in stock"
        elif reg_lower in seen_in_sheet:
            reason = f"Duplicate registration number '{reg}' (also on row {seen_in_sheet[reg_lower]})"
        else:
            reason = None
        if reason:
            errors.append({"row": row["row"], "reason": reason})
            row_results[ri] = {**row, "status": "error", "reason": reason}
        else:
            seen_in_sheet[reg_lower] = row["row"]
            kept_docs.append(doc)
    docs = kept_docs

    if errors:
        return {
            "committed": False,
            "all_success": False,
            "inserted": 0,
            "skipped": len(errors),
            "total_rows": total_data_rows,
            "rows": row_results,
            "errors": errors[:200],
            "message": f"{len(errors)} of {total_data_rows} row(s) failed validation. Fix them and re-upload — nothing was imported.",
        }

    if not confirm:
        return {
            "committed": False,
            "all_success": True,
            "inserted": 0,
            "skipped": 0,
            "total_rows": total_data_rows,
            "rows": row_results,
            "errors": [],
            "message": f"All {total_data_rows} row(s) validated successfully. Confirm to import.",
        }

    # A row imported with status=Sold bypasses both the Sales form and the Inventory
    # quick-status-edit path, so it needs the same auto-create-a-sale-record treatment
    # `update_vehicle` already does for a direct Sold edit (see there for the pattern).
    sold_docs = [d for d in docs if d.get("status") == "sold"]
    for d in sold_docs:
        sale_price = d.get("selling_price") or d.get("purchase_price", 0)
        d["sold_date"] = d.get("purchase_date") or datetime.now(timezone.utc).date().isoformat()
        if not d.get("selling_price"):
            d["selling_price"] = sale_price

    inserted = 0
    for i in range(0, len(docs), 500):
        batch = docs[i:i + 500]
        result = await db.vehicles.insert_many(batch, ordered=False)
        inserted += len(result.inserted_ids)

    if sold_docs:
        now_iso = datetime.now(timezone.utc).isoformat()
        await db.sales.insert_many([{
            "id": str(uuid.uuid4()),
            "vehicle_id": d["id"],
            "customer_id": None,
            "sale_price": d["selling_price"],
            "extra_expenses": [],
            "expenses_total": 0,
            "total_amount": d["selling_price"],
            "payment_method": "Due",
            "paid_cash": 0,
            "paid_bank": 0,
            "due_amount": d["selling_price"],
            "due_date": None,
            "payment_status": "Unpaid",
            "sale_date": d["sold_date"],
            "notes": "Auto-created: vehicle imported with status=Sold — review customer/payment details.",
            "created_by": cu.get("username"),
            "created_at": now_iso,
        } for d in sold_docs])

    await db.audit_logs.insert_one({"action": "vehicles_bulk_imported",
        "user": cu["username"], "timestamp": datetime.now(timezone.utc).isoformat(),
        "details": f"Imported {inserted} vehicles via bulk import from {file.filename}" + (f", {len(sold_docs)} auto-recorded as sales" if sold_docs else "")})

    return {
        "committed": True,
        "all_success": True,
        "inserted": inserted,
        "skipped": 0,
        "total_rows": total_data_rows,
        "rows": row_results,
        "errors": [],
        "message": f"Imported {inserted} vehicle{'s' if inserted != 1 else ''} successfully." + (f" {len(sold_docs)} marked Sold were added to the Sales tab too." if sold_docs else ""),
    }

@api_router.get("/vehicles/{vid}")
async def get_vehicle(vid: str, cu: dict = Depends(require("vehicles", "view"))):
    v = await db.vehicles.find_one({"id": vid}, {"_id": 0})
    if not v: raise HTTPException(404, "Vehicle not found")
    v = await enrich_vehicle(v)
    v["expenses"] = await db.expenses.find({"vehicle_id": vid}, {"_id": 0}).to_list(200)
    v["job_cards"] = await db.job_cards.find({"vehicle_id": vid}, {"_id": 0}).to_list(100)
    return _hide_financials_for_role(v, cu.get("role", "admin"))

@api_router.put("/vehicles/{vid}")
async def update_vehicle(vid: str, vehicle: VehicleUpdate, cu: dict = Depends(require("vehicles", "edit"))):
    existing = await db.vehicles.find_one({"id": vid}, {"_id": 0})
    if not existing: raise HTTPException(404, "Vehicle not found")
    upd = {k: val for k, val in vehicle.model_dump().items() if val is not None}
    if upd.get("registration_number") and upd["registration_number"].strip().lower() != (existing.get("registration_number") or "").strip().lower():
        dup = await db.vehicles.find_one(
            {"id": {"$ne": vid}, "registration_number": {"$regex": f"^{re.escape(upd['registration_number'].strip())}$", "$options": "i"}},
            {"_id": 0, "id": 1},
        )
        if dup:
            raise HTTPException(400, f"Registration number '{upd['registration_number']}' is already in stock")
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    if upd.get("status") == "sold" and "sold_date" not in upd:
        # Preserve the original sale date if this vehicle was sold before — otherwise
        # re-marking Sold (e.g. after a warranty repair stint) resets the sale date
        # and the warranty clock to today.
        upd["sold_date"] = existing.get("sold_date") or datetime.now(timezone.utc).date().isoformat()

    # Marking a vehicle Sold directly (Inventory/Edit/quick-status) bypasses the Sales form —
    # auto-create the matching sale record so it still shows up in the Sales tab.
    became_sold = upd.get("status") == "sold" and existing.get("status") != "sold"
    sale_price = upd.get("selling_price", existing.get("selling_price")) or existing.get("purchase_price", 0)
    if became_sold and not existing.get("selling_price") and "selling_price" not in upd:
        upd["selling_price"] = sale_price

    # Unlisted vehicles are hidden from the storefront specifically because they had
    # no price yet ("Move ... With No Price To Unlisted" in Inventory) — the moment a
    # price is added, it's ready to sell, so relist it automatically instead of making
    # someone remember to flip the status separately. The edit form always re-submits
    # the current status (even when untouched), so this checks the resulting status
    # rather than requiring it be absent from the payload — an explicit status change
    # to something other than "unlisted" in the same request still wins.
    if (upd.get("status", existing.get("status")) == "unlisted"
            and not existing.get("selling_price") and upd.get("selling_price")):
        upd["status"] = "available"

    r = await db.vehicles.update_one({"id": vid}, {"$set": upd})
    if r.matched_count == 0: raise HTTPException(404, "Vehicle not found")

    if became_sold and not await db.sales.find_one({"vehicle_id": vid, "returned": {"$ne": True}}):
        sale_date = upd.get("sold_date", existing.get("sold_date")) or datetime.now(timezone.utc).date().isoformat()
        await db.sales.insert_one({
            "id": str(uuid.uuid4()),
            "vehicle_id": vid,
            "customer_id": upd.get("customer_id", existing.get("customer_id")),
            "sale_price": sale_price,
            "extra_expenses": [],
            "expenses_total": 0,
            "total_amount": sale_price,
            "payment_method": "Due",
            "paid_cash": 0,
            "paid_bank": 0,
            "due_amount": sale_price,
            "due_date": None,
            "payment_status": "Unpaid",
            "sale_date": sale_date,
            "notes": "Auto-created: vehicle marked Sold directly from Inventory",
            "created_by": cu.get("username"),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    await db.audit_logs.insert_one({"action": "vehicle_updated", "vehicle_id": vid,
        "user": cu["username"], "timestamp": datetime.now(timezone.utc).isoformat(),
        "details": f"Updated fields: {list(upd.keys())}"})
    asyncio.create_task(_notify_storefront())
    return await db.vehicles.find_one({"id": vid}, {"_id": 0})

VEHICLE_STATUSES = {"available", "reserved", "sold", "unlisted", "scrap", "in_repair"}

# Status-only update, separate from the full PUT /vehicles/{vid} above: this is what lets
# parts_supervisor (edit_status only, not full edit — see ROLE_PERMISSIONS) move a vehicle
# through the repair pipeline without exposing the rest of the vehicle record to editing.
@api_router.patch("/vehicles/{vid}/status")
async def update_vehicle_status(vid: str, body: VehicleStatusUpdate, cu: dict = Depends(require_any("vehicles", {"edit", "edit_status"}))):
    if body.status not in VEHICLE_STATUSES:
        raise HTTPException(400, f"Invalid status '{body.status}'")

    role = cu.get("role", "admin")
    has_full_edit = role == "admin" or "edit" in ROLE_PERMISSIONS.get(role, {}).get("vehicles", set())

    existing = await db.vehicles.find_one({"id": vid}, {"_id": 0})
    if not existing: raise HTTPException(404, "Vehicle not found")

    if not has_full_edit:
        # Scoped-access roles (e.g. parts_supervisor) may only move a vehicle between the
        # statuses their permission covers — never touch Sold/Reserved/Unlisted vehicles.
        if existing.get("status") not in PARTS_ALLOWED_STATUSES or body.status not in PARTS_ALLOWED_STATUSES:
            raise HTTPException(403, "You do not have permission to set this vehicle status")

    upd = {"status": body.status, "updated_at": datetime.now(timezone.utc).isoformat()}
    if body.status == "sold":
        # Preserve the original sale date if this vehicle was sold before (e.g. it was
        # flipped to In Repair for a warranty return and is now being flipped back) —
        # otherwise this would reset the sale date and the warranty clock to today.
        upd["sold_date"] = existing.get("sold_date") or datetime.now(timezone.utc).date().isoformat()

    became_sold = body.status == "sold" and existing.get("status") != "sold"
    sale_price = existing.get("selling_price") or existing.get("purchase_price", 0)
    if became_sold and not existing.get("selling_price"):
        upd["selling_price"] = sale_price

    r = await db.vehicles.update_one({"id": vid}, {"$set": upd})
    if r.matched_count == 0: raise HTTPException(404, "Vehicle not found")

    if became_sold and not await db.sales.find_one({"vehicle_id": vid, "returned": {"$ne": True}}):
        await db.sales.insert_one({
            "id": str(uuid.uuid4()),
            "vehicle_id": vid,
            "customer_id": existing.get("customer_id"),
            "sale_price": sale_price,
            "extra_expenses": [],
            "expenses_total": 0,
            "total_amount": sale_price,
            "payment_method": "Due",
            "paid_cash": 0,
            "paid_bank": 0,
            "due_amount": sale_price,
            "due_date": None,
            "payment_status": "Unpaid",
            "sale_date": upd["sold_date"],
            "notes": "Auto-created: vehicle marked Sold directly from Inventory",
            "created_by": cu.get("username"),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    await db.audit_logs.insert_one({"action": "vehicle_status_updated", "vehicle_id": vid,
        "user": cu["username"], "timestamp": datetime.now(timezone.utc).isoformat(),
        "details": f"Status changed from {existing.get('status')} to {body.status}"})
    asyncio.create_task(_notify_storefront())
    v = await db.vehicles.find_one({"id": vid}, {"_id": 0})
    return _hide_financials_for_role(v, role)

# Lets Vehicle Detail / Sold Stock show the refund-preview numbers before opening the return
# modal, without the caller having to know which sale record is currently active for this
# vehicle — there's no time limit on this (unlike the warranty window), a vehicle sold years
# ago can still be returned, so this just looks for whichever sale hasn't been returned yet.
@api_router.get("/vehicles/{vid}/active-sale")
async def get_active_sale(vid: str, cu: dict = Depends(admin_only)):
    sale = await db.sales.find_one({"vehicle_id": vid, "returned": {"$ne": True}}, {"_id": 0}, sort=[("created_at", -1)])
    if not sale:
        raise HTTPException(404, "No active sale found for this vehicle")
    return sale

# Recondition-house return: unlike the plain status dropdown above (which the /sales/reconcile
# ribbon is specifically meant to catch if someone bypasses proper channels), this is the one
# sanctioned path for taking a sold vehicle back. The shop assesses the vehicle's condition and
# refunds only part of what the customer paid, keeping the rest — so the sale record is
# tagged "returned" (not deleted) with that split preserved, and the vehicle re-enters stock at
# whatever status the assessment calls for. There is no time limit — a vehicle can come back
# for return long after its warranty window has closed.
@api_router.post("/vehicles/{vid}/return")
async def return_vehicle(vid: str, body: VehicleReturnRequest, cu: dict = Depends(admin_only)):
    existing = await db.vehicles.find_one({"id": vid}, {"_id": 0})
    if not existing: raise HTTPException(404, "Vehicle not found")
    if existing.get("status") != "sold":
        raise HTTPException(400, f"Vehicle is not currently marked Sold (status: {existing.get('status')})")

    sale = await db.sales.find_one({"vehicle_id": vid, "returned": {"$ne": True}}, sort=[("created_at", -1)])
    if not sale:
        raise HTTPException(404, "No active sale found for this vehicle")

    total_amount = sale.get("total_amount", 0)
    if not (0 <= body.refund_amount <= total_amount):
        raise HTTPException(400, f"Refund amount must be between 0 and the sale total ({total_amount})")
    if body.new_status not in VEHICLE_STATUSES or body.new_status == "sold":
        raise HTTPException(400, f"Invalid return status '{body.new_status}'")

    refund_amount = round(body.refund_amount, 2)
    retained_amount = round(total_amount - refund_amount, 2)

    await db.sales.update_one({"id": sale["id"]}, {"$set": {
        "returned": True,
        "returned_at": datetime.now(timezone.utc).isoformat(),
        "returned_status": body.new_status,
        "refund_amount": refund_amount,
        "retained_amount": retained_amount,
        "return_notes": body.notes,
        "due_amount": 0,
    }})

    await db.vehicles.update_one({"id": vid}, {"$set": {
        "status": body.new_status,
        "sold_date": None,
        "customer_id": None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})

    await db.audit_logs.insert_one({"action": "vehicle_returned", "vehicle_id": vid,
        "user": cu["username"], "timestamp": datetime.now(timezone.utc).isoformat(),
        "details": f"Sale {sale['id']} returned with {refund_amount} refunded, vehicle set to {body.new_status}"})

    v = await db.vehicles.find_one({"id": vid}, {"_id": 0})
    return {"vehicle": v, "refund_amount": refund_amount, "retained_amount": retained_amount}

@api_router.delete("/vehicles/{vid}")
async def delete_vehicle(vid: str, cu: dict = Depends(admin_only)):
    r = await db.vehicles.delete_one({"id": vid})
    if r.deleted_count == 0: raise HTTPException(404, "Vehicle not found")
    await db.expenses.delete_many({"vehicle_id": vid})
    await db.job_cards.delete_many({"vehicle_id": vid})
    # Otherwise a sale record survives with no vehicle to resolve — it still counts in the
    # Sales tab total but can never show up in Sold Stock, silently desyncing the two counts.
    await db.sales.delete_many({"vehicle_id": vid})
    asyncio.create_task(_notify_storefront())
    return {"message": "Deleted"}

@api_router.get("/vehicles/{vid}/qr-data")
async def get_vehicle_qr(vid: str):
    v = await db.vehicles.find_one({"id": vid}, {"_id": 0})
    if not v: raise HTTPException(404, "Not found")
    return {"id": v["id"], "brand": v.get("brand"), "model": v.get("model"),
            "variant": v.get("variant"), "year": v.get("year"),
            "engine_cc": v.get("engine_cc"), "fuel_type": v.get("fuel_type"),
            "color": v.get("color"), "ownership_number": v.get("ownership_number"),
            "kilometer_run": v.get("kilometer_run"), "condition": v.get("condition"),
            "selling_price": v.get("selling_price"), "minimum_selling_price": v.get("minimum_selling_price"),
            "registration_number": v.get("registration_number"), "status": v.get("status"),
            "contact": "Hamro G&G Auto Enterprises"}

# ── EXPENSES ──────────────────────────────────────────────────────────
@api_router.get("/vehicles/{vid}/expenses")
async def get_expenses(vid: str, cu: dict = Depends(require("expenses", "view"))):
    return await db.expenses.find({"vehicle_id": vid}, {"_id": 0}).to_list(200)

@api_router.post("/expenses")
async def create_expense(exp: ExpenseCreate, cu: dict = Depends(require("expenses", "create"))):
    if not await db.vehicles.find_one({"id": exp.vehicle_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "Vehicle not found")
    e = exp.model_dump()
    e["id"] = str(uuid.uuid4())
    e["date"] = e.get("date") or datetime.now(timezone.utc).date().isoformat()
    e["added_by"] = cu["username"]
    e["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.expenses.insert_one(e)
    e.pop("_id", None)
    return e

@api_router.delete("/expenses/{eid}")
async def delete_expense(eid: str, cu: dict = Depends(require("expenses", "delete"))):
    r = await db.expenses.delete_one({"id": eid})
    if r.deleted_count == 0: raise HTTPException(404, "Expense not found")
    return {"message": "Deleted"}

# ── JOB CARDS ─────────────────────────────────────────────────────────
@api_router.get("/jobs")
async def get_jobs(status: Optional[str] = None, vehicle_id: Optional[str] = None, vehicle_status: Optional[str] = None, cu: dict = Depends(require("jobs", "view"))):
    q = {}
    if status and status != "all": q["status"] = status
    if vehicle_id: q["vehicle_id"] = vehicle_id
    jobs = await db.job_cards.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    if not jobs:
        return []
    # Batch-load current vehicle status in one query (avoids N+1) so the frontend can show/filter
    # by the vehicle's live pipeline stage (available/sold/in_repair/...) alongside the job's own status.
    ids = list({j["vehicle_id"] for j in jobs if j.get("vehicle_id")})
    status_by_vehicle: dict = {}
    if ids:
        vs = await db.vehicles.find({"id": {"$in": ids}}, {"_id": 0, "id": 1, "status": 1}).to_list(10000)
        status_by_vehicle = {v["id"]: v.get("status") for v in vs}
    for j in jobs:
        j["vehicle_status"] = status_by_vehicle.get(j.get("vehicle_id"))
    if vehicle_status and vehicle_status != "all":
        jobs = [j for j in jobs if j.get("vehicle_status") == vehicle_status]
    return jobs

async def _job_part_current_stock(part_id: str, component_name: Optional[str]):
    """Looks up the current usable stock for a job-card part row — either a plain part's own
    `quantity`, or (when component_name is set) one named item inside a Set's checklist, which
    is tracked independently of the set's own whole-set quantity. Returns (current, label) or
    (None, None) if the part/component no longer exists."""
    if component_name:
        pp = await db.spare_parts.find_one({"id": part_id}, {"_id": 0, "name": 1, "set_components": 1})
        if not pp: return None, None
        comp = next((c for c in pp.get("set_components", []) if c.get("name") == component_name), None)
        if not comp: return None, None
        return comp.get("stock", 0), f"{pp.get('name')} — {component_name}"
    pp = await db.spare_parts.find_one({"id": part_id}, {"_id": 0, "name": 1, "quantity": 1})
    if not pp: return None, None
    return pp.get("quantity", 0), pp.get("name")

async def _job_part_set_stock(part_id: str, component_name: Optional[str], new_value: int):
    if component_name:
        await db.spare_parts.update_one({"id": part_id, "set_components.name": component_name}, {"$set": {"set_components.$.stock": new_value}})
    else:
        await db.spare_parts.update_one({"id": part_id}, {"$set": {"quantity": new_value}})

@api_router.post("/jobs")
async def create_job(job: JobCardCreate, cu: dict = Depends(require("jobs", "create"))):
    count = await db.job_cards.count_documents({})
    jc = job.model_dump()
    jc["id"] = str(uuid.uuid4())
    jc["job_number"] = f"JC-{datetime.now(timezone.utc).year}-{str(count + 1).zfill(3)}"
    jc["status"] = "pending"; jc["actual_cost"] = None
    jc["created_at"] = datetime.now(timezone.utc).isoformat()
    jc["completed_at"] = None; jc["created_by"] = cu["username"]
    if job.is_external:
        jc["vehicle_id"] = None
        if not job.vehicle_brand or not job.vehicle_model or not job.registration_number:
            raise HTTPException(400, "Brand, model and registration number are required for external vehicles")
    else:
        if not job.vehicle_id: raise HTTPException(400, "vehicle_id is required")
        v = await db.vehicles.find_one({"id": job.vehicle_id}, {"_id": 0})
        if not v: raise HTTPException(404, "Vehicle not found")
        if v.get("status") == "sold":
            if not _within_warranty(v):
                raise HTTPException(400, "This vehicle's 6-month warranty period has expired")
            jc["is_warranty"] = True
        elif v.get("status") != "in_repair":
            raise HTTPException(400, "Job cards can only be created for vehicles in the Repair stage, or sold vehicles still under warranty")
        jc["vehicle_brand"] = v.get("brand"); jc["vehicle_model"] = v.get("model")
        jc["vehicle_year"] = v.get("year"); jc["registration_number"] = v.get("registration_number")
    await db.job_cards.insert_one(jc)
    jc.pop("_id", None)
    # Deduct parts (or, for a Set, one specific component within it) from spare parts inventory and log transactions
    for part in jc.get("parts", []):
        part_id = part.get("part_id")
        component_name = part.get("component_name")
        qty = int(part.get("quantity", 0))
        if part_id and qty > 0:
            current, part_label = await _job_part_current_stock(part_id, component_name)
            if current is not None:
                await _job_part_set_stock(part_id, component_name, max(0, current - qty))
                txn = {
                    "id": str(uuid.uuid4()), "part_id": part_id, "part_name": part_label,
                    "type": "out", "quantity": qty, "reason": "Used in Job Card",
                    "date": datetime.now(timezone.utc).isoformat()[:10],
                    "job_id": jc["id"], "notes": f"Job {jc['job_number']}",
                    "created_by": cu.get("username"), "created_at": datetime.now(timezone.utc).isoformat(),
                }
                await db.part_transactions.insert_one(txn)
    return jc

@api_router.put("/jobs/{jid}")
async def update_job(jid: str, job: JobCardUpdate, cu: dict = Depends(require("jobs", "edit"))):
    existing = await db.job_cards.find_one({"id": jid}, {"_id": 0})
    if not existing: raise HTTPException(404, "Job not found")
    upd = {k: v for k, v in job.model_dump().items() if v is not None}
    if upd.get("status") == "completed":
        upd["completed_at"] = datetime.now(timezone.utc).isoformat()
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()

    if "parts" in upd:
        old_parts = existing.get("parts", []) or []
        new_parts = upd["parts"] or []
        old_sig = {(p.get("part_id"), p.get("part_name"), int(p.get("quantity", 0))) for p in old_parts}
        new_sig = {(p.get("part_id"), p.get("part_name"), int(p.get("quantity", 0))) for p in new_parts}
        # A completed job that gets its parts list changed (e.g. another item added)
        # wasn't actually finished — reopen it, unless the caller already set a status
        # explicitly in this same request. Keeps every other field untouched.
        if old_sig != new_sig and existing.get("status") == "completed" and "status" not in upd:
            upd["status"] = "in_progress"
            upd["completed_at"] = None

        # Keyed by (part_id, component_name) — a plain part's key has component_name=None, a
        # specific item used out of a Set is keyed separately per item so two different items
        # from the same set (or a component alongside the same set's other components) each
        # diff independently instead of colliding on a shared part_id.
        old_qtys, new_qtys = {}, {}
        for p in old_parts:
            if p.get("part_id"):
                key = (p["part_id"], p.get("component_name"))
                old_qtys[key] = old_qtys.get(key, 0) + int(p.get("quantity", 0))
        for p in new_parts:
            if p.get("part_id"):
                key = (p["part_id"], p.get("component_name"))
                new_qtys[key] = new_qtys.get(key, 0) + int(p.get("quantity", 0))
        diffs = {}
        stock_info = {}
        for key in set(old_qtys) | set(new_qtys):
            diff = new_qtys.get(key, 0) - old_qtys.get(key, 0)
            if diff == 0: continue
            part_id, component_name = key
            current, part_label = await _job_part_current_stock(part_id, component_name)
            if current is None: continue
            if diff > 0 and current < diff:
                raise HTTPException(400, f"Not enough stock for {part_label}: only {current} left")
            diffs[key] = diff
            stock_info[key] = (current, part_label)
        for key, diff in diffs.items():
            part_id, component_name = key
            current, part_label = stock_info[key]
            await _job_part_set_stock(part_id, component_name, max(0, current - diff))
            txn = {
                "id": str(uuid.uuid4()), "part_id": part_id, "part_name": part_label,
                "type": "out" if diff > 0 else "in", "quantity": abs(diff),
                "reason": "Used in Job Card" if diff > 0 else "Removed from Job Card",
                "date": datetime.now(timezone.utc).date().isoformat(),
                "job_id": jid, "notes": f"Job {existing.get('job_number')} updated",
                "created_by": cu.get("username"), "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.part_transactions.insert_one(txn)

    r = await db.job_cards.update_one({"id": jid}, {"$set": upd})
    if r.matched_count == 0: raise HTTPException(404, "Job not found")
    return await db.job_cards.find_one({"id": jid}, {"_id": 0})

@api_router.delete("/jobs/{jid}")
async def delete_job(jid: str, cu: dict = Depends(require("jobs", "delete"))):
    r = await db.job_cards.delete_one({"id": jid})
    if r.deleted_count == 0: raise HTTPException(404, "Job not found")
    return {"message": "Deleted"}

# ── CUSTOMERS ─────────────────────────────────────────────────────────
@api_router.get("/customers")
async def get_customers(cu: dict = Depends(require("customers", "view"))):
    customers = await db.customers.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    customer_ids = [c["id"] for c in customers]
    sales_by_customer: dict = {}
    if customer_ids:
        all_sales = await db.sales.find(
            {"customer_id": {"$in": customer_ids}}, {"_id": 0, "customer_id": 1, "due_amount": 1, "due_date": 1, "returned": 1}
        ).to_list(20000)
        for s in all_sales:
            sales_by_customer.setdefault(s["customer_id"], []).append(s)
    today_iso2 = datetime.now(timezone.utc).date().isoformat()
    for c in customers:
        cust_sales = sales_by_customer.get(c["id"], [])
        active_cust_sales = [cs for cs in cust_sales if not cs.get("returned")]
        c["purchase_count"] = len(active_cust_sales); c["is_repeat_customer"] = len(active_cust_sales) > 1
        c["total_due"] = round(sum(cs.get("due_amount", 0) for cs in cust_sales), 2)
        c["has_overdue"] = any((cs.get("due_amount", 0) > 0 and cs.get("due_date") and cs.get("due_date") < today_iso2) for cs in cust_sales)
    return customers

@api_router.post("/customers")
async def create_customer(cust: CustomerCreate, cu: dict = Depends(require("customers", "create"))):
    c = cust.model_dump()
    c["id"] = str(uuid.uuid4())
    c["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.customers.insert_one(c)
    c.pop("_id", None)
    return c

@api_router.get("/customers/{cid}")
async def get_customer(cid: str, cu: dict = Depends(require("customers", "view"))):
    c = await db.customers.find_one({"id": cid}, {"_id": 0})
    if not c: raise HTTPException(404, "Not found")
    sales = await db.sales.find({"customer_id": cid}, {"_id": 0}).sort("sale_date", -1).to_list(200)
    vehicle_ids = list({s["vehicle_id"] for s in sales if s.get("vehicle_id")})
    vehicles_by_id = {}
    if vehicle_ids:
        vs = await db.vehicles.find({"id": {"$in": vehicle_ids}}, {"_id": 0, "id": 1, "brand": 1, "model": 1, "year": 1, "registration_number": 1}).to_list(len(vehicle_ids))
        vehicles_by_id = {v["id"]: v for v in vs}
    for s in sales:
        v = vehicles_by_id.get(s.get("vehicle_id"))
        s["vehicle_info"] = f"{v['brand']} {v['model']} {v.get('year','')}" + (f" ({v['registration_number']})" if v.get("registration_number") else "") if v else "Vehicle removed"
    c["sales"] = sales
    active_sales = [s for s in sales if not s.get("returned")]
    c["purchase_count"] = len(active_sales)
    c["is_repeat_customer"] = len(active_sales) > 1
    return c

@api_router.put("/customers/{cid}")
async def update_customer(cid: str, cust: CustomerCreate, cu: dict = Depends(require("customers", "edit"))):
    r = await db.customers.update_one({"id": cid}, {"$set": cust.model_dump()})
    if r.matched_count == 0: raise HTTPException(404, "Not found")
    return await db.customers.find_one({"id": cid}, {"_id": 0})

@api_router.delete("/customers/{cid}")
async def delete_customer(cid: str, cu: dict = Depends(require("customers", "delete"))):
    r = await db.customers.delete_one({"id": cid})
    if r.deleted_count == 0: raise HTTPException(404, "Not found")
    return {"message": "Deleted"}

# ── SALES ─────────────────────────────────────────────────────────────
@api_router.get("/sales")
async def get_sales(start_date: Optional[str] = None, end_date: Optional[str] = None, returned: Optional[bool] = None, cu: dict = Depends(require("sales", "view"))):
    """start_date/end_date (both required together) filter by sale_date server-side —
    lets callers like the Dashboard's sales ribbon fetch just "today" or "this week"
    instead of the full history. `returned` lets callers like the Sold Stock page's
    Returned tab fetch just the sales that were later returned. Vehicle/customer lookups
    are batched with $in instead of one query per sale, which is what made this endpoint
    slow to begin with."""
    query = {}
    if start_date and end_date:
        query["sale_date"] = {"$gte": start_date, "$lte": end_date}
    if returned is not None:
        query["returned"] = True if returned else {"$ne": True}
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)

    # Clean up any sale a now-removed code path already added job-card cost to — see
    # _strip_job_card_extra_expenses. No-ops (no extra query/write) for anything unaffected.
    to_clean = [s for s in sales if any(str(e.get("name", "")).startswith("Job Card ") for e in s.get("extra_expenses", []))]
    if to_clean:
        await asyncio.gather(*(_strip_job_card_extra_expenses(s) for s in to_clean))
        sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)

    vehicle_ids = list({s["vehicle_id"] for s in sales if s.get("vehicle_id")})
    customer_ids = list({s["customer_id"] for s in sales if s.get("customer_id")})
    vehicles_by_id, customers_by_id = {}, {}
    if vehicle_ids:
        vs = await db.vehicles.find({"id": {"$in": vehicle_ids}}, {"_id": 0, "id": 1, "brand": 1, "model": 1, "year": 1, "registration_number": 1}).to_list(len(vehicle_ids))
        vehicles_by_id = {v["id"]: v for v in vs}
    if customer_ids:
        cs = await db.customers.find({"id": {"$in": customer_ids}}, {"_id": 0, "id": 1, "name": 1, "contact_number": 1}).to_list(len(customer_ids))
        customers_by_id = {c["id"]: c for c in cs}
    job_cost_by_vehicle: dict = {}
    if vehicle_ids:
        all_jobs = await db.job_cards.find({"vehicle_id": {"$in": vehicle_ids}, "is_warranty": {"$ne": True}}, {"_id": 0}).to_list(20000)
        for j in all_jobs:
            job_cost_by_vehicle[j["vehicle_id"]] = job_cost_by_vehicle.get(j["vehicle_id"], 0) + _job_card_cost(j)

    for s in sales:
        v = vehicles_by_id.get(s.get("vehicle_id"))
        if v:
            s["vehicle_info"] = f"{v['brand']} {v['model']} {v.get('year','')}" + (f" ({v['registration_number']})" if v.get("registration_number") else "")
            s["vehicle_brand"] = v.get("brand"); s["vehicle_model"] = v.get("model")
            s["vehicle_year"] = v.get("year"); s["registration_number"] = v.get("registration_number")
        c = customers_by_id.get(s.get("customer_id"))
        s["customer_name"] = c["name"] if c else "Walk-in Customer"
        s["customer_contact"] = c.get("contact_number") if c else None
        s["job_card_cost"] = job_cost_by_vehicle.get(s.get("vehicle_id"), 0)
    return sales

@api_router.post("/sales")
async def create_sale(sale: SaleCreate, cu: dict = Depends(require("sales", "create"))):
    v = await db.vehicles.find_one({"id": sale.vehicle_id}, {"_id": 0})
    if not v: raise HTTPException(404, "Vehicle not found")
    if v.get("status") != "available": raise HTTPException(400, f"Vehicle is already {v.get('status')}")
    expenses_total = sum(float(e.get("amount", 0)) for e in sale.extra_expenses)
    total_amount = sale.sale_price + expenses_total
    paid_total = (sale.paid_cash or 0) + (sale.paid_bank or 0)
    due_amount = max(round(total_amount - paid_total, 2), 0)
    payment_status = "Paid" if due_amount <= 0 else ("Partial" if paid_total > 0 else "Unpaid")
    sale_date = sale.sale_date or datetime.now(timezone.utc).date().isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "vehicle_id": sale.vehicle_id,
        "customer_id": sale.customer_id,
        "sale_price": sale.sale_price,
        "extra_expenses": sale.extra_expenses,
        "expenses_total": expenses_total,
        "total_amount": total_amount,
        "payment_method": sale.payment_method,
        "paid_cash": sale.paid_cash or 0,
        "paid_bank": sale.paid_bank or 0,
        "due_amount": due_amount,
        "due_date": sale.due_date,
        "payment_status": payment_status,
        "sale_date": sale_date,
        "notes": sale.notes,
        "created_by": cu.get("username"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.sales.insert_one(doc)
    # Defense in depth: strip out any "Job Card " item the client sent — job card cost must
    # never be billed to the customer, see _strip_job_card_extra_expenses.
    cleaned = await _strip_job_card_extra_expenses(doc)
    if cleaned:
        doc.update(cleaned)
    doc.pop("_id", None)
    # Mark vehicle as sold
    await db.vehicles.update_one({"id": sale.vehicle_id}, {"$set": {
        "status": "sold",
        "selling_price": sale.sale_price,
        "sold_date": sale_date,
        "customer_id": sale.customer_id,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    # Update customer purchase history
    if sale.customer_id:
        await db.customers.update_one({"id": sale.customer_id}, {"$set": {"last_purchase_date": sale_date}})
    return doc

@api_router.get("/sales/summary")
async def get_sales_summary(cu: dict = Depends(require("sales", "view"))):
    sales = await db.sales.find({}, {"_id": 0}).to_list(1000)
    total_revenue = sum(_sale_revenue(s) for s in sales)
    this_month = datetime.now(timezone.utc).strftime("%Y-%m")
    monthly = [s for s in sales if s.get("sale_date", "").startswith(this_month)]
    avg = total_revenue / len(sales) if sales else 0
    total_due = round(sum(s.get("due_amount", 0) for s in sales), 2)
    today_iso = datetime.now(timezone.utc).date().isoformat()
    due_count = sum(1 for s in sales if s.get("due_amount", 0) > 0)
    overdue_count = sum(1 for s in sales if s.get("due_amount", 0) > 0 and s.get("due_date") and s.get("due_date") < today_iso)
    return {"total_sales": len(sales), "total_revenue": total_revenue, "this_month_sales": len(monthly), "this_month_revenue": sum(_sale_revenue(s) for s in monthly), "avg_sale_price": round(avg, 2), "total_due": total_due, "due_count": due_count, "overdue_count": overdue_count}

@api_router.get("/sales/reconcile")
async def reconcile_sales(cu: dict = Depends(admin_only)):
    """Diagnostic: finds sales whose linked vehicle no longer has status "sold" — the
    usual cause of the Sales-tab total drifting from the Sold Stock count. Happens when
    a sold vehicle's status gets changed (or the vehicle deleted) directly from Inventory
    instead of through "Delete Sale", which is the only path that keeps both in sync.

    Two legitimate exceptions are excluded from this check, both from PATCH /vehicles/{vid}/status:
    - "in_repair": the warranty-repair cycle, where a sold vehicle is temporarily brought
      back into the recondition pipeline and later flipped back to "sold" with its original
      sold_date preserved. The sale stays active throughout.
    - sales flagged "returned": a recondition-house return, where the vehicle is handed
      back and re-enters stock (e.g. flipped to "available") instead of going through
      "Delete Sale". The sale is intentionally kept (with its "returned" flag set) for
      history/revenue instead of deleted, so it's excluded here entirely — its vehicle's
      current status (and even later deletion) no longer implies a desync.
    Only genuine mismatches (deleted vehicle, or any other unflagged status drift) surface here.

    This only catches sales drifting away from their vehicle. The opposite drift — a vehicle
    sitting at status "sold" (so it counts in Sold Stock) with no active sale record backing it
    (so it doesn't count in Sales) — happens when a vehicle's status is set to "sold" by a path
    that skips the auto-create-sale logic (e.g. bulk import, or a direct DB edit), and is checked
    separately below."""
    sales = await db.sales.find({}, {"_id": 0}).to_list(1000)
    mismatches = []
    for s in sales:
        if s.get("returned"):
            continue
        v = await db.vehicles.find_one({"id": s.get("vehicle_id")}, {"_id": 0})
        if not v:
            mismatches.append({
                "sale_id": s["id"], "vehicle_id": s.get("vehicle_id"), "issue": "vehicle_deleted",
                "vehicle_info": None, "vehicle_status": None,
                "sale_date": s.get("sale_date"), "total_amount": s.get("total_amount"),
            })
        elif v.get("status") != "sold" and v.get("status") != "in_repair":
            mismatches.append({
                "sale_id": s["id"], "vehicle_id": s.get("vehicle_id"), "issue": "vehicle_status_mismatch",
                "vehicle_info": f"{v.get('brand','')} {v.get('model','')} {v.get('year','')}".strip(),
                "vehicle_status": v.get("status"),
                "sale_date": s.get("sale_date"), "total_amount": s.get("total_amount"),
            })

    sold_vehicles = await db.vehicles.find({"status": "sold"}, {"_id": 0}).to_list(1000)
    for v in sold_vehicles:
        active_sale = await db.sales.find_one({"vehicle_id": v["id"], "returned": {"$ne": True}}, {"_id": 0, "id": 1})
        if not active_sale:
            mismatches.append({
                "sale_id": None, "vehicle_id": v["id"], "issue": "orphan_sold_vehicle",
                "vehicle_info": f"{v.get('brand','')} {v.get('model','')} {v.get('year','')}".strip(),
                "vehicle_status": v.get("status"),
                "sale_date": v.get("sold_date"), "total_amount": v.get("selling_price"),
            })
    return {"count": len(mismatches), "mismatches": mismatches}

# Job card cost must never be billed to the customer — it reduces margin purely by being
# counted in _vehicle_investment (see enrich_vehicle / _vehicle_investment), the same way a
# recon expense does. It must NOT also appear as a Sale "extra expense" line item: that field
# feeds total_amount/due_amount — what the customer actually owes — so adding it there raises
# the price they're billed instead of just lowering the margin the shop reports internally.
#
# An earlier version of this code briefly did add job cards to extra_expenses, which live
# sale records may still carry. This strips any such line item back out and recomputes the
# sale's totals, so an already-affected sale self-heals the next time it's read.
async def _strip_job_card_extra_expenses(s: dict) -> Optional[dict]:
    if s.get("returned"):
        return None
    kept = [e for e in s.get("extra_expenses", []) if not str(e.get("name", "")).startswith("Job Card ")]
    if kept == s.get("extra_expenses", []):
        return None
    expenses_total = round(sum(e.get("amount", 0) for e in kept), 2)
    total_amount = round(s.get("sale_price", 0) + expenses_total, 2)
    paid_total = (s.get("paid_cash", 0) or 0) + (s.get("paid_bank", 0) or 0)
    due_amount = max(round(total_amount - paid_total, 2), 0)
    payment_status = "Paid" if due_amount <= 0 else ("Partial" if paid_total > 0 else "Unpaid")
    update = {
        "extra_expenses": kept, "expenses_total": expenses_total,
        "total_amount": total_amount, "due_amount": due_amount, "payment_status": payment_status,
    }
    await db.sales.update_one({"id": s["id"]}, {"$set": update})
    s.update(update)
    return update

@api_router.get("/sales/{sid}")
async def get_sale(sid: str, cu: dict = Depends(require("sales", "view"))):
    s = await db.sales.find_one({"id": sid}, {"_id": 0})
    if not s: raise HTTPException(404, "Not found")
    # Self-heal: strip out any job-card item a now-removed code path already billed onto
    # this sale — see _strip_job_card_extra_expenses.
    await _strip_job_card_extra_expenses(s)
    v = await db.vehicles.find_one({"id": s.get("vehicle_id")}, {"_id": 0})
    if v:
        s["vehicle_info"] = f"{v.get('brand','')} {v.get('model','')} {v.get('year','')}".strip()
        s["vehicle_brand"] = v.get("brand"); s["vehicle_model"] = v.get("model")
        s["vehicle_year"] = v.get("year"); s["registration_number"] = v.get("registration_number")
        s["engine_cc"] = v.get("engine_cc"); s["fuel_type"] = v.get("fuel_type")
        s["vehicle_status"] = v.get("status")
    s["job_card_cost"] = await _job_card_cost_total(s["vehicle_id"]) if s.get("vehicle_id") else 0
    c = await db.customers.find_one({"id": s.get("customer_id")}, {"_id": 0}) if s.get("customer_id") else None
    s["customer_name"] = c["name"] if c else "Walk-in Customer"
    s["customer_contact"] = c.get("contact_number") if c else None
    s["customer_address"] = c.get("address") if c else None
    # Margin/profit reveal what the shop paid for the vehicle — restricted to Admin
    # to keep that number away from front desk, same as the vehicle-level fields.
    if cu.get("role", "admin") == "admin" and v:
        investment = await _vehicle_investment(s["vehicle_id"], v)
        total = _sale_revenue(s)
        s["total_investment"] = investment
        s["profit"] = total - investment
        s["profit_margin"] = round(((total - investment) / total) * 100, 2) if total else None
    return s

@api_router.put("/sales/{sid}")
async def update_sale(sid: str, sale: SaleUpdate, cu: dict = Depends(get_current_user)):
    if cu.get("role") != "admin":
        raise HTTPException(403, "Only admin accounts can edit sales records")
    existing = await db.sales.find_one({"id": sid}, {"_id": 0})
    if not existing: raise HTTPException(404, "Not found")
    v = await db.vehicles.find_one({"id": sale.vehicle_id}, {"_id": 0})
    if not v: raise HTTPException(404, "Vehicle not found")
    expenses_total = sum(float(e.get("amount", 0)) for e in sale.extra_expenses)
    total_amount = sale.sale_price + expenses_total
    paid_total = (sale.paid_cash or 0) + (sale.paid_bank or 0)
    due_amount = max(round(total_amount - paid_total, 2), 0)
    payment_status = "Paid" if due_amount <= 0 else ("Partial" if paid_total > 0 else "Unpaid")
    sale_date = sale.sale_date or existing.get("sale_date")
    update_doc = {
        "vehicle_id": sale.vehicle_id,
        "customer_id": sale.customer_id,
        "sale_price": sale.sale_price,
        "extra_expenses": sale.extra_expenses,
        "expenses_total": expenses_total,
        "total_amount": total_amount,
        "payment_method": sale.payment_method,
        "paid_cash": sale.paid_cash or 0,
        "paid_bank": sale.paid_bank or 0,
        "due_amount": due_amount,
        "due_date": sale.due_date,
        "payment_status": payment_status,
        "sale_date": sale_date,
        "notes": sale.notes,
        "updated_by": cu.get("username"),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.sales.update_one({"id": sid}, {"$set": update_doc})
    # Keep vehicle record in sync (price/date/customer may have changed)
    await db.vehicles.update_one({"id": sale.vehicle_id}, {"$set": {
        "selling_price": sale.sale_price,
        "sold_date": sale_date,
        "customer_id": sale.customer_id,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    if sale.customer_id:
        await db.customers.update_one({"id": sale.customer_id}, {"$set": {"last_purchase_date": sale_date}})
    updated = await db.sales.find_one({"id": sid}, {"_id": 0})
    return updated

@api_router.delete("/sales/{sid}")
async def delete_sale(sid: str, cu: dict = Depends(get_current_user)):
    if cu.get("role") != "admin":
        raise HTTPException(403, "Only admin accounts can delete sales records")
    s = await db.sales.find_one({"id": sid}, {"_id": 0})
    if not s: raise HTTPException(404, "Not found")
    await db.sales.delete_one({"id": sid})
    # Restore vehicle to available
    await db.vehicles.update_one({"id": s["vehicle_id"]}, {"$set": {
        "status": "available", "sold_date": None, "customer_id": None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    return {"message": "Sale deleted, vehicle restored to available"}

# ── TEAM ──────────────────────────────────────────────────────────────
@api_router.get("/team")
async def get_team(cu: dict = Depends(require("team", "view"))):
    members = await db.team_members.find({}, {"_id": 0}).to_list(100)
    for m in members:
        if m.get("role") == "mechanic":
            m["total_jobs"] = await db.job_cards.count_documents({"mechanic_id": m["id"]})
            m["completed_jobs"] = await db.job_cards.count_documents({"mechanic_id": m["id"], "status": "completed"})
            m["completion_rate"] = round(m["completed_jobs"] / m["total_jobs"] * 100) if m["total_jobs"] > 0 else 0
    return members

@api_router.get("/team/leaderboard")
async def get_leaderboard(cu: dict = Depends(require("team", "view"))):
    sales_staff = await db.team_members.find({"role": "sales"}, {"_id": 0}).to_list(50)
    sales_ids = [s["id"] for s in sales_staff]
    sold_by_salesperson: dict = {}
    if sales_ids:
        sold = await db.vehicles.find({"salesperson_id": {"$in": sales_ids}}, {"_id": 0, "salesperson_id": 1, "selling_price": 1}).to_list(20000)
        for v in sold:
            sold_by_salesperson.setdefault(v["salesperson_id"], []).append(v)
    sales_board = []
    for s in sales_staff:
        sold_vehicles = sold_by_salesperson.get(s["id"], [])
        revenue = sum((v.get("selling_price") or 0) for v in sold_vehicles)
        sales_board.append({**s, "vehicles_sold": len(sold_vehicles), "revenue_generated": revenue})
    sales_board.sort(key=lambda x: x["vehicles_sold"], reverse=True)

    mechanics = await db.team_members.find({"role": "mechanic"}, {"_id": 0}).to_list(50)
    mech_ids = [m["id"] for m in mechanics]
    jobs_by_mechanic: dict = {}
    if mech_ids:
        jobs = await db.job_cards.find({"mechanic_id": {"$in": mech_ids}}, {"_id": 0, "mechanic_id": 1, "status": 1}).to_list(20000)
        for j in jobs:
            jobs_by_mechanic.setdefault(j["mechanic_id"], []).append(j)
    mech_board = []
    for m in mechanics:
        jobs = jobs_by_mechanic.get(m["id"], [])
        total = len(jobs)
        done = sum(1 for j in jobs if j.get("status") == "completed")
        mech_board.append({**m, "total_jobs": total, "completed_jobs": done,
                           "completion_rate": round(done/total*100) if total > 0 else 0})
    mech_board.sort(key=lambda x: x["completed_jobs"], reverse=True)
    return {"sales_leaderboard": sales_board, "mechanics_leaderboard": mech_board}

@api_router.post("/team")
async def create_team_member(member: TeamMemberCreate, cu: dict = Depends(require("team", "create"))):
    m = member.model_dump()
    m["id"] = str(uuid.uuid4()); m["is_active"] = True
    m["joining_date"] = m.get("joining_date") or datetime.now(timezone.utc).date().isoformat()
    m["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.team_members.insert_one(m)
    m.pop("_id", None)
    return m

@api_router.put("/team/{mid}")
async def update_team_member(mid: str, member: TeamMemberCreate, cu: dict = Depends(require("team", "edit"))):
    r = await db.team_members.update_one({"id": mid}, {"$set": member.model_dump()})
    if r.matched_count == 0: raise HTTPException(404, "Not found")
    return await db.team_members.find_one({"id": mid}, {"_id": 0})

@api_router.delete("/team/{mid}")
async def delete_team_member(mid: str, cu: dict = Depends(require("team", "delete"))):
    r = await db.team_members.delete_one({"id": mid})
    if r.deleted_count == 0: raise HTTPException(404, "Not found")
    return {"message": "Deleted"}

# ── PARTNERS ──────────────────────────────────────────────────────────
@api_router.get("/partners")
async def get_partners(cu: dict = Depends(admin_only)):
    return await db.partners.find({}, {"_id": 0}).to_list(100)

@api_router.post("/partners")
async def create_partner(partner: PartnerCreate, cu: dict = Depends(admin_only)):
    p = partner.model_dump()
    p["id"] = str(uuid.uuid4()); p["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.partners.insert_one(p)
    p.pop("_id", None)
    return p

@api_router.put("/partners/{pid}")
async def update_partner(pid: str, partner: PartnerCreate, cu: dict = Depends(admin_only)):
    r = await db.partners.update_one({"id": pid}, {"$set": partner.model_dump()})
    if r.matched_count == 0: raise HTTPException(404, "Not found")
    return await db.partners.find_one({"id": pid}, {"_id": 0})

@api_router.delete("/partners/{pid}")
async def delete_partner(pid: str, cu: dict = Depends(admin_only)):
    r = await db.partners.delete_one({"id": pid})
    if r.deleted_count == 0: raise HTTPException(404, "Not found")
    return {"message": "Deleted"}

# ── VENDORS ───────────────────────────────────────────────────────────
@api_router.get("/vendors")
async def get_vendors(cu: dict = Depends(require("vendors", "view"))):
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(200)
    vendor_ids = [v["id"] for v in vendors]

    vehicles_by_vendor, parts_by_vendor, payments_by_vendor = {}, {}, {}
    if vendor_ids:
        all_vehicles = await db.vehicles.find(
            {"$or": [{"vendor_id": {"$in": vendor_ids}}, {"linked_contact_type": "vendor", "linked_contact_id": {"$in": vendor_ids}}]},
            {"_id": 0},
        ).to_list(20000)
        for vh in all_vehicles:
            vid = vh.get("vendor_id") or (vh.get("linked_contact_id") if vh.get("linked_contact_type") == "vendor" else None)
            if vid: vehicles_by_vendor.setdefault(vid, []).append(vh)

        all_parts = await db.spare_parts.find({"vendor_id": {"$in": vendor_ids}}, {"_id": 0}).to_list(20000)
        for p in all_parts:
            parts_by_vendor.setdefault(p["vendor_id"], []).append(p)

        all_payments = await db.vendor_payments.find({"vendor_id": {"$in": vendor_ids}}, {"_id": 0}).to_list(20000)
        for p in all_payments:
            payments_by_vendor.setdefault(p["vendor_id"], []).append(p)

    for v in vendors:
        vehicles = vehicles_by_vendor.get(v["id"], [])
        parts = parts_by_vendor.get(v["id"], [])
        vehicle_owed = sum(vh.get("purchase_price", 0) for vh in vehicles)
        parts_owed = sum(p.get("quantity", 0) * p.get("unit_cost", 0) for p in parts)
        total_owed = vehicle_owed + parts_owed
        payments = payments_by_vendor.get(v["id"], [])
        total_paid = sum(p["amount"] for p in payments)
        v["total_purchased"] = total_owed; v["total_paid"] = total_paid
        v["remaining_due"] = max(0, total_owed - total_paid)
        v["vehicle_count"] = len(vehicles)
        v["parts_count"] = len(parts)
        v["parts_purchased"] = parts_owed
        v["overdue"] = v["remaining_due"] > 0
    return vendors

@api_router.post("/vendors")
async def create_vendor(vendor: VendorCreate, cu: dict = Depends(require("vendor_lookup", "create"))):
    v = vendor.model_dump()
    v["id"] = str(uuid.uuid4()); v["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.vendors.insert_one(v)
    v.pop("_id", None)
    return v

@api_router.put("/vendors/{vid}")
async def update_vendor(vid: str, vendor: VendorCreate, cu: dict = Depends(require("vendors", "edit"))):
    r = await db.vendors.update_one({"id": vid}, {"$set": vendor.model_dump()})
    if r.matched_count == 0: raise HTTPException(404, "Not found")
    return await db.vendors.find_one({"id": vid}, {"_id": 0})

@api_router.delete("/vendors/{vid}")
async def delete_vendor(vid: str, cu: dict = Depends(require("vendors", "delete"))):
    r = await db.vendors.delete_one({"id": vid})
    if r.deleted_count == 0: raise HTTPException(404, "Not found")
    return {"message": "Deleted"}

@api_router.get("/vendors/search")
async def search_vendors(q: str = "", vendor_type: Optional[str] = None, cu: dict = Depends(require("vendor_lookup", "view"))):
    """Fast vendor name search for autocomplete."""
    vendors = await db.vendors.find({}, {"_id": 0, "id": 1, "name": 1, "phone": 1, "vendor_type": 1}).to_list(200)
    if vendor_type:
        vendors = [v for v in vendors if v.get("vendor_type", "both") in (vendor_type, "both")]
    if q:
        q_lower = q.lower()
        vendors = [v for v in vendors if q_lower in v.get("name", "").lower()]
    return vendors[:8] if q else vendors

@api_router.get("/vendors/{vid}/payments")
async def get_vendor_payments(vid: str, cu: dict = Depends(require("vendors", "view"))):
    payments = await db.vendor_payments.find({"vendor_id": vid}, {"_id": 0}).sort("payment_date", -1).to_list(500)
    vehicles = await db.vehicles.find(_vendor_vehicle_filter(vid), {"_id": 0}).to_list(200)
    parts = await db.spare_parts.find({"vendor_id": vid}, {"_id": 0}).to_list(1000)
    vehicle_owed = sum(v.get("purchase_price", 0) for v in vehicles)
    parts_owed = sum(p.get("quantity", 0) * p.get("unit_cost", 0) for p in parts)
    total_owed = vehicle_owed + parts_owed
    total_paid = sum(p["amount"] for p in payments)
    bills = {}
    for p in parts:
        key = p.get("bill_no") or "No Bill No."
        b = bills.setdefault(key, {"bill_no": key, "entry_date": p.get("entry_date") or (p.get("created_at", "")[:10]), "items": [], "total": 0})
        b["items"].append(p)
        b["total"] += p.get("quantity", 0) * p.get("unit_cost", 0)
    parts_bills = sorted(bills.values(), key=lambda b: b["entry_date"] or "", reverse=True)
    return {"payments": payments, "total_paid": total_paid,
            "total_owed": total_owed, "remaining_due": max(0, total_owed - total_paid),
            "vehicles": vehicles, "parts_bills": parts_bills}

@api_router.post("/vendor-payments")
async def create_vendor_payment(payment: VendorPaymentCreate, cu: dict = Depends(require("vendors", "manage_payments"))):
    if not await db.vendors.find_one({"id": payment.vendor_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "Vendor not found")
    p = payment.model_dump()
    p["id"] = str(uuid.uuid4())
    p["payment_date"] = p.get("payment_date") or datetime.now(timezone.utc).date().isoformat()
    p["recorded_by"] = cu["username"]
    p["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.vendor_payments.insert_one(p)
    p.pop("_id", None)
    return p

@api_router.delete("/vendor-payments/{pid}")
async def delete_vendor_payment(pid: str, cu: dict = Depends(require("vendors", "manage_payments"))):
    r = await db.vendor_payments.delete_one({"id": pid})
    if r.deleted_count == 0: raise HTTPException(404, "Not found")
    return {"message": "Deleted"}

# ── EMI ───────────────────────────────────────────────────────────────
@api_router.get("/emi")
async def get_emi_list(cu: dict = Depends(admin_only)):
    emis = await db.emi_records.find({}, {"_id": 0}).to_list(500)
    for e in emis:
        payments = await db.emi_payments.find({"emi_id": e["id"]}, {"_id": 0}).to_list(200)
        paid = sum(p["amount"] for p in payments)
        e["total_paid"] = paid
        e["remaining_balance"] = max(0, e.get("loan_amount", 0) - paid)
        e["payments_made"] = len(payments)
        e["is_active"] = e["remaining_balance"] > 0
    return emis

@api_router.post("/emi")
async def create_emi(emi: EMICreate, cu: dict = Depends(admin_only)):
    e = emi.model_dump()
    e["id"] = str(uuid.uuid4())
    # Calculate monthly installment: EMI = P * r * (1+r)^n / ((1+r)^n - 1)
    p_val = e["loan_amount"]
    r = e["interest_rate"] / 100 / 12
    n = e["tenure_months"]
    if r > 0:
        monthly = p_val * r * ((1 + r) ** n) / (((1 + r) ** n) - 1)
    else:
        monthly = p_val / n
    e["monthly_installment"] = round(monthly, 2)
    e["total_payable"] = round(monthly * n, 2)
    e["total_interest"] = round(monthly * n - p_val, 2)
    e["status"] = "active"
    e["created_at"] = datetime.now(timezone.utc).isoformat()
    # Get customer info
    customer = await db.customers.find_one({"id": emi.customer_id}, {"_id": 0})
    if customer:
        e["customer_name"] = customer.get("name")
        e["customer_phone"] = customer.get("contact_number")
    # Get vehicle info
    vehicle = await db.vehicles.find_one({"id": emi.vehicle_id}, {"_id": 0})
    if vehicle:
        e["vehicle_name"] = f"{vehicle.get('brand')} {vehicle.get('model')} {vehicle.get('year')}"
    await db.emi_records.insert_one(e)
    e.pop("_id", None)
    return e

@api_router.post("/emi-payments")
async def add_emi_payment(payment: EMIPaymentCreate, cu: dict = Depends(admin_only)):
    if not await db.emi_records.find_one({"id": payment.emi_id}, {"_id": 0, "id": 1}):
        raise HTTPException(404, "EMI record not found")
    p = payment.model_dump()
    p["id"] = str(uuid.uuid4())
    p["payment_date"] = p.get("payment_date") or datetime.now(timezone.utc).date().isoformat()
    p["recorded_by"] = cu["username"]
    p["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.emi_payments.insert_one(p)
    p.pop("_id", None)
    return p

# ── FINANCE ───────────────────────────────────────────────────────────
@api_router.get("/finance/summary")
async def finance_summary(cu: dict = Depends(admin_only)):
    # Inventory value (available vehicles)
    avail = await db.vehicles.find({"status": "available"}, {"_id": 0}).to_list(1000)
    inventory_value = sum((await _batch_vehicle_investment(avail)).values())
    # Revenue & COGS from Sales table (single source of truth)
    sales_records = await db.sales.find({}, {"_id": 0}).to_list(1000)
    total_revenue = sum(_sale_revenue(s) for s in sales_records)
    # Dedupe: a vehicle sold, returned, then resold appears in two sales rows but its
    # investment (purchase price + expenses) should only count once toward COGS.
    sold_vehicle_ids = list({s["vehicle_id"] for s in sales_records})
    sold_vehicles = (
        await db.vehicles.find({"id": {"$in": sold_vehicle_ids}}, {"_id": 0}).to_list(5000)
        if sold_vehicle_ids else []
    )
    total_cogs = sum((await _batch_vehicle_investment(sold_vehicles)).values())
    gross_profit = total_revenue - total_cogs

    # Vendor payables
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(100)
    vendor_payables = sum((await _batch_vendor_payable([v["id"] for v in vendors])).values())

    # EMI receivables
    emis = await db.emi_records.find({"status": "active"}, {"_id": 0}).to_list(200)
    emi_receivables = sum((await _batch_emi_remaining(emis)).values())

    partners = await db.partners.find({}, {"_id": 0}).to_list(100)
    total_capital = sum(p.get("capital_contribution", 0) for p in partners)

    return {
        "inventory_value": inventory_value,
        "total_revenue": total_revenue,
        "total_cogs": total_cogs,
        "gross_profit": gross_profit,
        "profit_margin_pct": round((gross_profit / total_revenue * 100), 2) if total_revenue > 0 else 0,
        "vendor_payables": vendor_payables,
        "emi_receivables": emi_receivables,
        "total_partner_capital": total_capital,
        "vehicles_in_stock": len(avail),
        "vehicles_sold": len(sales_records),
    }

# ── REPORTS ───────────────────────────────────────────────────────────
@api_router.get("/reports/dashboard")
async def dashboard_stats(cu: dict = Depends(admin_only)):
    # These 4 counts are independent — running them concurrently instead of one
    # await after another turns 4 round-trips into the time of the slowest one.
    total, available, reserved, in_repair = await asyncio.gather(
        db.vehicles.count_documents({}),
        db.vehicles.count_documents({"status": "available"}),
        db.vehicles.count_documents({"status": "reserved"}),
        db.vehicles.count_documents({"status": "in_repair"}),
    )

    avail_v = await db.vehicles.find({"status": "available"}, {"_id": 0}).to_list(1000)
    investment_by_vehicle = await _batch_vehicle_investment(avail_v)
    locked_capital = sum(investment_by_vehicle.values())
    aging = _aging_counts(avail_v)

    # Use Sales table as single source of truth for sold count & profit
    sales_records = await db.sales.find({}, {"_id": 0}).to_list(1000)
    sold = len(sales_records)

    # Sold vehicles' investment feeds both the profit loop below and total_cogs —
    # fetch + batch-compute once instead of once per sale/vehicle.
    sold_vehicle_ids = list({s["vehicle_id"] for s in sales_records})
    sold_vehicles = (
        await db.vehicles.find({"id": {"$in": sold_vehicle_ids}}, {"_id": 0}).to_list(5000)
        if sold_vehicle_ids else []
    )
    sold_investment_by_vehicle = await _batch_vehicle_investment(sold_vehicles)
    sold_vehicle_ids_existing = {v["id"] for v in sold_vehicles}

    total_profit = 0
    for s in sales_records:
        if s.get("returned"):
            # The vehicle's investment cost stays attributed to inventory (it'll be
            # subtracted again on its actual resale) — only the retained refund fee
            # counts as profit here, so it isn't double-charged against COGS.
            total_profit += _sale_revenue(s)
            continue
        if s["vehicle_id"] in sold_vehicle_ids_existing:
            total_profit += _sale_revenue(s) - sold_investment_by_vehicle.get(s["vehicle_id"], 0)

    total_revenue = sum(_sale_revenue(s) for s in sales_records)
    # Dedupe: a vehicle sold, returned, then resold appears in two sales rows but its
    # investment (purchase price + expenses) should only count once toward COGS.
    total_cogs = sum(sold_investment_by_vehicle.values())

    vendors = await db.vendors.find({}, {"_id": 0}).to_list(100)
    payable_by_vendor = await _batch_vendor_payable([v["id"] for v in vendors])
    total_vendor_due = sum(payable_by_vendor.values())

    pending_jobs, in_progress_jobs, total_customers, total_vendors = await asyncio.gather(
        db.job_cards.count_documents({"status": "pending"}),
        db.job_cards.count_documents({"status": "in_progress"}),
        db.customers.count_documents({}),
        db.vendors.count_documents({}),
    )

    return {
        "total_vehicles": total, "available": available, "sold": sold,
        "reserved": reserved, "in_repair": in_repair,
        "locked_capital": locked_capital, "total_realized_profit": total_profit,
        "dead_stock_count": aging["dead"], "slow_moving_count": aging["slow"],
        "normal_count": aging["normal"], "fresh_count": aging["fresh"],
        "pending_jobs": pending_jobs,
        "in_progress_jobs": in_progress_jobs,
        "total_customers": total_customers,
        "total_vendor_due": total_vendor_due,
        "total_vendors": total_vendors,
        "total_revenue": total_revenue, "inventory_value": locked_capital,
        "total_cogs": total_cogs,
    }

@api_router.get("/reports/inventory")
async def inventory_report(cu: dict = Depends(admin_only)):
    vehicles = await db.vehicles.find({}, {"_id": 0}).to_list(1000)
    report = {"by_brand": {}, "by_status": {}, "by_aging": {"fresh": 0, "normal": 0, "slow": 0, "dead": 0},
              "by_source": {}, "slow_moving": [], "dead_stock": [], "by_fuel": {}}
    for v in vehicles:
        brand = v.get("brand", "Unknown")
        report["by_brand"][brand] = report["by_brand"].get(brand, 0) + 1
        st = v.get("status", "available")
        report["by_status"][st] = report["by_status"].get(st, 0) + 1
        src = v.get("purchase_source", "Unknown")
        if src not in report["by_source"]: report["by_source"][src] = {"count": 0}
        report["by_source"][src]["count"] += 1
        fuel = v.get("fuel_type", "Petrol")
        report["by_fuel"][fuel] = report["by_fuel"].get(fuel, 0) + 1
        if v.get("status") == "available":
            ag = stock_aging(v.get("purchase_date", ""))
            cat = ag["category"]
            if cat in report["by_aging"]: report["by_aging"][cat] += 1
            item = {"id": v["id"], "brand": v.get("brand"), "model": v.get("model"),
                    "year": v.get("year"), "days": ag["days"],
                    "purchase_price": v.get("purchase_price"), "selling_price": v.get("selling_price")}
            if cat == "slow": report["slow_moving"].append(item)
            elif cat == "dead": report["dead_stock"].append(item)
    return report

@api_router.get("/reports/financial")
async def financial_report(cu: dict = Depends(admin_only)):
    sold = await db.vehicles.find({"status": "sold"}, {"_id": 0}).to_list(1000)
    investment_by_vehicle = await _batch_vehicle_investment(sold)
    monthly = {}
    for v in sold:
        sd = v.get("sold_date") or v.get("updated_at", "")
        month = sd[:7] if sd else "unknown"
        if month not in monthly: monthly[month] = {"revenue": 0, "investment": 0, "profit": 0, "count": 0}
        inv = investment_by_vehicle.get(v["id"], 0)
        sp = v.get("selling_price") or 0
        monthly[month]["revenue"] += sp
        monthly[month]["investment"] += inv
        monthly[month]["profit"] += sp - inv
        monthly[month]["count"] += 1
    partners = await db.partners.find({}, {"_id": 0}).to_list(100)
    total_profit = sum(m["profit"] for m in monthly.values())
    partner_shares = [{"name": p["name"], "stake": p["stake_percentage"],
                       "capital": p["capital_contribution"],
                       "profit_share": round(total_profit * p["stake_percentage"] / 100, 2)} for p in partners]
    return {"monthly_breakdown": monthly, "total_profit": total_profit, "partner_shares": partner_shares}

# Sale rows enriched with everything the Finance tab's Monthly Breakdown (grouped into
# real BS months client-side, since all BS/AD conversion already lives in the frontend —
# see nepali-date.js) and its per-month closing-report export need. "extra_expense" here
# means repair/prep cost only (vehicle_photos... no — db.expenses + non-warranty job cards),
# deliberately excluding purchase_price: that split matches the approved Excel template's
# own definition ("not part of what the customer owes"), not the broader "investment"
# figure (purchase price + accessories + repair cost) used for profit elsewhere.
async def _enriched_sales_for_closing(start_date: Optional[str] = None, end_date: Optional[str] = None):
    query = {}
    if start_date and end_date:
        query["sale_date"] = {"$gte": start_date, "$lte": end_date}
    sales = await db.sales.find(query, {"_id": 0}).sort("sale_date", 1).to_list(5000)
    vehicle_ids = list({s["vehicle_id"] for s in sales if s.get("vehicle_id")})
    customer_ids = list({s["customer_id"] for s in sales if s.get("customer_id")})
    vehicles_by_id, customers_by_id = {}, {}
    if vehicle_ids:
        vs = await db.vehicles.find({"id": {"$in": vehicle_ids}}, {"_id": 0}).to_list(len(vehicle_ids))
        vehicles_by_id = {v["id"]: v for v in vs}
    if customer_ids:
        cs = await db.customers.find({"id": {"$in": customer_ids}}, {"_id": 0, "id": 1, "name": 1, "contact_number": 1}).to_list(len(customer_ids))
        customers_by_id = {c["id"]: c for c in cs}
    exps_by_vehicle: dict = {}
    jobs_by_vehicle: dict = {}
    if vehicle_ids:
        all_exps = await db.expenses.find({"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0}).to_list(20000)
        for e in all_exps:
            exps_by_vehicle.setdefault(e["vehicle_id"], []).append(e)
        all_jobs = await db.job_cards.find({"vehicle_id": {"$in": vehicle_ids}, "is_warranty": {"$ne": True}}, {"_id": 0}).to_list(20000)
        for j in all_jobs:
            jobs_by_vehicle.setdefault(j["vehicle_id"], []).append(j)

    out = []
    for s in sales:
        v = vehicles_by_id.get(s.get("vehicle_id"))
        c = customers_by_id.get(s.get("customer_id"))
        repair_cost = (sum(e.get("amount", 0) for e in exps_by_vehicle.get(s.get("vehicle_id"), []))
                       + sum(_job_card_cost(j) for j in jobs_by_vehicle.get(s.get("vehicle_id"), [])))
        investment = (v.get("purchase_price", 0) + v.get("accessories_cost", 0) + repair_cost) if v else 0
        vehicle_label = "Vehicle removed"
        if v:
            parts = [v["brand"], v["model"]]
            if v.get("year"): parts.append(str(v["year"]))
            vehicle_label = " ".join(parts) + (f" ({v['registration_number']})" if v.get("registration_number") else "")
        out.append({
            "id": s["id"], "sale_date": s.get("sale_date"),
            "sale_price": s.get("sale_price", 0) or 0,
            "due_amount": s.get("due_amount", 0) or 0,
            "returned": bool(s.get("returned")),
            "retained_amount": s.get("retained_amount"),
            "vehicle_label": vehicle_label,
            "customer_name": c["name"] if c else "Walk-in",
            "customer_contact": c.get("contact_number") if c else None,
            "extra_expense": round(repair_cost, 2),
            "investment": round(investment, 2),
        })
    return out

@api_router.get("/reports/monthly-breakdown-bs")
async def monthly_breakdown_bs(cu: dict = Depends(admin_only)):
    """All sales, enriched — the frontend buckets these into real Bikram Sambat months
    (see Finance.jsx) instead of the Gregorian-month grouping /reports/financial uses."""
    return await _enriched_sales_for_closing()

def _build_closing_report_xlsx(rows: list, month_label: str, prepared_by: str) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill

    template_path = ROOT_DIR / "report_templates" / "sales_closing_template.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Sales Closing"]

    def cell_style(row, col):
        c = ws.cell(row=row, column=col)
        return {"font": copy.copy(c.font), "fill": copy.copy(c.fill), "border": copy.copy(c.border),
                "alignment": copy.copy(c.alignment), "number_format": c.number_format}

    # Template rows alternate white/light-gray banding (row 5 white, row 6 gray, ...) —
    # capture both so generated rows keep that banding instead of all coming out white.
    data_row_style_odd = {col: cell_style(5, col) for col in range(1, 11)}
    data_row_style_even = {col: cell_style(6, col) for col in range(1, 11)}
    total_row_style = {col: cell_style(36, col) for col in range(1, 11)}
    legend_header_style = cell_style(38, 1)
    legend_line_styles = [cell_style(39 + i, 1) for i in range(4)]
    legend_texts = [ws.cell(row=39 + i, column=1).value for i in range(4)]
    total_label = ws.cell(row=36, column=1).value

    def apply_style(cell, style):
        cell.font = style["font"]; cell.fill = style["fill"]; cell.border = style["border"]
        cell.alignment = style["alignment"]; cell.number_format = style["number_format"]

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= 5:
            ws.unmerge_cells(str(rng))
    for row in ws.iter_rows(min_row=5, max_row=42, max_col=10):
        for c in row:
            c.value = None

    ws["A1"] = f"G&G Auto – Closing Report for the Month of {month_label}"
    today = datetime.now(timezone.utc).date().isoformat()
    ws["A2"] = f"Report Month:  {month_label}        Prepared By:  {prepared_by or '________'}        Date Prepared:  {today}"

    RETURNED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    FIRST_ROW = 5
    n = max(len(rows), 1)
    last_row = FIRST_ROW + n - 1
    total_row = last_row + 1

    # The template's Status-column conditional formatting (green "Paid" / red "Due") is
    # scoped to its original J5:J35 range — widen it to match however many rows we
    # actually write, or anything past row 35 renders with no color at all. openpyxl's
    # ConditionalFormatting objects are keyed by identity in an internal dict, so the
    # range can't be mutated in place; remove the old entry and re-add the same rules
    # under the new range instead.
    status_cf_rules = []
    for cf in list(ws.conditional_formatting):
        if str(cf.sqref).startswith("J5"):
            status_cf_rules = list(cf.rules)
            del ws.conditional_formatting._cf_rules[cf]
    for rule in status_cf_rules:
        ws.conditional_formatting.add(f"J{FIRST_ROW}:J{last_row}", rule)

    for i in range(n):
        row = FIRST_ROW + i
        s = rows[i] if i < len(rows) else None
        if s:
            is_returned = s["returned"]
            selling_price = float(s["retained_amount"] or 0) if is_returned else float(s["sale_price"] or 0)
            due = 0.0 if is_returned else float(s["due_amount"] or 0)
            values = {
                1: f'=IF(C{row}="","",ROW()-{FIRST_ROW}+1)', 2: s["sale_date"], 3: s["vehicle_label"],
                4: s["customer_name"] or "Walk-in", 5: s["customer_contact"] or "",
                6: selling_price, 7: s["extra_expense"], 8: due,
                9: f'=IF(C{row}="","",F{row}-H{row})',
                10: "Returned" if is_returned else f'=IF(C{row}="","",IF(H{row}=0,"Paid","Due"))',
            }
        else:
            is_returned = False
            values = {1: f'=IF(C{row}="","",ROW()-{FIRST_ROW}+1)', 9: f'=IF(C{row}="","",F{row}-H{row})',
                      10: f'=IF(C{row}="","",IF(H{row}=0,"Paid","Due"))'}
        row_style = data_row_style_odd if i % 2 == 0 else data_row_style_even
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col, value=values.get(col))
            apply_style(cell, row_style[col])
            if is_returned:
                cell.fill = RETURNED_FILL

    ws.cell(row=total_row, column=1, value=total_label)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    ws.cell(row=total_row, column=6, value=f"=SUM(F{FIRST_ROW}:F{last_row})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G{FIRST_ROW}:G{last_row})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H{FIRST_ROW}:H{last_row})")
    ws.cell(row=total_row, column=9, value=f"=SUM(I{FIRST_ROW}:I{last_row})")
    ws.cell(row=total_row, column=10, value=f'=COUNTIF(C{FIRST_ROW}:C{last_row},"?*")&" bikes"')
    for col in range(1, 11):
        apply_style(ws.cell(row=total_row, column=col), total_row_style[col])

    lr = total_row + 2
    ws.cell(row=lr, column=1, value="Legend:")
    apply_style(ws.cell(row=lr, column=1), legend_header_style)
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=10)
    for i, text in enumerate(legend_texts):
        r = lr + 1 + i
        ws.cell(row=r, column=1, value=text)
        apply_style(ws.cell(row=r, column=1), legend_line_styles[i])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    note_row = lr + 1 + len(legend_texts)
    ws.cell(row=note_row, column=1,
            value="• Rows highlighted in yellow are sales that were later returned/refunded — Selling Price shown is the amount actually retained, not the original sale price.")
    apply_style(ws.cell(row=note_row, column=1), legend_line_styles[-1])
    ws.cell(row=note_row, column=1).fill = RETURNED_FILL
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@api_router.get("/reports/monthly-closing-export")
async def monthly_closing_export(start_date: str, end_date: str, label: str, cu: dict = Depends(admin_only)):
    rows = await _enriched_sales_for_closing(start_date, end_date)
    xlsx_bytes = _build_closing_report_xlsx(rows, label, cu.get("username", ""))
    safe_label = re.sub(r"[^A-Za-z0-9]+", "_", label).strip("_")
    filename = f"GG_Auto_Closing_Report_{safe_label}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# Pipeline order mirrors VEHICLE_STATUS_OPTIONS in frontend/src/utils/helpers.js (scrap
# excluded there too — it's a terminal/do-not-disturb stage, not part of the live pipeline).
_INVENTORY_PIPELINE_STATUSES = ["unlisted", "in_repair", "available", "reserved", "sold"]
_INVENTORY_STATUS_LABELS = {"unlisted": "Unlisted", "in_repair": "In Repair", "available": "Available",
                            "reserved": "Reserved", "sold": "Sold"}

def _build_inventory_pipeline_xlsx(vehicles: list, prepared_by: str) -> bytes:
    import openpyxl
    template_path = ROOT_DIR / "report_templates" / "inventory_pipeline_template.xlsx"
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Inventory Pipeline"]

    def cell_style(row, col):
        c = ws.cell(row=row, column=col)
        return {"font": copy.copy(c.font), "fill": copy.copy(c.fill), "border": copy.copy(c.border),
                "alignment": copy.copy(c.alignment), "number_format": c.number_format}

    data_row_style_odd = {col: cell_style(5, col) for col in range(1, 10)}
    data_row_style_even = {col: cell_style(6, col) for col in range(1, 10)}

    def apply_style(cell, style):
        cell.font = style["font"]; cell.fill = style["fill"]; cell.border = style["border"]
        cell.alignment = style["alignment"]; cell.number_format = style["number_format"]

    FIRST_ROW = 5
    ORIG_LAST_ROW = 40  # template's built-in range before any extension
    n = max(len(vehicles), 1)
    last_row = FIRST_ROW + n - 1
    extra_rows = max(last_row - ORIG_LAST_ROW, 0)

    # Capture summary/legend styles + text before touching anything below the original range.
    orig_summary_title_row, orig_summary_labels_row, orig_summary_values_row = 42, 43, 44
    orig_legend_header_row = 46
    summary_title_style = cell_style(orig_summary_title_row, 1)
    summary_title_text = ws.cell(row=orig_summary_title_row, column=1).value
    summary_label_styles = [cell_style(orig_summary_labels_row, c) for c in range(1, 7)]
    summary_label_texts = [ws.cell(row=orig_summary_labels_row, column=c).value for c in range(1, 7)]
    summary_value_styles = [cell_style(orig_summary_values_row, c) for c in range(1, 7)]
    legend_header_style = cell_style(orig_legend_header_row, 1)
    legend_line_styles = [cell_style(47 + i, 1) for i in range(5)]
    legend_texts = [ws.cell(row=47 + i, column=1).value for i in range(5)]

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= 5:
            ws.unmerge_cells(str(rng))
    for row in ws.iter_rows(min_row=5, max_row=51, max_col=9):
        for c in row:
            c.value = None

    ws["A1"] = "G&G Auto – Vehicle Inventory Pipeline"
    today = datetime.now(timezone.utc).date().isoformat()
    ws["A2"] = f"As of:  {today}        Prepared By:  {prepared_by or '________'}"

    def vehicle_label(v):
        parts = [v["brand"].strip(), v["model"].strip()]
        if v.get("year"): parts.append(str(v["year"]))
        return " ".join(parts)

    for i in range(n):
        row = FIRST_ROW + i
        v = vehicles[i] if i < len(vehicles) else None
        if v:
            reg_or_chassis = v.get("registration_number") or v.get("chassis_number") or ""
            values = {
                1: f'=IF(C{row}="","",ROW()-{FIRST_ROW}+1)',
                2: v.get("purchase_date") or "", 3: vehicle_label(v), 4: reg_or_chassis,
                5: v.get("purchase_price") or 0, 6: _INVENTORY_STATUS_LABELS.get(v["status"], v["status"]),
                7: f'=IF(OR(C{row}="",B{row}=""),"",TODAY()-B{row})',
                8: v.get("selling_price") or 0, 9: v.get("notes") or "",
            }
        else:
            values = {1: f'=IF(C{row}="","",ROW()-{FIRST_ROW}+1)',
                      7: f'=IF(OR(C{row}="",B{row}=""),"",TODAY()-B{row})'}
        row_style = data_row_style_odd if i % 2 == 0 else data_row_style_even
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value=values.get(col))
            apply_style(cell, row_style[col])

    # Status dropdown + conditional-formatting range, widened to match however many rows
    # were actually written (template ships scoped to F5:F40 only) — see the identical
    # issue/fix for the Sales Closing template's Status column, above.
    for dv in ws.data_validations.dataValidation:
        if str(dv.sqref).startswith("F5"):
            dv.sqref = f"F{FIRST_ROW}:F{last_row}"
    status_cf_rules = []
    for cf in list(ws.conditional_formatting):
        if str(cf.sqref).startswith("F5"):
            status_cf_rules = list(cf.rules)
            del ws.conditional_formatting._cf_rules[cf]
    for rule in status_cf_rules:
        ws.conditional_formatting.add(f"F{FIRST_ROW}:F{last_row}", rule)

    # ── Summary + legend, shifted down by however many extra rows we needed ────────────
    summary_title_row = orig_summary_title_row + extra_rows
    summary_labels_row = orig_summary_labels_row + extra_rows
    summary_values_row = orig_summary_values_row + extra_rows
    legend_header_row = orig_legend_header_row + extra_rows

    ws.cell(row=summary_title_row, column=1, value=summary_title_text)
    apply_style(ws.cell(row=summary_title_row, column=1), summary_title_style)
    ws.merge_cells(start_row=summary_title_row, start_column=1, end_row=summary_title_row, end_column=9)

    for c in range(1, 7):
        cell = ws.cell(row=summary_labels_row, column=c, value=summary_label_texts[c - 1])
        apply_style(cell, summary_label_styles[c - 1])
    status_formulas = {
        1: f'=COUNTIF(F{FIRST_ROW}:F{last_row},"Available")',
        2: f'=COUNTIF(F{FIRST_ROW}:F{last_row},"Unlisted")',
        3: f'=COUNTIF(F{FIRST_ROW}:F{last_row},"In Repair")',
        4: f'=COUNTIF(F{FIRST_ROW}:F{last_row},"Reserved")',
        5: f'=COUNTIF(F{FIRST_ROW}:F{last_row},"Sold")',
        6: f'=COUNTIF(C{FIRST_ROW}:C{last_row},"?*")-COUNTIF(F{FIRST_ROW}:F{last_row},"Sold")',
    }
    for c in range(1, 7):
        cell = ws.cell(row=summary_values_row, column=c, value=status_formulas[c])
        apply_style(cell, summary_value_styles[c - 1])

    lr = legend_header_row
    ws.cell(row=lr, column=1, value="Legend:")
    apply_style(ws.cell(row=lr, column=1), legend_header_style)
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=9)
    for i, text in enumerate(legend_texts):
        r = lr + 1 + i
        ws.cell(row=r, column=1, value=text)
        apply_style(ws.cell(row=r, column=1), legend_line_styles[i])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@api_router.get("/reports/inventory-pipeline-export")
async def inventory_pipeline_export(cu: dict = Depends(require("vehicles", "view"))):
    vehicles = await db.vehicles.find({"status": {"$ne": "scrap"}}, {"_id": 0}).to_list(5000)
    order = {s: i for i, s in enumerate(_INVENTORY_PIPELINE_STATUSES)}
    vehicles.sort(key=lambda v: (order.get(v.get("status"), 99), v.get("purchase_date") or ""))
    xlsx_bytes = _build_inventory_pipeline_xlsx(vehicles, cu.get("username", ""))
    today = datetime.now(timezone.utc).date().isoformat()
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="GG_Auto_Inventory_Pipeline_{today}.xlsx"'},
    )

# ── ACCOUNTING SUMMARY ────────────────────────────────────────────────
@api_router.get("/reports/accounting-summary")
async def accounting_summary(start_date: str, end_date: str, cu: dict = Depends(admin_only)):
    """Returns cost/sales/profit for vehicles purchased/sold within [start_date, end_date]."""
    # Vehicles purchased in period (total cost = purchase + expenses)
    purchased = await db.vehicles.find(
        {"purchase_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).to_list(5000)
    purchase_count = len(purchased)
    total_cost = sum((await _batch_vehicle_investment(purchased)).values())

    # Sales in period — use Sales table as source of truth
    sales_in_period = await db.sales.find(
        {"sale_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).to_list(5000)
    sold_count = len(sales_in_period)
    total_sales = sum(_sale_revenue(s) for s in sales_in_period)

    # Retained refund fee (returned sales) has no cost basis subtracted — the vehicle's
    # investment is still attributed to inventory until it's actually resold. Note this
    # intentionally does NOT dedupe by vehicle_id: if the same vehicle was sold, returned,
    # and resold (all non-returned sales) within the period, its investment is meant to
    # count once per sale here — this mirrors the original per-sale-record calculation.
    non_returned = [s for s in sales_in_period if not s.get("returned")]
    sold_vehicle_ids = list({s["vehicle_id"] for s in non_returned})
    sold_vehicles = (
        await db.vehicles.find({"id": {"$in": sold_vehicle_ids}}, {"_id": 0}).to_list(5000)
        if sold_vehicle_ids else []
    )
    investment_by_vehicle = await _batch_vehicle_investment(sold_vehicles)
    total_investment_sold = sum(
        investment_by_vehicle[s["vehicle_id"]] for s in non_returned if s["vehicle_id"] in investment_by_vehicle
    )

    net_profit = total_sales - total_investment_sold
    return {
        "period": {"start": start_date, "end": end_date},
        "total_cost": total_cost,
        "purchase_count": purchase_count,
        "total_sales": total_sales,
        "sold_count": sold_count,
        "net_profit": net_profit,
        "total_investment_sold": total_investment_sold,
    }

# ── AUDIT LOGS ────────────────────────────────────────────────────────
@api_router.get("/audit-logs")
async def get_audit_logs(cu: dict = Depends(admin_only)):
    logs = await db.audit_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(200)
    return logs

# ── LEADS (storefront Sell / Exchange / Book Service submissions) ─────
@api_router.get("/leads")
async def get_leads(cu: dict = Depends(admin_only)):
    return await db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)

@api_router.put("/leads/{lid}")
async def update_lead(lid: str, lead: LeadUpdate, cu: dict = Depends(admin_only)):
    r = await db.leads.update_one({"id": lid}, {"$set": {"status": lead.status}})
    if r.matched_count == 0: raise HTTPException(404, "Lead not found")
    return await db.leads.find_one({"id": lid}, {"_id": 0})

@api_router.delete("/leads/{lid}")
async def delete_lead(lid: str, cu: dict = Depends(admin_only)):
    r = await db.leads.delete_one({"id": lid})
    if r.deleted_count == 0: raise HTTPException(404, "Lead not found")
    return {"message": "Deleted"}

# ── SETTINGS (storefront branding/contact info) ────────────────────────
@api_router.get("/settings")
async def get_settings(cu: dict = Depends(admin_only)):
    # No "id" filter: settings is looked up purely by the company_id the wrapper scopes
    # this query to -- one row per company, "id" is just an arbitrary row identifier.
    s = await db.settings.find_one({}, {"_id": 0})
    return s or {}

@api_router.put("/settings")
async def update_settings(settings: SettingsUpdate, cu: dict = Depends(admin_only)):
    updates = {k: v for k, v in settings.model_dump().items() if v is not None}
    await db.settings.update_one({}, {"$set": updates}, upsert=True)
    asyncio.create_task(_notify_storefront())
    return await db.settings.find_one({}, {"_id": 0})

# ── STARTUP ───────────────────────────────────────────────────────────
@app.on_event("startup")
async def startup():
    # Every real deployment already has a company (created via /auth/signup, or via the
    # one-time migration_add_company_id.sql backfill run when multi-tenancy was introduced).
    # A fresh install / CI database has neither, so this bootstraps one -- the
    # admin/frontdesk/parts accounts, seed partners, and default settings row below all need
    # SOME company_id to satisfy the NOT NULL + FOREIGN KEY hardening on those tables (see
    # migration_harden_company_id.sql). Runs every startup but only inserts once: idempotent
    # on the "does any company exist yet" check, same pattern as the accounts below it.
    default_company = await db.companies.find_one({}, {"_id": 0, "id": 1}, sort=[("created_at", 1)])
    if not default_company:
        default_company = {"id": str(uuid.uuid4()), "name": "Default Company",
                            "created_at": datetime.now(timezone.utc).isoformat()}
        await db.companies.insert_one(dict(default_company))
        logger.info("Default company created (fresh install): %s", default_company["id"])
    default_company_id = default_company["id"]

    if not await db.users.find_one({"username": "admin"}):
        await db.users.insert_one({
            "id": str(uuid.uuid4()), "username": "admin", "company_id": default_company_id,
            "password_hash": hash_pw("admin123"), "name": "Admin", "role": "admin",
            "email_verified_at": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info("Default admin created: admin / admin123")
    else:
        # Admin: the original account, unchanged except display name.
        await db.users.update_one({"username": "admin"}, {"$set": {"name": "Admin"}})

    ADDITIONAL_ACCOUNTS = [
        {"username": "frontdesk", "password": "frontdesk123", "name": "Front desk stock", "role": "stock_supervisor"},
        {"username": "parts", "password": "parts123", "name": "Parts department", "role": "parts_supervisor"},
    ]
    for acct in ADDITIONAL_ACCOUNTS:
        if not await db.users.find_one({"username": acct["username"]}):
            await db.users.insert_one({
                "id": str(uuid.uuid4()), "username": acct["username"], "company_id": default_company_id,
                "password_hash": hash_pw(acct["password"]), "name": acct["name"], "role": acct["role"],
                "email_verified_at": datetime.now(timezone.utc).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            logger.info(f"Default account created: {acct['username']} / {acct['password']}")

    # Email verification (shipped 2026-08-20) only applies to brand-new self-signups going
    # forward -- every account that existed before this feature must never be retroactively
    # locked out of login. Filtered on created_at, not just "$exists: False", because on
    # the MySQL backend a real signup's intentional email_verified_at=NULL (still pending
    # verification) is indistinguishable from "column never set" -- sqldb.py's $exists
    # translates to a plain IS NULL there, so without the created_at bound this would
    # silently re-verify every not-yet-confirmed signup on every server restart. Mongo's
    # $exists doesn't have that ambiguity, but the same filter is correct on both.
    EMAIL_VERIFICATION_CUTOFF = "2026-08-21T00:00:00+00:00"  # feature ship date, end-of-day UTC
    backfilled = await db.users.update_many(
        {"email_verified_at": {"$exists": False}, "created_at": {"$lte": EMAIL_VERIFICATION_CUTOFF}},
        {"$set": {"email_verified_at": datetime.now(timezone.utc).isoformat()}}
    )
    if backfilled.matched_count:
        logger.info(f"Backfilled email_verified_at for {backfilled.matched_count} pre-existing user(s)")
    if await db.partners.count_documents({}) == 0:
        now = datetime.now(timezone.utc).isoformat()
        await db.partners.insert_many([
            {"id": str(uuid.uuid4()), "company_id": default_company_id, "name": "Partner A", "capital_contribution": 500000, "stake_percentage": 33.33, "contact": "", "created_at": now},
            {"id": str(uuid.uuid4()), "company_id": default_company_id, "name": "Partner B", "capital_contribution": 500000, "stake_percentage": 33.33, "contact": "", "created_at": now},
            {"id": str(uuid.uuid4()), "company_id": default_company_id, "name": "You (Owner)", "capital_contribution": 500000, "stake_percentage": 33.34, "contact": "", "created_at": now},
        ])
    await db.kit_components.create_index([("kit_part_id", 1), ("component_part_id", 1)], unique=True)
    await db.kit_components.create_index("component_part_id")

    # Core collections had no indexes at all, so every find() below was a full
    # collection scan — the biggest lever on query latency once the DB is remote
    # (Atlas) from the app host, since scans get worse as data grows while an
    # indexed lookup stays flat. Mirrors the fields these collections are most
    # commonly filtered/joined on across server.py ($in batches, per-record lookups).
    await db.vehicles.create_index("status")
    await db.vehicles.create_index("id")
    await db.vehicles.create_index("vendor_id")
    await db.sales.create_index("sale_date")
    await db.sales.create_index("vehicle_id")
    await db.sales.create_index("customer_id")
    await db.expenses.create_index("vehicle_id")
    await db.customers.create_index("id")
    await db.vendor_payments.create_index("vendor_id")
    await db.job_cards.create_index("status")
    if not await db.settings.find_one({"id": "general"}):
        await db.settings.insert_one({
            "id": "general", "company_id": default_company_id,
            "logo_url": "https://images.unsplash.com/photo-1777288411485-1eb05bd4a289?q=80&w=880&auto=format&fit=crop&ixlib=rb-4.1.0&ixid=M3wxMjA3fDB8MHxwaG90by1wYWdlfHx8fGVufDB8fHx8fA%3D%3D",
            "business_name": "G&G AUTO Enterprises",
            "contact_phone": "9860087161",
            "contact_email": "info@ggautonp.com",
            "address": "Nayabasti, Boudha",
            "hero_image_url": "https://images.unsplash.com/photo-1622185135505-2d795003994a?q=80&w=1470&auto=format&fit=crop",
            "service_image_url": "",
        })

    # Run in the background so neither ever delays the app becoming ready — shrinks
    # any vehicle photo or legal document over their 2MB cap left over from before
    # those caps existed, or from a run that got interrupted last time. See
    # _compress_oversized_photos / _compress_oversized_documents.
    asyncio.create_task(_compress_oversized_photos())
    asyncio.create_task(_compress_oversized_documents())

_cors_origins_raw = os.environ.get('CORS_ORIGINS')
if not _cors_origins_raw:
    logger.warning(
        "CORS_ORIGINS is not set -- defaulting to '*' (any origin). This is fine for local "
        "dev but should never be left unset in production: set it to the frontend's exact "
        "URL(s), comma-separated, in the environment. See DEPLOYMENT.md."
    )
app.add_middleware(CORSMiddleware, allow_credentials=True,
                   allow_origins=(_cors_origins_raw or '*').split(','),
                   allow_methods=["*"], allow_headers=["*"])
# Compresses JSON responses over 1KB (list endpoints especially) before they hit the
# network — no API/behavior change, just a smaller wire payload.
app.add_middleware(GZipMiddleware, minimum_size=1000)


# ══════════════════════════════════════════════════════════════════════
# ── VEHICLE PHOTO UPLOAD ──────────────────────────────────────────────
# Stored as base64 in MongoDB (db.vehicle_photos), NOT local disk — Render's
# free-tier filesystem is ephemeral and wipes local files on every restart,
# which was causing uploaded photos/docs to vanish after the backend slept.
# ══════════════════════════════════════════════════════════════════════
ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif", "image/heic", "image/heif"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic", ".heif"}

def _upload_type_ok(file: UploadFile, allowed_types: set, allowed_extensions: set) -> bool:
    """True if the upload's content-type OR filename extension is on the allow-list.
    Browsers (Windows Chrome/Edge especially) frequently can't identify HEIC/HEIF —
    e.g. photos extracted from a zip lose the extended-attribute MIME hint — and report
    a generic "application/octet-stream" instead. Falling back to the extension avoids
    rejecting legitimate photos just because the browser guessed wrong; the bytes are
    still decoded (and validated) for real by Pillow at compression time below."""
    if file.content_type in allowed_types:
        return True
    ext = os.path.splitext(file.filename or "")[1].lower()
    return ext in allowed_extensions

_EXT_CONTENT_TYPES = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp",
    ".gif": "image/gif", ".heic": "image/heic", ".heif": "image/heif", ".pdf": "application/pdf",
}

def _resolved_content_type(file: UploadFile) -> str:
    """The content-type to actually store. Falls back to a lookup by extension when the
    browser reported a generic/missing type — otherwise the wrong MIME ends up baked into
    the stored `data:` URI and browsers refuse to render or download it correctly."""
    if file.content_type and file.content_type != "application/octet-stream":
        return file.content_type
    ext = os.path.splitext(file.filename or "")[1].lower()
    return _EXT_CONTENT_TYPES.get(ext, file.content_type or "application/octet-stream")

# Photo/document compression decodes untrusted phone-camera files (HEIC in particular,
# via pillow-heif's native libheif binding) — a malformed or unusual file can segfault
# the decoder. asyncio.to_thread alone doesn't protect against that: a thread shares the
# same process memory, so a native crash there takes the entire server down with it (this
# is what was behind the intermittent 502s / PM2 restart-loop on photo uploads). Running
# this work in a real OS subprocess means a crash there only kills that one subprocess —
# concurrent.futures raises BrokenProcessPool in the awaiting request instead of the whole
# server dying, and _run_isolated below replaces the pool so the next upload gets a fresh one.
_compress_pool = ProcessPoolExecutor(max_workers=2)

async def _run_isolated(func, *args):
    global _compress_pool
    loop = asyncio.get_running_loop()
    try:
        return await loop.run_in_executor(_compress_pool, func, *args)
    except BrokenProcessPool:
        logger.error(f"{func.__name__} crashed its worker process — recreating the pool", exc_info=True)
        _compress_pool.shutdown(wait=False, cancel_futures=True)
        _compress_pool = ProcessPoolExecutor(max_workers=2)
        raise

RAW_UPLOAD_MAX = 25 * 1024 * 1024  # 25 MB cap on the original file straight off a phone camera
PHOTO_MAX_DIMENSION = 1600  # px, longest side
PHOTO_JPEG_QUALITY = 80
PHOTO_MAX_BYTES = 2 * 1024 * 1024  # hard cap — same 2MB ceiling as legal documents

def _compress_photo(content: bytes) -> tuple[bytes, str]:
    """Downscales and re-encodes an uploaded photo as JPEG. Raw phone-camera
    uploads were routinely 2-3MB+, which meant next/image on the storefront
    had to download the full original from Render on every cache miss just
    to produce a ~30KB resized thumbnail — this is why the storefront felt
    slow. Compressing once at upload time fixes it for every consumer. Quality
    is stepped down and, if that's still not enough, the image is shrunk
    further in a loop — same approach as _compress_document_image — so no
    photo leaves this function larger than PHOTO_MAX_BYTES."""
    img = Image.open(io.BytesIO(content))
    img = ImageOps.exif_transpose(img)  # respect phone camera orientation before resizing
    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")
    img.thumbnail((PHOTO_MAX_DIMENSION, PHOTO_MAX_DIMENSION), Image.LANCZOS)

    def encode(quality):
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=quality, optimize=True)
        return out.getvalue()

    quality = PHOTO_JPEG_QUALITY
    data = encode(quality)
    while len(data) > PHOTO_MAX_BYTES and quality > 30:
        quality -= 15
        data = encode(quality)
    while len(data) > PHOTO_MAX_BYTES and max(img.size) > 500:
        img.thumbnail((int(img.width * 0.8), int(img.height * 0.8)), Image.LANCZOS)
        data = encode(max(quality, 50))
    return data, "image/jpeg"

def _photo_out(p: dict) -> dict:
    return {"id": p["id"], "filename": p["filename"], "uploaded_at": p["uploaded_at"],
            "size": p["size"], "url": f"data:{p['content_type']};base64,{p['data']}"}

async def _compress_oversized_photos():
    """Same idea as _compress_oversized_documents (below), for vehicle_photos —
    kicked off once from the startup event, no button. Shrinks every photo
    currently over PHOTO_MAX_BYTES via _compress_photo, so anything uploaded
    before that cap (or before compression existed at all) gets caught too.
    Safe to run every startup — already-small photos are skipped."""
    photos = await db.vehicle_photos.find({}, {"_id": 0, "id": 1, "size": 1}).to_list(100000)
    oversized_ids = [p["id"] for p in photos if (p.get("size") or 0) > PHOTO_MAX_BYTES]
    if not oversized_ids:
        return
    logger.info(f"Compressing {len(oversized_ids)} oversized vehicle photo(s)...")
    succeeded = 0
    for photo_id in oversized_ids:
        p = await db.vehicle_photos.find_one({"id": photo_id})
        if not p:
            continue
        raw = base64.b64decode(p["data"])
        try:
            # Isolated subprocess, not a thread: this is the actual CPU-bound work (LANCZOS
            # resize + JPEG encode) and must not run inline (would freeze every other request
            # on this single-worker server), and a crash in it must not take the whole
            # process down with it — see _run_isolated.
            compressed, new_content_type = await _run_isolated(_compress_photo, raw)
        except Exception:
            logger.warning(f"Failed to compress vehicle photo {photo_id}", exc_info=True)
            continue
        if len(compressed) >= len(raw):
            continue  # didn't actually help — leave the original in place
        await db.vehicle_photos.update_one(
            {"id": photo_id},
            {"$set": {"data": base64.b64encode(compressed).decode("ascii"),
                      "content_type": new_content_type, "size": len(compressed)}},
        )
        succeeded += 1
    logger.info(f"Photo compression sweep done: {succeeded}/{len(oversized_ids)} shrunk.")

@api_router.get("/vehicles/{vid}/photos")
async def get_vehicle_photos(vid: str, cu: dict = Depends(require("vehicle_media", "view"))):
    v = await db.vehicles.find_one({"id": vid}, {"_id": 0, "id": 1})
    if not v: raise HTTPException(404, "Vehicle not found")
    photos = await db.vehicle_photos.find({"vehicle_id": vid}, {"_id": 0}).sort("uploaded_at", 1).to_list(200)
    return [_photo_out(p) for p in photos]

@api_router.post("/vehicles/{vid}/photos")
async def upload_vehicle_photo(vid: str, file: UploadFile = File(...), cu: dict = Depends(require("vehicle_media", "create"))):
    v = await db.vehicles.find_one({"id": vid})
    if not v: raise HTTPException(404, "Vehicle not found")
    if not _upload_type_ok(file, ALLOWED_IMAGE_TYPES, IMAGE_EXTENSIONS):
        raise HTTPException(400, f"File type {file.content_type} not allowed. Use JPEG/PNG/WebP/HEIC.")
    content = await file.read()
    if len(content) > RAW_UPLOAD_MAX:
        raise HTTPException(400, "File too large. Max 25MB.")
    content_type = _resolved_content_type(file)
    try:
        # Isolated subprocess — same reasoning as hash_pw_async re: not blocking the single
        # worker's event loop, plus crash containment: a native decoder crash on a bad phone
        # photo must not take the whole server down (see _run_isolated).
        content, content_type = await _run_isolated(_compress_photo, content)
    except Exception:
        logger.warning(f"Photo compression failed for upload to vehicle {vid}, storing original", exc_info=True)
    if len(content) > PHOTO_MAX_BYTES:
        # Only reachable if compression above failed and fell back to the raw upload —
        # the compressor itself guarantees output under PHOTO_MAX_BYTES on success.
        raise HTTPException(400, "File too large and could not be compressed. Try a smaller photo.")
    photo_id = str(uuid.uuid4())
    photo = {
        "id": photo_id, "vehicle_id": vid, "filename": file.filename or f"{photo_id}.jpg",
        "content_type": content_type, "data": base64.b64encode(content).decode("ascii"),
        "uploaded_at": datetime.now(timezone.utc).isoformat(), "size": len(content),
    }
    await db.vehicle_photos.insert_one(photo)
    asyncio.create_task(_notify_storefront())
    return _photo_out(photo)

@api_router.delete("/vehicles/{vid}/photos/{photo_id}")
async def delete_vehicle_photo(vid: str, photo_id: str, cu: dict = Depends(require("vehicle_media", "delete"))):
    r = await db.vehicle_photos.delete_one({"id": photo_id, "vehicle_id": vid})
    if r.deleted_count == 0: raise HTTPException(404, "Photo not found")
    asyncio.create_task(_notify_storefront())
    return {"message": "Photo deleted"}

# ══════════════════════════════════════════════════════════════════════
# ── LEGAL DOCUMENT UPLOAD ─────────────────────────────────────────────
# Also stored as base64 in MongoDB (db.legal_documents) for the same reason.
# ══════════════════════════════════════════════════════════════════════
ALLOWED_DOC_TYPES = {"application/pdf", "image/jpeg", "image/png", "image/webp", "image/heic", "image/heif"}
DOC_EXTENSIONS = IMAGE_EXTENSIONS | {".pdf"}
DOC_TYPES = ["bluebook", "insurance", "tax_clearance", "transfer", "other"]
DOC_MAX_DIMENSION = 2000  # px, longest side — higher than photos so scanned text/fine print stays legible
DOC_JPEG_QUALITY = 85
DOC_MAX_BYTES = 2 * 1024 * 1024  # hard cap — every stored document must land under this, however big the source

def _compress_document_image(content: bytes) -> tuple[bytes, str]:
    """Same starting point as _compress_photo but tuned for document scans
    (bluebook, insurance card, etc.) — a higher resolution/quality ceiling since
    these get zoomed in on to read fine print. On top of that, quality is
    stepped down and, if that alone isn't enough, the image is shrunk further
    in a loop until the encoded size is under DOC_MAX_BYTES — no document
    leaves this function larger than the 2MB cap, regardless of the source."""
    img = Image.open(io.BytesIO(content))
    img = ImageOps.exif_transpose(img)
    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")
    img.thumbnail((DOC_MAX_DIMENSION, DOC_MAX_DIMENSION), Image.LANCZOS)

    def encode(quality):
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=quality, optimize=True)
        return out.getvalue()

    quality = DOC_JPEG_QUALITY
    data = encode(quality)
    while len(data) > DOC_MAX_BYTES and quality > 30:
        quality -= 15
        data = encode(quality)
    while len(data) > DOC_MAX_BYTES and max(img.size) > 600:
        img.thumbnail((int(img.width * 0.8), int(img.height * 0.8)), Image.LANCZOS)
        data = encode(max(quality, 50))
    return data, "image/jpeg"

def _compress_pdf(content: bytes) -> bytes:
    """Recompresses every embedded raster image inside a PDF in place (JPEG
    re-encode via pypdf's ImageFile.replace) instead of rasterizing whole
    pages, so any vector text/annotations survive untouched. Runs a second,
    more aggressive pass if the first doesn't get under DOC_MAX_BYTES. Falls
    back to the original bytes if pypdf can't parse the file, there are no
    images to shrink (born-digital PDF), or compressing didn't actually help —
    a document is never made worse by running it through this."""
    def recompress(source: bytes, dimension: int, quality: int) -> bytes:
        writer = PdfWriter()
        writer.append(PdfReader(io.BytesIO(source)))
        for page in writer.pages:
            for pimg in page.images:
                try:
                    pil_img = pimg.image
                    if pil_img.mode not in ("RGB", "L"):
                        pil_img = pil_img.convert("RGB")
                    pil_img.thumbnail((dimension, dimension), Image.LANCZOS)
                    pimg.replace(pil_img, quality=quality)
                except Exception:
                    continue  # leave this one image as-is rather than risk corrupting the PDF
        out = io.BytesIO()
        writer.write(out)
        return out.getvalue()

    try:
        data = recompress(content, DOC_MAX_DIMENSION, DOC_JPEG_QUALITY)
        if len(data) > DOC_MAX_BYTES:
            harder = recompress(content, 1200, 50)
            if len(harder) < len(data):
                data = harder
        return data if len(data) < len(content) else content
    except Exception:
        logger.warning("PDF compression failed, storing original", exc_info=True)
        return content

def _doc_out(d: dict) -> dict:
    return {"id": d["id"], "filename": d["filename"], "doc_type": d["doc_type"],
            "original_name": d["original_name"], "uploaded_at": d["uploaded_at"], "size": d["size"],
            "url": f"data:{d['content_type']};base64,{d['data']}"}

@api_router.get("/vehicles/{vid}/legal-documents")
async def get_legal_documents(vid: str, cu: dict = Depends(require("vehicle_media", "view"))):
    v = await db.vehicles.find_one({"id": vid}, {"_id": 0, "id": 1})
    if not v: raise HTTPException(404, "Vehicle not found")
    docs = await db.legal_documents.find({"vehicle_id": vid}, {"_id": 0}).sort("uploaded_at", 1).to_list(200)
    return [_doc_out(d) for d in docs]

async def _compress_oversized_documents():
    """Background sweep, kicked off once from the startup event (see below) —
    no button, no admin endpoint. Recompresses every legal_document currently
    over DOC_MAX_BYTES (images via _compress_document_image, PDFs via
    _compress_pdf) so old uploads and anything from before this cap existed
    get shrunk down too, not just new ones (which are already capped at
    upload time). Awaits between documents so this never blocks the event
    loop for other requests (the actual compression work runs via asyncio.to_thread,
    not just an inter-item sleep(0)), and is safe to run every startup — anything
    already under the cap, or where compressing wouldn't actually shrink it,
    is skipped, so a repeat run does almost nothing."""
    docs = await db.legal_documents.find({}, {"_id": 0, "id": 1, "size": 1}).to_list(100000)
    oversized_ids = [d["id"] for d in docs if (d.get("size") or 0) > DOC_MAX_BYTES]
    if not oversized_ids:
        return
    logger.info(f"Compressing {len(oversized_ids)} oversized legal document(s)...")
    succeeded = 0
    for doc_id in oversized_ids:
        d = await db.legal_documents.find_one({"id": doc_id})
        if not d:
            continue
        raw = base64.b64decode(d["data"])
        content_type = d.get("content_type", "")
        try:
            if content_type.startswith("image/"):
                compressed, new_content_type = await _run_isolated(_compress_document_image, raw)
            elif content_type == "application/pdf":
                compressed, new_content_type = await _run_isolated(_compress_pdf, raw), content_type
            else:
                continue
        except Exception:
            logger.warning(f"Failed to compress legal document {doc_id}", exc_info=True)
            continue
        if len(compressed) >= len(raw):
            continue  # didn't actually help — leave the original in place
        await db.legal_documents.update_one(
            {"id": doc_id},
            {"$set": {"data": base64.b64encode(compressed).decode("ascii"),
                      "content_type": new_content_type, "size": len(compressed)}},
        )
        succeeded += 1
        await asyncio.sleep(0)  # yield to the event loop between documents
    logger.info(f"Document compression sweep done: {succeeded}/{len(oversized_ids)} shrunk.")

@api_router.post("/vehicles/{vid}/legal-documents")
async def upload_legal_document(vid: str, file: UploadFile = File(...), doc_type: str = Form("other"), cu: dict = Depends(require("vehicle_media", "create"))):
    v = await db.vehicles.find_one({"id": vid})
    if not v: raise HTTPException(404, "Vehicle not found")
    if not _upload_type_ok(file, ALLOWED_DOC_TYPES, DOC_EXTENSIONS):
        raise HTTPException(400, "Only PDF/JPEG/PNG/HEIC allowed for documents.")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 10MB.")
    content_type = _resolved_content_type(file)
    if content_type.startswith("image/"):
        try:
            content, content_type = await _run_isolated(_compress_document_image, content)
        except Exception:
            logger.warning(f"Document image compression failed for upload to vehicle {vid}, storing original", exc_info=True)
    elif content_type == "application/pdf":
        try:
            content = await _run_isolated(_compress_pdf, content)
        except Exception:
            logger.warning(f"PDF compression failed for upload to vehicle {vid}, storing original", exc_info=True)
    doc_id = str(uuid.uuid4())
    doc = {
        "id": doc_id, "vehicle_id": vid, "filename": file.filename or f"{doc_id}.pdf",
        "content_type": content_type, "data": base64.b64encode(content).decode("ascii"),
        "doc_type": doc_type, "original_name": file.filename,
        "uploaded_at": datetime.now(timezone.utc).isoformat(), "size": len(content),
    }
    await db.legal_documents.insert_one(doc)
    # Update status field
    status_field = f"{doc_type}_status" if doc_type != "other" else None
    if status_field: await db.vehicles.update_one({"id": vid}, {"$set": {status_field: "ok"}})
    return _doc_out(doc)

@api_router.delete("/vehicles/{vid}/legal-documents/{doc_id}")
async def delete_legal_document(vid: str, doc_id: str, cu: dict = Depends(require("vehicle_media", "delete"))):
    r = await db.legal_documents.delete_one({"id": doc_id, "vehicle_id": vid})
    if r.deleted_count == 0: raise HTTPException(404, "Document not found")
    return {"message": "Document deleted"}

@api_router.get("/admin/storage-usage")
async def storage_usage(cu: dict = Depends(admin_only)):
    """Aggregates storage used by uploaded vehicle photos and legal documents —
    the only two collections holding binary file data (base64 in Mongo, not disk).
    `bytes` here is the decoded file size recorded at upload time; the base64 text
    actually stored is ~33% larger than that on top."""
    photos = await db.vehicle_photos.find({}, {"_id": 0, "vehicle_id": 1, "size": 1}).to_list(100000)
    docs = await db.legal_documents.find({}, {"_id": 0, "vehicle_id": 1, "size": 1}).to_list(100000)
    photos_bytes = sum(p.get("size", 0) or 0 for p in photos)
    docs_bytes = sum(d.get("size", 0) or 0 for d in docs)

    by_vehicle = {}
    for item in photos + docs:
        vid = item.get("vehicle_id")
        by_vehicle[vid] = by_vehicle.get(vid, 0) + (item.get("size", 0) or 0)

    top = sorted(by_vehicle.items(), key=lambda kv: kv[1], reverse=True)[:10]
    top_vehicles = []
    for vid, size in top:
        v = await db.vehicles.find_one({"id": vid}, {"_id": 0, "brand": 1, "model": 1, "registration_number": 1})
        label = f"{v['brand']} {v['model']}" if v else "Unknown vehicle"
        if v and v.get("registration_number"): label += f" ({v['registration_number']})"
        top_vehicles.append({"vehicle_id": vid, "label": label, "bytes": size})

    return {
        "photos": {"count": len(photos), "bytes": photos_bytes},
        "documents": {"count": len(docs), "bytes": docs_bytes},
        "total_bytes": photos_bytes + docs_bytes,
        "top_vehicles": top_vehicles,
    }

# ══════════════════════════════════════════════════════════════════════
# ── SPARE PARTS ───────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════
class SetComponentIn(BaseModel):
    name: str
    stock: int = 0  # independently tracked, decremented when this specific item (not the whole set) is used/sold
    rate: float = 0  # this item's own price; sum across all components must not exceed the set's own unit_cost

class SparePartCreate(BaseModel):
    name: str
    category: str = "General"
    brand_compatibility: Optional[str] = None
    part_number: Optional[str] = None
    quantity: int = 0
    unit_cost: float = 0  # net cost, i.e. before vat_rate is applied
    vat_rate: Optional[float] = None  # e.g. 13 for 13% — applied server-side so unit_cost is always saved VAT-inclusive
    selling_price: Optional[float] = None
    vendor_id: Optional[str] = None
    bill_no: Optional[str] = None
    entry_date: Optional[str] = None
    supplier: Optional[str] = None  # kept for backward compat
    min_stock_alert: int = 2
    location: Optional[str] = None
    notes: Optional[str] = None
    is_kit: bool = False
    stock_type: str = "singular"  # "singular" | "set" | "kit" — purely descriptive label; is_kit still drives all BOM/breakdown logic below and is only true for "kit"
    # "Set" components are a lightweight named stock count (e.g. individual stickers in a sticker
    # set) — unlike kit components they aren't links to other spare_parts rows, so they're stored
    # inline here rather than through the kit_components collection/API. This still lets a single
    # item be used/sold out of the set without touching the set's own (whole-set) `quantity`.
    set_components: List[SetComponentIn] = []

class SparePartUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    brand_compatibility: Optional[str] = None
    part_number: Optional[str] = None
    quantity: Optional[int] = None
    unit_cost: Optional[float] = None
    selling_price: Optional[float] = None
    vendor_id: Optional[str] = None
    bill_no: Optional[str] = None
    entry_date: Optional[str] = None
    supplier: Optional[str] = None
    min_stock_alert: Optional[int] = None
    location: Optional[str] = None
    notes: Optional[str] = None
    is_kit: Optional[bool] = None
    stock_type: Optional[str] = None
    set_components: Optional[List[SetComponentIn]] = None

class KitComponentIn(BaseModel):
    component_part_id: str
    qty_per_kit: int

class KitComponentsUpdate(BaseModel):
    components: List[KitComponentIn]

class BreakKitRequest(BaseModel):
    quantity: int = 1

@api_router.get("/spare-parts")
async def get_spare_parts(category: Optional[str] = None, low_stock: Optional[bool] = None, cu: dict = Depends(require("spare_parts", "view"))):
    query = {}
    if category: query["category"] = category
    parts = await db.spare_parts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    if low_stock: parts = [p for p in parts if p.get("quantity", 0) <= p.get("min_stock_alert", 2)]
    # Batch-fetch vendor names
    vendor_ids = {p["vendor_id"] for p in parts if p.get("vendor_id")}
    vendor_map = {}
    if vendor_ids:
        vdocs = await db.vendors.find({"id": {"$in": list(vendor_ids)}}, {"_id": 0, "id": 1, "name": 1}).to_list(200)
        vendor_map = {v["id"]: v["name"] for v in vdocs}
    for p in parts:
        p["total_value"] = p.get("quantity", 0) * p.get("unit_cost", 0)
        p["low_stock"] = p.get("quantity", 0) <= p.get("min_stock_alert", 2)
        p["margin"] = round(((p.get("selling_price", 0) - p.get("unit_cost", 0)) / p.get("unit_cost", 1)) * 100, 1) if p.get("selling_price") and p.get("unit_cost") else None
        p["vendor_name"] = vendor_map.get(p.get("vendor_id", ""))
    return parts

def _check_set_component_rates(unit_cost: float, set_components: list):
    if not set_components: return
    total = round(sum(c.get("rate", 0) for c in set_components), 2)
    if total > unit_cost:
        raise HTTPException(400, f"Component rates add up to NPR {total}, which is more than the set's own rate (NPR {unit_cost})")

def _apply_vat(unit_cost: float, vat_rate: Optional[float]) -> float:
    if not vat_rate:
        return unit_cost
    return round(unit_cost * (1 + vat_rate / 100), 2)

@api_router.post("/spare-parts")
async def create_spare_part(part: SparePartCreate, cu: dict = Depends(require("spare_parts", "create"))):
    doc = {"id": str(uuid.uuid4()), **part.dict(), "created_at": datetime.now(timezone.utc).isoformat()}
    doc["unit_cost"] = _apply_vat(doc.get("unit_cost", 0), doc.pop("vat_rate", None))
    _check_set_component_rates(doc.get("unit_cost", 0), doc.get("set_components", []))
    if not doc.get("entry_date"): doc["entry_date"] = datetime.now(timezone.utc).date().isoformat()
    await db.spare_parts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/spare-parts/{pid}")
async def update_spare_part(pid: str, part: SparePartUpdate, cu: dict = Depends(require("spare_parts", "edit"))):
    upd = {k: v for k, v in part.dict().items() if v is not None}
    if not upd: raise HTTPException(400, "No fields to update")
    if "set_components" in upd or "unit_cost" in upd:
        existing = await db.spare_parts.find_one({"id": pid}, {"_id": 0, "unit_cost": 1, "set_components": 1})
        if not existing: raise HTTPException(404, "Not found")
        _check_set_component_rates(
            upd.get("unit_cost", existing.get("unit_cost", 0)),
            upd.get("set_components", existing.get("set_components", [])),
        )
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    r = await db.spare_parts.update_one({"id": pid}, {"$set": upd})
    if r.matched_count == 0: raise HTTPException(404, "Not found")
    doc = await db.spare_parts.find_one({"id": pid}, {"_id": 0})
    return doc

@api_router.delete("/spare-parts/{pid}")
async def delete_spare_part(pid: str, cu: dict = Depends(require("spare_parts", "delete"))):
    r = await db.spare_parts.delete_one({"id": pid})
    if r.deleted_count == 0: raise HTTPException(404, "Not found")
    await db.kit_components.delete_many({"$or": [{"kit_part_id": pid}, {"component_part_id": pid}]})
    return {"message": "Deleted"}

@api_router.post("/spare-parts/{pid}/adjust-stock")
async def adjust_spare_stock(pid: str, req: dict, cu: dict = Depends(require("spare_parts", "edit"))):
    delta = req.get("delta", 0)
    p = await db.spare_parts.find_one({"id": pid}, {"_id": 0, "quantity": 1})
    if not p: raise HTTPException(404, "Not found")
    new_qty = max(0, p.get("quantity", 0) + int(delta))
    await db.spare_parts.update_one({"id": pid}, {"$set": {"quantity": new_qty}})
    return {"quantity": new_qty}

@api_router.post("/spare-parts/{pid}/stock-out")
async def stock_out_part(pid: str, req: PartStockOut, cu: dict = Depends(require("spare_parts", "edit"))):
    p = await db.spare_parts.find_one({"id": pid}, {"_id": 0})
    if not p: raise HTTPException(404, "Not found")
    if req.quantity <= 0: raise HTTPException(400, "Quantity must be positive")
    # Atomic conditional decrement (see break_kit for why) instead of the old
    # find-then-set, so two simultaneous stock-outs can't both pass the check.
    updated = await db.spare_parts.find_one_and_update(
        {"id": pid, "quantity": {"$gte": req.quantity}},
        {"$inc": {"quantity": -req.quantity}},
        return_document=ReturnDocument.AFTER,
    )
    if not updated:
        raise HTTPException(400, f"Insufficient stock. Available: {p.get('quantity', 0)}")
    new_qty = updated["quantity"]
    txn = {
        "id": str(uuid.uuid4()), "part_id": pid, "part_name": p.get("name"),
        "type": "out", "quantity": req.quantity, "reason": req.reason,
        "date": req.date or datetime.now(timezone.utc).isoformat()[:10],
        "job_id": req.job_id, "notes": req.notes,
        "created_by": cu.get("username"), "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.part_transactions.insert_one(txn)
    txn.pop("_id", None)
    return {"quantity": new_qty, "transaction": txn}

class SetComponentStockOut(BaseModel):
    component_name: str
    quantity: int
    reason: str
    date: Optional[str] = None
    job_id: Optional[str] = None
    notes: Optional[str] = None

@api_router.post("/spare-parts/{pid}/set-components/stock-out")
async def stock_out_set_component(pid: str, req: SetComponentStockOut, cu: dict = Depends(require("spare_parts", "edit"))):
    """Deducts from a single named item inside a Set's component checklist, independently of
    the set's own (whole-set) `quantity` — lets one sticker/item be used without touching the
    count of complete sets on hand."""
    p = await db.spare_parts.find_one({"id": pid}, {"_id": 0})
    if not p: raise HTTPException(404, "Not found")
    if req.quantity <= 0: raise HTTPException(400, "Quantity must be positive")
    comp = next((c for c in p.get("set_components", []) if c.get("name") == req.component_name), None)
    if not comp: raise HTTPException(404, f"Component '{req.component_name}' not found in this set")
    updated = await db.spare_parts.find_one_and_update(
        {"id": pid, "set_components.name": req.component_name, "set_components.stock": {"$gte": req.quantity}},
        {"$inc": {"set_components.$.stock": -req.quantity}},
        return_document=ReturnDocument.AFTER,
    )
    if not updated:
        raise HTTPException(400, f"Insufficient stock. Available: {comp.get('stock', 0)}")
    updated_comp = next(c for c in updated["set_components"] if c["name"] == req.component_name)
    txn = {
        "id": str(uuid.uuid4()), "part_id": pid, "part_name": f"{p.get('name')} — {req.component_name}",
        "type": "out", "quantity": req.quantity, "reason": req.reason,
        "date": req.date or datetime.now(timezone.utc).date().isoformat(),
        "job_id": req.job_id, "notes": req.notes,
        "created_by": cu.get("username"), "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.part_transactions.insert_one(txn)
    txn.pop("_id", None)
    return {"component": updated_comp, "transaction": txn}

@api_router.get("/spare-parts/summary")
async def spare_parts_summary(cu: dict = Depends(require("spare_parts", "view"))):
    parts = await db.spare_parts.find({}, {"_id": 0}).to_list(1000)
    total_value = sum(p.get("quantity", 0) * p.get("unit_cost", 0) for p in parts)
    low_stock = [p for p in parts if p.get("quantity", 0) <= p.get("min_stock_alert", 2)]
    categories = list({p.get("category", "General") for p in parts})
    return {"total_parts": len(parts), "total_value": total_value, "low_stock_count": len(low_stock), "categories": categories}

@api_router.get("/spare-parts/{pid}/transactions")
async def get_part_transactions(pid: str, cu: dict = Depends(require("spare_parts", "view"))):
    txns = await db.part_transactions.find({"part_id": pid}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return txns

# ══════════════════════════════════════════════════════════════════════
# ── KITS / BILL OF MATERIALS ──────────────────────────────────────────
# A spare_parts doc with is_kit=True keeps its own `quantity` (sealed kits on
# hand). kit_components maps that kit to the component parts it breaks down
# into, each of which also keeps its own independent `quantity` (loose stock).
# ══════════════════════════════════════════════════════════════════════

@api_router.get("/spare-parts/{pid}/kit-components")
async def get_kit_components(pid: str, cu: dict = Depends(require("spare_parts", "view"))):
    rows = await db.kit_components.find({"kit_part_id": pid}, {"_id": 0}).to_list(200)
    if not rows: return []
    comp_ids = [r["component_part_id"] for r in rows]
    comps = await db.spare_parts.find({"id": {"$in": comp_ids}}, {"_id": 0}).to_list(200)
    comp_map = {c["id"]: c for c in comps}
    out = []
    for r in rows:
        c = comp_map.get(r["component_part_id"], {})
        out.append({**r, "component_name": c.get("name"), "component_quantity": c.get("quantity", 0)})
    return out

@api_router.post("/spare-parts/{pid}/kit-components")
async def set_kit_components(pid: str, body: KitComponentsUpdate, cu: dict = Depends(require("spare_parts", "edit"))):
    kit = await db.spare_parts.find_one({"id": pid}, {"_id": 0})
    if not kit: raise HTTPException(404, "Kit part not found")
    comp_ids = [c.component_part_id for c in body.components]
    if pid in comp_ids: raise HTTPException(400, "A kit cannot contain itself")
    if len(comp_ids) != len(set(comp_ids)): raise HTTPException(400, "Duplicate component in kit")
    if comp_ids:
        comps = await db.spare_parts.find({"id": {"$in": comp_ids}}, {"_id": 0}).to_list(200)
        comp_map = {c["id"]: c for c in comps}
        for c_id in comp_ids:
            c = comp_map.get(c_id)
            if not c: raise HTTPException(400, f"Component part {c_id} not found")
            if c.get("is_kit"): raise HTTPException(400, f"{c.get('name')} is itself a kit — nested kits aren't supported")
    await db.kit_components.delete_many({"kit_part_id": pid})
    if body.components:
        now = datetime.now(timezone.utc).isoformat()
        docs = [{"id": str(uuid.uuid4()), "kit_part_id": pid, "component_part_id": c.component_part_id,
                 "qty_per_kit": c.qty_per_kit, "created_by": cu.get("username"), "created_at": now}
                for c in body.components]
        await db.kit_components.insert_many(docs)
    await db.spare_parts.update_one({"id": pid}, {"$set": {"is_kit": bool(body.components)}})
    return await get_kit_components(pid, cu)

@api_router.get("/spare-parts/{pid}/containing-kits")
async def get_containing_kits(pid: str, cu: dict = Depends(require("spare_parts", "view"))):
    """Reverse lookup: which kits (if any) this part is a component of, with each
    kit's current sealed-stock — used to offer 'break a kit' when loose stock runs out."""
    rows = await db.kit_components.find({"component_part_id": pid}, {"_id": 0}).to_list(200)
    if not rows: return []
    kit_ids = list({r["kit_part_id"] for r in rows})
    kits = await db.spare_parts.find({"id": {"$in": kit_ids}}, {"_id": 0}).to_list(200)
    kit_map = {k["id"]: k for k in kits}
    out = []
    for r in rows:
        k = kit_map.get(r["kit_part_id"])
        if not k: continue
        out.append({"kit_part_id": k["id"], "kit_name": k.get("name"), "kit_quantity": k.get("quantity", 0), "qty_per_kit": r["qty_per_kit"]})
    return out

@api_router.post("/spare-parts/{pid}/break-kit")
async def break_kit(pid: str, req: BreakKitRequest, cu: dict = Depends(require("spare_parts", "edit"))):
    if req.quantity <= 0: raise HTTPException(400, "Quantity must be positive")
    kit = await db.spare_parts.find_one({"id": pid}, {"_id": 0})
    if not kit: raise HTTPException(404, "Not found")
    components = await db.kit_components.find({"kit_part_id": pid}, {"_id": 0}).to_list(200)
    if not components: raise HTTPException(400, "This part has no kit components defined")

    # Atomic conditional decrement — the $gte guard blocks the update entirely (rather
    # than clamping to 0) if concurrent breaks/sales already dropped kit stock below
    # what this request needs, closing the race window a plain find-then-update has.
    updated_kit = await db.spare_parts.find_one_and_update(
        {"id": pid, "quantity": {"$gte": req.quantity}},
        {"$inc": {"quantity": -req.quantity}},
        return_document=ReturnDocument.AFTER,
    )
    if not updated_kit:
        raise HTTPException(400, f"Insufficient kit stock. Available: {kit.get('quantity', 0)}")

    now_iso = datetime.now(timezone.utc).isoformat()
    txns = [{
        "id": str(uuid.uuid4()), "part_id": pid, "part_name": kit.get("name"),
        "type": "out", "quantity": req.quantity, "reason": "Kit Broken",
        "date": now_iso[:10], "job_id": None,
        "notes": f"Broken into {len(components)} component part(s)",
        "created_by": cu.get("username"), "created_at": now_iso,
    }]
    for comp in components:
        add_qty = comp["qty_per_kit"] * req.quantity
        comp_doc = await db.spare_parts.find_one_and_update(
            {"id": comp["component_part_id"]},
            {"$inc": {"quantity": add_qty}},
            return_document=ReturnDocument.AFTER,
        )
        if not comp_doc: continue
        txns.append({
            "id": str(uuid.uuid4()), "part_id": comp["component_part_id"], "part_name": comp_doc.get("name"),
            "type": "in", "quantity": add_qty, "reason": "From Broken Kit",
            "date": now_iso[:10], "job_id": None,
            "notes": f"From breaking {req.quantity}x {kit.get('name')}",
            "created_by": cu.get("username"), "created_at": now_iso,
        })
    await db.part_transactions.insert_many(txns)
    for t in txns: t.pop("_id", None)
    return {"kit_quantity": updated_kit["quantity"], "components_updated": len(components), "transactions": txns}

# ══════════════════════════════════════════════════════════════════════
# ── WEBSITE SYNC (hamroauto.com.np) ───────────────────────────────────
# ══════════════════════════════════════════════════════════════════════
@api_router.get("/sync/export")
async def export_for_website(cu: dict = Depends(admin_only)):
    """Export available inventory in hamroauto.com.np listing format."""
    vehicles = await db.vehicles.find({"status": "available"}, {"_id": 0}).to_list(200)
    vehicle_ids = [v["id"] for v in vehicles]
    photos_by_vehicle: dict = {}
    if vehicle_ids:
        all_photos = await db.vehicle_photos.find({"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0}).sort("uploaded_at", 1).to_list(5000)
        for p in all_photos:
            photos_by_vehicle.setdefault(p["vehicle_id"], []).append(p)
    listings = []
    for v in vehicles:
        photos = photos_by_vehicle.get(v["id"], [])
        listings.append({
            "title": f"{v.get('brand')} {v.get('model')} {v.get('year')}",
            "brand": v.get("brand"), "model": v.get("model"), "year": v.get("year"),
            "price": v.get("selling_price"), "engine_cc": v.get("engine_cc"),
            "fuel_type": v.get("fuel_type"), "ownership": v.get("ownership_number"),
            "color": v.get("color"), "condition": v.get("condition"),
            "km_run": v.get("kilometer_run"), "registration": v.get("registration_number"),
            "notes": v.get("notes"),
            "docs": {
                "bluebook": v.get("bluebook_status"), "insurance": v.get("insurance_status"),
                "tax": v.get("tax_clearance_status"), "transfer": v.get("transfer_status"),
            },
            "photos": [f"data:{p['content_type']};base64,{p['data']}" for p in photos],
            "contact": "Hamro G&G Auto · Kathmandu · 98XXXXXXXX",
            "source": "hamro_gng_auto",
            "exported_at": datetime.now(timezone.utc).isoformat(),
        })
    return {"count": len(listings), "listings": listings, "exported_at": datetime.now(timezone.utc).isoformat()}

@api_router.post("/sync/push")
async def push_to_website(cu: dict = Depends(admin_only)):
    """Simulate push to hamroauto.com.np — in production connect to their API."""
    vehicles = await db.vehicles.find({"status": "available"}, {"_id": 0}).to_list(200)
    sync_log = {"pushed_at": datetime.now(timezone.utc).isoformat(), "count": len(vehicles),
                "status": "exported", "message": f"Ready to sync {len(vehicles)} vehicles to hamroauto.com.np"}
    await db.sync_logs.insert_one({**sync_log, "id": str(uuid.uuid4())})
    sync_log.pop("_id", None)
    return sync_log

# ══════════════════════════════════════════════════════════════════════
# ── PUBLIC SHOP API (no auth — safe for an external storefront site) ──
# Read-only. Only ever returns the explicit allowlist below — never spread
# a raw vehicle dict here. Fields intentionally EXCLUDED as internal/
# sensitive: purchase_price, accessories_cost, minimum_selling_price,
# vendor_id, purchase_from, purchase_source, chassis_number, engine_number,
# customer_id, salesperson_id/name, discount, notes, created_by/at.
# ══════════════════════════════════════════════════════════════════════
def _public_vehicle_fields(v: dict) -> dict:
    return {
        "id": v.get("id"),
        "title": f"{v.get('brand', '')} {v.get('model', '')} {v.get('year', '')}".strip(),
        "brand": v.get("brand"),
        "model": v.get("model"),
        "variant": v.get("variant"),
        "type": v.get("vehicle_type", "bike"),
        "year": v.get("year"),
        "engine_cc": v.get("engine_cc"),
        "fuel_type": v.get("fuel_type"),
        "ownership_number": v.get("ownership_number"),
        "kilometer_run": v.get("kilometer_run"),
        "condition": v.get("condition"),
        "condition_rating": v.get("condition_rating"),
        "color": v.get("color"),
        "registration_number": v.get("registration_number"),
        "price": v.get("selling_price"),
        "status": "available",
        "created_at": v.get("created_at"),
    }

def _public_photo_url(request: Request, vid: str, photo_id: str) -> str:
    """Builds an absolute URL from the client-facing scheme/host, not request.base_url —
    that reflects whatever uvicorn itself saw, which is plain http behind an
    nginx/Caddy TLS-terminating reverse proxy (the typical VPS setup) unless
    X-Forwarded-Proto is both sent by the proxy and trusted by uvicorn's
    --forwarded-allow-ips. Get that wrong and every photo URL silently becomes
    http:// on an https:// site, which browsers and next/image's remotePatterns
    block outright. Reading the headers directly sidesteps needing that trust
    config to be right — safe here since this only affects a display URL, not a
    security decision."""
    scheme = request.headers.get("x-forwarded-proto", request.url.scheme).split(",")[0].strip()
    host = (request.headers.get("x-forwarded-host") or request.headers.get("host") or request.url.netloc).split(",")[0].strip()
    return f"{scheme}://{host}/api/public/vehicles/{vid}/photos/{photo_id}"

_default_company_id_cache: Optional[str] = None

async def _scope_to_default_company():
    """FastAPI dependency for the /public/* routes below. They're unauthenticated (no
    get_current_user, so current_company_id never gets set by a JWT) but the storefront
    (hamroauto.com.np) is one specific business's public site -- without this, every
    company that ever signs up via /auth/signup would have its inventory mixed into that
    one public storefront. Scopes to whichever company was created first (the original,
    pre-multi-tenant business) rather than building a whole public-storefront-per-company
    routing scheme, which nothing has asked for yet."""
    global _default_company_id_cache
    if _default_company_id_cache is None:
        c = await db.companies.find_one({}, {"_id": 0, "id": 1}, sort=[("created_at", 1)])
        _default_company_id_cache = c["id"] if c else None
    current_company_id.set(_default_company_id_cache)

@api_router.get("/public/vehicles", dependencies=[Depends(_scope_to_default_company)])
async def public_list_vehicles(request: Request):
    """Public, unauthenticated listing of available vehicles for an external shop frontend.
    Returns one cover photo URL per vehicle (the first uploaded) to keep the payload light —
    use /public/vehicles/{id} for the full photo gallery of a single vehicle."""
    vehicles = await db.vehicles.find({"status": "available"}, {"_id": 0}).sort("created_at", -1).to_list(200)
    vehicle_ids = [v["id"] for v in vehicles]
    # One batched fetch (id/vehicle_id/uploaded_at only — never the base64 `data`) instead of
    # a per-vehicle find_one, then keep the earliest photo per vehicle as its cover.
    all_photos = (
        await db.vehicle_photos.find(
            {"vehicle_id": {"$in": vehicle_ids}}, {"_id": 0, "id": 1, "vehicle_id": 1, "uploaded_at": 1}
        ).sort("uploaded_at", 1).to_list(20000)
        if vehicle_ids else []
    )
    cover_by_vehicle = {}
    for p in all_photos:
        cover_by_vehicle.setdefault(p["vehicle_id"], p)  # sorted ascending, so first seen = earliest
    out = []
    for v in vehicles:
        item = _public_vehicle_fields(v)
        cover = cover_by_vehicle.get(v["id"])
        item["cover_photo"] = _public_photo_url(request, v["id"], cover["id"]) if cover else None
        item["image_urls"] = [item["cover_photo"]] if cover else []
        out.append(item)
    return {"count": len(out), "vehicles": out}

@api_router.get("/public/vehicles/{vid}", dependencies=[Depends(_scope_to_default_company)])
async def public_get_vehicle(vid: str, request: Request):
    """Public, unauthenticated single-vehicle detail with the full photo gallery."""
    v = await db.vehicles.find_one({"id": vid, "status": "available"}, {"_id": 0})
    if not v: raise HTTPException(404, "Vehicle not found or not available")
    item = _public_vehicle_fields(v)
    photos = await db.vehicle_photos.find({"vehicle_id": vid}, {"_id": 0}).sort("uploaded_at", 1).to_list(50)
    item["image_urls"] = [_public_photo_url(request, vid, p["id"]) for p in photos]
    item["photos"] = item["image_urls"]  # kept for backwards compatibility with earlier consumers
    return item

_PHOTO_BYTES_CACHE: dict = {}  # photo_id -> (raw_bytes, content_type), see public_get_photo
_PHOTO_BYTES_CACHE_SIZE = 0  # running total of cached raw bytes — the cap below is a memory budget, not a photo count
PHOTO_CACHE_MAX_BYTES = 300 * 1024 * 1024  # ~300MB — 5000 uncapped photos could otherwise run into the GBs on a small VPS

@api_router.get("/public/vehicles/{vid}/photos/{photo_id}", dependencies=[Depends(_scope_to_default_company)])
async def public_get_photo(vid: str, photo_id: str):
    """Serves a single vehicle photo as a real image response (not JSON) — this gives
    hamroauto.com.np (or any consumer) a stable HTTPS URL per photo, so it can be added
    to an image-CDN allowlist (e.g. Next.js next/image remotePatterns) instead of having
    to handle inline base64 data URIs. Also backs the inventory list's thumbnails, where a
    single page can reference dozens of photos at once — each one now a separate request,
    so an in-process cache keeps that from becoming dozens of MySQL round-trips competing
    for the connection pool on every load. Safe because a photo_id's bytes never change
    after upload (no "replace photo" endpoint, only upload-new/delete); capped so a very
    large photo library can't grow this unbounded."""
    global _PHOTO_BYTES_CACHE_SIZE
    cached = _PHOTO_BYTES_CACHE.get(photo_id)
    if cached is None:
        p = await db.vehicle_photos.find_one({"id": photo_id, "vehicle_id": vid}, {"_id": 0})
        if not p: raise HTTPException(404, "Photo not found")
        raw = base64.b64decode(p["data"])
        cached = (raw, p["content_type"])
        if _PHOTO_BYTES_CACHE_SIZE + len(raw) <= PHOTO_CACHE_MAX_BYTES:
            _PHOTO_BYTES_CACHE[photo_id] = cached
            _PHOTO_BYTES_CACHE_SIZE += len(raw)
    content, content_type = cached
    return Response(
        content=content, media_type=content_type,
        headers={"Cache-Control": "public, max-age=31536000, immutable"},
    )

@api_router.get("/public/settings", dependencies=[Depends(_scope_to_default_company)])
async def public_get_settings():
    """Public, unauthenticated site branding/contact info for the storefront."""
    s = await db.settings.find_one({}, {"_id": 0, "id": 0})
    return s or {}

@api_router.post("/public/leads", dependencies=[Depends(_scope_to_default_company)])
async def public_create_lead(lead: LeadCreate):
    """Public, unauthenticated — Sell / Exchange / Book Service form submissions
    from the storefront. Reviewed and managed from the admin Leads screen."""
    l = lead.model_dump()
    l["id"] = str(uuid.uuid4())
    l["status"] = "new"
    l["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.leads.insert_one(l)
    l.pop("_id", None)
    return l

@app.on_event("shutdown")
async def shutdown():
    client.close()
    _compress_pool.shutdown(wait=False, cancel_futures=True)

app.include_router(api_router)